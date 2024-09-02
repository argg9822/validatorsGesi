import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import re
import shutil
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import pandas as pd
import tkinter.simpledialog as simpledialog
from colorama import init, Fore, Style
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import datetime
import json
import sys
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import re
import shutil
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import pandas as pd
import tkinter.simpledialog as simpledialog
from colorama import init, Fore, Style
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import datetime

import tkinter as tk
from tkinter import ttk
import re
import time


def validarVacias(sheet, CeldasVacias):
    celdas_pintadas_rojo = 0
    ultima_fila = sheet.max_row
    Num_celTexto = len(CeldasVacias["vacias"])
    columns = list(CeldasVacias["vacias"])
    for a in range(Num_celTexto):
        for i in range(2, ultima_fila + 1):
            if sheet.cell(row=i, column=columns[a]).value == " " :  
                    celdas_pintadas_rojo += 1
                    colum["column"] = {columns[a], 2}
                    colum["row"] = i
                    pintar(colum, sheet) 
    return  celdas_pintadas_rojo

def validarCeldasTexto(sheet, celTexto):
    celdas_pintadas_rojo = 0
    regex = re.compile("^[a-zA-ZÑñáéíóúÁÉÍÓÚ\s]+$")
    ultima_fila = sheet.max_row
    Num_celTexto = len(celTexto["ColumText"])
    columns = list(celTexto["ColumText"])
    for a in range(Num_celTexto):
        for i in range(2, ultima_fila + 1):
            if sheet.cell(row=i, column=columns[a]).value and not regex.match(sheet.cell(row=i, column=columns[a]).value):  
                    celdas_pintadas_rojo += 1
                    colum["column"] = {columns[a], 2}
                    colum["row"] = i
                    pintar(colum, sheet) 
    return  celdas_pintadas_rojo


# funcio para remplazar comillas
def remplazarComillas(sheet):
    try :
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and '`' in cell.value:
                    # Elimina las comillas
                    cell.value = cell.value.replace('`', '')

                    # Verifica si el valor es un número
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = numbers.FORMAT_NUMBER
                    # Verifica si el valor es una fecha
                    elif isinstance(cell.value, datetime.date):
                        cell.number_format = numbers.FORMAT_DATE_XLSX15
                    # Verifica si el valor es texto
                    else:
                        cell.number_format = numbers.FORMAT_TEXT      
    except Exception as e:
        print("Error", f"Se produjo un error de comillas: {str(e)}")      
         
#funcion para pintar celdas 
def pintar(colum, sheet):
    colorRed = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    number_colum = len(colum["column"])
    columns = list(colum["column"])
    for i in range(number_colum):
        sheet.cell(row=colum["row"], column=columns[i]).fill = colorRed
  
def saveFile():
    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        file_path_modificado = file_path.replace('.xlsx', '_errores.xlsx')
        
        # Guardar el libro de trabajo original con los cambios realizados
        #workbook.save(file_path)
        # Guardar el archivo modificado con el nombre específico para errores
        workbook.save(file_path_modificado)
        print("Archivo guardado", "El archivo ha sido guardado correctamente.")
        # Preguntar al usuario si desea abrir el archivo guardado
        open_file = messagebox.askquestion("Abrir Archivo", "¿Desea abrir el archivo guardado?")
        if open_file == 'yes':
            os.startfile(file_path_modificado)  # Abre el archivo guardado
    except Exception as e:
        print("Error", f"No se pudo guardar el archivo: {str(e)}")

def preguntaDescarga():
    try:
        respuesta = messagebox.askquestion("Abrir Archivo", "¿Guardar el archivo generado?")
        if respuesta == "yes":
            cadenaGuardar = "Guardando archivo..."
            print(cadenaGuardar)
            saveFile()
        else:
            print("Tu archivo no será descargado")
    except Exception as e:
        print(f"No se pudo guardar el archivo: {str(e)}")