import threading
import time
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Selenium e integraciones
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

import customtkinter as ctk
from tkinter import filedialog, messagebox

# --- Configuración de Estilo ---
COLORS = {
    "bg_dark":      "#0D1117",
    "bg_card":      "#161B22",
    "bg_input":     "#010409",
    "accent":       "#238636",
    "accent_hover": "#2EA043",
    "text_main":    "#E6EDF3",
    "text_dim":     "#8B949E",
    "border":       "#30363D",
    "blue_btn":     "#1F6FEB",
    "red_btn":      "#DA3633"
}

# CAMBIO 1: Heredar de CTkToplevel para ser ventana secundaria
class OPENUI(ctk.CTkToplevel):
    def __init__(self, master=None):
        super().__init__(master)

        self.title("GESI - Validador sesiones")
        self.geometry("1000x750")
        
        # CAMBIO 2: Usar self directamente para configurar (ya no existe self.root)
        self.configure(fg_color=COLORS["bg_dark"])
        
        # Mantener al frente al abrir
        self.after(100, self.lift)
        self.focus_force()
        
        self.nombres = None
        self.fichas = []
        self.driver = None
        self.captcha_listo = threading.Event()
        self.confirmacion_si = threading.Event()
        self.stop_event = threading.Event()
        self._tiempo_inicio = None
        self._tick_activo   = False
        self._eta_segundos  = None

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._build_sidebar()
        self._build_main_panel()

    def _build_sidebar(self):
        self.sidebar = ctk.CTkFrame(self, width=320, fg_color=COLORS["bg_card"], corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")

        ctk.CTkLabel(self.sidebar, text="Credenciales GESIForm",
                    font=ctk.CTkFont("Segoe UI", 18, "bold")).pack(pady=(30, 20))

        self.user_var = ctk.StringVar()
        self.pass_var = ctk.StringVar()

        ctk.CTkLabel(self.sidebar, text="Usuario", text_color=COLORS["text_dim"]).pack(anchor="w", padx=25)
        self.entry_user = ctk.CTkEntry(self.sidebar, textvariable=self.user_var, height=35, fg_color=COLORS["bg_input"])
        self.entry_user.pack(fill="x", padx=25, pady=(0, 10))

        ctk.CTkLabel(self.sidebar, text="Contraseña", text_color=COLORS["text_dim"]).pack(anchor="w", padx=25)
        self.entry_pass = ctk.CTkEntry(self.sidebar, textvariable=self.pass_var, show="*", height=35, fg_color=COLORS["bg_input"])
        self.entry_pass.pack(fill="x", padx=25, pady=(0, 20))
        
        # ── Captcha ───────────────────────────────────────────────────────────────
        self.btn_captcha = ctk.CTkButton(
            self.sidebar,
            text="Confirmar Captcha ✅",
            fg_color=COLORS["blue_btn"],
            state="disabled",
            command=lambda: self.captcha_listo.set(),
        )
        self.btn_captcha.pack(fill="x", padx=25, pady=15)

        ctk.CTkFrame(self.sidebar, height=2, fg_color=COLORS["border"]).pack(fill="x", padx=20, pady=10)

        # ── Entorno ───────────────────────────────────────────────────────────────
        ctk.CTkLabel(self.sidebar, text="ENTORNO",
                    font=ctk.CTkFont("Segoe UI", 12, "bold"),
                    text_color=COLORS["text_dim"]).pack(anchor="w", padx=25, pady=(10, 6))

        self.entorno_var = ctk.StringVar(value="")  
        self.entorno_xpath = {
            "Laboral": {
                "entorno": "/html/body/div[1]/div/nav/div/div[4]/ul/li[6]/a",
                "base":    "/html/body/div[1]/div/nav/div/div[4]/ul/li[6]/div/ul/form[6]/li/a",
                "seccion_1": "controlBotonSeccion322",
                "seccion_2": "controlBotonSeccion323"
            },
            "Educativo": {
                "entorno": "/html/body/div[1]/div/nav/div/div[4]/ul/li[5]/a",
                "base":    "/html/body/div[1]/div/nav/div/div[4]/ul/li[5]/div/ul/form[6]/li/a",
                "seccion_1": "controlBotonSeccion287",
                "seccion_2": "controlBotonSeccion288"
            },
            "Comunitario": {
                "entorno": "/html/body/div[1]/div/nav/div/div[4]/ul/li[4]/a",
                "base":    "/html/body/div[1]/div/nav/div/div[4]/ul/li[4]/div/ul/form[5]/li/a",
                "seccion_1": "controlBotonSeccion318",
                "seccion_2": "controlBotonSeccion319"
            },
            "Institucional": {
                "entorno": "/html/body/div/div/nav/div/div[4]/ul/li[3]/a",
                "base":    "/html/body/div/div/nav/div/div[4]/ul/li[3]/div/ul/form[10]/li/a",
                "seccion_1": "controlBotonSeccion314",
                "seccion_2": "controlBotonSeccion315",
            },
        }
        
        # vacío = ninguno seleccionado aún

        ENTORNOS = [
            ("🏭  Laboral",       "Laboral"),
            ("🎓  Educativo",     "Educativo"),
            ("🏘️  Comunitario",   "Comunitario"),
            ("🏥  Institucional", "Institucional"),
        ]

        self._entorno_btns = {}
        entorno_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        entorno_frame.pack(fill="x", padx=25, pady=(0, 6))

        for label, valor in ENTORNOS:
            btn = ctk.CTkButton(
                entorno_frame,
                text=label,
                height=34,
                fg_color=COLORS["bg_input"],
                text_color=COLORS["text_dim"],
                hover_color=COLORS["border"],
                command=lambda v=valor: self._seleccionar_entorno(v),
            )
            btn.pack(fill="x", pady=3)
            self._entorno_btns[valor] = btn

        ctk.CTkFrame(self.sidebar, height=2, fg_color=COLORS["border"]).pack(fill="x", padx=20, pady=10)

        # ── Opciones adicionales ──────────────────────────────────────────────────
        ctk.CTkLabel(self.sidebar, text="OPCIONES",
                    font=ctk.CTkFont("Segoe UI", 12, "bold"),
                    text_color=COLORS["text_dim"]).pack(anchor="w", padx=25, pady=(0, 6))

        self.solo_calidad_var    = ctk.BooleanVar(value=False)
        self.solo_comprobador_var = ctk.BooleanVar(value=False)

        self.chk_calidad = ctk.CTkCheckBox(
            self.sidebar,
            text="Solo calidad",
            variable=self.solo_calidad_var,
            fg_color=COLORS["blue_btn"],
            hover_color=COLORS["border"],
            command=self._on_opcion_calidad,
        )
        self.chk_calidad.pack(anchor="w", padx=25, pady=(0, 8))

        self.chk_comprobador = ctk.CTkCheckBox(
            self.sidebar,
            text="Solo comprobador",
            variable=self.solo_comprobador_var,
            fg_color=COLORS["blue_btn"],
            hover_color=COLORS["border"],
            command=self._on_opcion_comprobador,
        )
        self.chk_comprobador.pack(anchor="w", padx=25, pady=(0, 10))

    # ── Helpers de estado ─────────────────────────────────────────────────────────

    def _seleccionar_entorno(self, valor: str):
        """
        Resalta el botón del entorno elegido y apaga los demás.
        self.entorno_var queda con el valor seleccionado.
        Si se pulsa el mismo entorno que ya estaba activo, lo deselecciona.
        """
        if self.entorno_var.get() == valor:
            # Deseleccionar
            self.entorno_var.set("")
            self._entorno_btns[valor].configure(
                fg_color=COLORS["bg_input"],
                text_color=COLORS["text_dim"],
            )
            return

        # Apagar todos
        for v, btn in self._entorno_btns.items():
            btn.configure(fg_color=COLORS["bg_input"], text_color=COLORS["text_dim"])

        # Encender el seleccionado
        self.entorno_var.set(valor)
        self._entorno_btns[valor].configure(
            fg_color=COLORS["blue_btn"],
            text_color="#FFFFFF",
        )

    def _on_opcion_calidad(self):
        """Solo calidad y Solo comprobador son mutuamente excluyentes."""
        if self.solo_calidad_var.get():
            self.solo_comprobador_var.set(False)

    def _on_opcion_comprobador(self):
        """Solo comprobador y Solo calidad son mutuamente excluyentes."""
        if self.solo_comprobador_var.get():
            self.solo_calidad_var.set(False)

    def _build_main_panel(self):
        # CAMBIO 4: Anclar a 'self'
        self.main_view = ctk.CTkFrame(self, fg_color="transparent")
        self.main_view.grid(row=0, column=1, sticky="nsew", padx=25, pady=25)

        self.btn_file = ctk.CTkButton(self.main_view, text="📂 Seleccionar archivo de excel",
                                     fg_color=COLORS["bg_card"], height=45,
                                     command=self.seleccionar_archivo)
        self.btn_file.pack(fill="x", pady=(0, 8))

        # ── Barra de progreso ─────────────────────────────────────────────────
        self._lbl_progreso = ctk.CTkLabel(
            self.main_view, text="",
            text_color=COLORS["text_dim"],
            font=ctk.CTkFont("Segoe UI", 11),
            anchor="e",
        )
        self._lbl_progreso.pack(fill="x")

        self._progress_container = ctk.CTkFrame(
            self.main_view,
            fg_color=COLORS["bg_input"],
            height=28,
            corner_radius=6,
        )
        self._progress_container.pack(fill="x", pady=(2, 8))
        self._progress_container.pack_propagate(False)

        self.progress_bar = ctk.CTkProgressBar(
            self._progress_container,
            fg_color=COLORS["bg_input"],
            progress_color=COLORS["accent"],
            height=28,
            corner_radius=6,
        )
        self.progress_bar.set(0)
        self.progress_bar.pack(fill="both", expand=True)

        self._lbl_tiempo = ctk.CTkLabel(
            self._progress_container,
            text="",
            text_color="#FFFFFF",
            font=ctk.CTkFont("Consolas", 11),
            fg_color="transparent",
        )
        self._lbl_tiempo.place(relx=0.5, rely=0.5, anchor="center")

        self.txt_log = ctk.CTkTextbox(self.main_view, fg_color="#010409", border_width=1,
                                     font=ctk.CTkFont("Consolas", 12), text_color="#76e1fe")
        self.txt_log.pack(fill="both", expand=True, pady=10)
        self.txt_log.configure(state="disabled")

        btn_row = ctk.CTkFrame(self.main_view, fg_color="transparent")
        btn_row.pack(fill="x")

        self.btn_action = ctk.CTkButton(btn_row, text="Iniciar validación ▶️",
                                        fg_color=COLORS["accent"], height=55,
                                        font=ctk.CTkFont("Segoe UI", 16, "bold"),
                                        command=self.start_thread)
        self.btn_action.pack(side="left", fill="x", expand=True, padx=(0, 6))

        self.btn_stop = ctk.CTkButton(btn_row, text="⏹ Detener",
                                      fg_color=COLORS["red_btn"], height=55,
                                      font=ctk.CTkFont("Segoe UI", 16, "bold"),
                                      state="disabled",
                                      command=self.detener_validacion)
        self.btn_stop.pack(side="left", ipadx=16)

    def log(self, message):
        t = datetime.now().strftime("%H:%M:%S")
        self.txt_log.configure(state="normal")
        self.txt_log.insert("end", f"[{t}] {message}\n")
        self.txt_log.see("end")
        self.txt_log.configure(state="disabled")
        self.update_idletasks()

    # --- LÓGICA DE AUTOMATIZACIÓN (Sin cambios necesarios aquí) ---
    def seleccionar_archivo(self):
        archivo = filedialog.askopenfilename(title="Seleccionar archivo de Excel", filetypes=[("Archivos de Excel", "*.xlsx;*.xls")])
        if archivo:
            try:
                wb = load_workbook(archivo)
                self.nombres = wb['Hoja1']
                self.log(f"Archivo cargado: {Path(archivo).name}")
                self.btn_file.configure(text=f"✅ {Path(archivo).name}")
                self.fichas = [str(r[0]).strip() for r in self.nombres.iter_rows(min_col=1, max_col=1, values_only=True)
                  if r[0] is not None and str(r[0]).strip()]
                
                messagebox.showinfo("Carga exitosa", f"Se cargaron {len(self.fichas)} fichas")
            except Exception as e:
                self.log(f"Error cargando Excel: {e}")

    def wait_for_element(self, by, value, timeout=10):
        return WebDriverWait(self.driver, timeout).until(EC.presence_of_element_located((by, value)))

    def start_thread(self):
        if not self.nombres:
            return messagebox.showwarning("Falta Info", "Carga el Excel primero.")
        self.stop_event.clear()
        self._eta_segundos  = None
        self._tiempo_inicio = datetime.now()
        self._tick_activo   = True
        self._lbl_tiempo.configure(text="⏱  00:00    ETA: --:--")
        self._lbl_progreso.configure(text="")
        self.progress_bar.set(0)
        self.btn_action.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self._tick()
        threading.Thread(target=self.ingresoAplicativo, daemon=True).start()

    def detener_validacion(self):
        self.log("⏹ Deteniendo... se completará la ficha en curso y luego se generará el reporte.")
        self.stop_event.set()
        self.btn_stop.configure(state="disabled")

    def _reset_botones(self):
        self._tick_activo = False
        self.btn_action.configure(state="normal")
        self.btn_stop.configure(state="disabled")

    # ── Cronómetro ────────────────────────────────────────────────────────────

    @staticmethod
    def _fmt_tiempo(seg: int) -> str:
        h = seg // 3600
        m = (seg % 3600) // 60
        s = seg % 60
        return f"{h:02d}:{m:02d}:{s:02d}" if h else f"{m:02d}:{s:02d}"

    def _tick(self):
        if not self._tick_activo or not self._tiempo_inicio:
            return
        elapsed = int((datetime.now() - self._tiempo_inicio).total_seconds())
        eta_part = (f"    ETA: ~{self._fmt_tiempo(self._eta_segundos)}"
                    if self._eta_segundos is not None else "    ETA: calculando...")
        self._lbl_tiempo.configure(
            text=f"⏱  {self._fmt_tiempo(elapsed)}{eta_part}"
        )
        self.after(1000, self._tick)

    def actualizar_progreso(self, actual: int, total: int):
        pct = actual / total if total else 0
        texto = f"Fichas validadas: {actual} / {total}  ({round(pct * 100)}%)"
        if actual > 0 and self._tiempo_inicio:
            elapsed = (datetime.now() - self._tiempo_inicio).total_seconds()
            self._eta_segundos = int(elapsed / actual * (total - actual))
        self.after(0, lambda: (
            self.progress_bar.set(pct),
            self._lbl_progreso.configure(text=texto),
        ))

    def ingresoAplicativo(self):
        try:
            self.log("Abriendo Chrome...")
            options = webdriver.ChromeOptions()
            options.add_argument('--start-maximized')
            self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

            self.driver.get("https://gesiapps.saludcapital.gov.co/GESI_sistemas/login")
            time.sleep(2)

            try:
                self.wait_for_element(By.XPATH, '/html/body/div/div[2]/button[3]').click()
                time.sleep(1)
                self.wait_for_element(By.XPATH, '/html/body/div/div[3]/p[2]/a').click()
                time.sleep(3)
            except:
                self.log("Omitiendo bypass de seguridad inicial.")

            self.log("Enviando credenciales de usuario...")
            self.wait_for_element(By.ID, 'usuario').send_keys(self.user_var.get())
            self.wait_for_element(By.ID, 'password').send_keys(self.pass_var.get())

            codigo = self.wait_for_element(By.ID, 'tokenAcceso').get_attribute('value')
            self.wait_for_element(By.ID, 'valorCodigoSeguridad').send_keys(codigo)

            self.log("⚠️ POR FAVOR, RESUELVE EL CAPTCHA Y PRESIONA 'Confirmar Captcha'.")
            self.btn_captcha.configure(state="normal")

            self.captcha_listo.wait()
            self.captcha_listo.clear()
            self.btn_captcha.configure(state="disabled")

            self.next_step()

        except Exception as e:
            self.log(f"Error: {e}")
            self.after(0, self._reset_botones)

    def next_step(self):
        self.wait_for_element(By.XPATH, '/html/body/section/div/div/form/div/div/div[7]/div/div/button').click()

        try:
            from liveValidator.main import ejecutarValidacion
            ejecutarValidacion(self)
        except Exception as e:
            messagebox.showerror("Error", f"Fallo al abrir ventana: {e}")
        finally:
            self.after(0, self._reset_botones)
    
# --- Función de arranque ---
def main(master=None):
    if master is None:
        root = ctk.CTk()
        app = OPENUI(root)
        root.mainloop()
    else:
        app = OPENUI(master)
        return app

if __name__ == "__main__":
    main()