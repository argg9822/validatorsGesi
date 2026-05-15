# main.py

import time
from datetime import datetime, date
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import tkinter as tk
from tkinter import filedialog

from liveValidator.secciones.seccion1 import validarSeccion1
from liveValidator.secciones.seccion2 import validarSeccion2
from liveValidator.helpers.validarDocumentos import validarDocumentos

# ── Estado global ─────────────────────────────────────────────────────────────
selfLocal = None
driver    = None
reporte   = []
reporte_documentos = []

PAGES_PER_ROW = 20


# ── Helpers de lectura ────────────────────────────────────────────────────────

def leerTextoElemento(element_id: str) -> str:
    if not element_id:
        return ""
    try:
        el = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, element_id))
        )
        if el.tag_name.lower() == "select":
            return Select(el).first_selected_option.text.strip()
        return (el.get_attribute("value") or "").strip()
    except:
        return ""

def leerValueElemento(element_id: str) -> str:
    if not element_id:
        return ""
    try:
        el = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, element_id))
        )
        if el.tag_name.lower() == "select":
            return Select(el).first_selected_option.get_attribute("value").strip()
        return (el.get_attribute("value") or "").strip()
    except:
        return ""

def leerValorOculto(element_id: str) -> str:
    try:
        el = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, element_id))
        )
        return (el.get_attribute("value") or "").strip()
    except:
        return ""

def leerValuesMultiple(element_id: str) -> list:
    if not element_id:
        return []
    try:
        el = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, element_id))
        )
        return [o.get_attribute("value") for o in Select(el).all_selected_options]
    except:
        return []

def leerTextosMultiple(element_id: str) -> list:
    if not element_id:
        return []
    try:
        el = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, element_id))
        )
        return [o.text.strip() for o in Select(el).all_selected_options]
    except:
        return []

def calcularEdad(fecha_nac: date, fecha_int: date) -> int:
    edad = fecha_int.year - fecha_nac.year
    if (fecha_int.month, fecha_int.day) < (fecha_nac.month, fecha_nac.day):
        edad -= 1
    return edad

def parsearFecha(valor: str):
    if not valor:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(valor.strip(), fmt).date()
        except ValueError:
            continue
    return None

def _acumularDocumentos(driver, ficha: str, page_num: int,
                         documentos_ficha: list):
    """
    En modo 'solo comprobador', lee únicamente los campos necesarios
    para la validación externa (tipo doc, num doc, nombre, apellido, fecha nac).
    No ejecuta ningún validador de calidad.
    """
    from liveValidator.helpers.camposPorLabel import resolverIDs, LABELS_SECCION_2
    from liveValidator.helpers.validarTipoDocumento import extraerCodigoDocumento

    CAMPOS_MINIMOS = {
        "tipo_doc":       LABELS_SECCION_2["tipo_doc"],
        "num_doc":        LABELS_SECCION_2["num_doc"],
        "primer_nombre":  LABELS_SECCION_2["primer_nombre"],
        "primer_apellido":LABELS_SECCION_2["primer_apellido"],
        "fecha_nac":      LABELS_SECCION_2["fecha_nac"],
    }

    ids = resolverIDs(driver, CAMPOS_MINIMOS)

    documentos_ficha.append({
        "ficha":           ficha,
        "pagina":          page_num,
        "tipo_doc":        extraerCodigoDocumento(leerTextoElemento(ids.get("tipo_doc", ""))),
        "num_doc":         leerTextoElemento(ids.get("num_doc", "")),
        "primer_nombre":   leerTextoElemento(ids.get("primer_nombre", "")),
        "primer_apellido": leerTextoElemento(ids.get("primer_apellido", "")),
        "fecha_nac":       leerTextoElemento(ids.get("fecha_nac", "")),
    })

# Dict que se pasa a cada sección para no duplicar funciones
HELPERS = {
    "leerTexto":          leerTextoElemento,
    "leerValue":          leerValueElemento,
    "leerOculto":         leerValorOculto,
    "leerMultipleValues": leerValuesMultiple,
    "leerMultipleTextos": leerTextosMultiple,
    "calcularEdad":       calcularEdad,
    "parsearFecha":       parsearFecha,
    "log":                lambda msg: selfLocal.log(msg),  # ← nuevo
}

# ── Registro de errores ───────────────────────────────────────────────────────

def registrarError(ficha, page_num, digitador, campo, valor, error,
                   seccion="", corregido=False, datos_extra=None):
    reg = {
        "ficha":     ficha,
        "pagina":    page_num,
        "digitador": digitador,
        "seccion":   seccion,
        "campo":     campo,
        "valor":     valor,
        "error":     error,
        "corregido": "Sí" if corregido else "No",
        "estado":    "Error",
    }
    if datos_extra:
        reg.update(datos_extra)
    reporte.append(reg)
    ico = "🔧" if corregido else "❌"
    selfLocal.log(f"   {ico} p.{page_num} [{seccion}] [{campo}]: {error}")

def _registrarErroresPagina(ficha: str, page_num: int, digitador: str,
                              errores: list, seccion: str):
    """
    Recibe la lista de errores de una sección y los registra en el reporte.
    Acepta tanto 3-tuplas (campo, valor, msg) como 4-tuplas (campo, valor, msg, corregido).
    Si no hay errores registra una fila OK para esa sección/página.
    """
    if errores:
        for item in errores:
            if len(item) == 4:
                campo, valor, msg, corregido = item
            else:
                campo, valor, msg = item
                corregido = False
            registrarError(ficha, page_num, digitador,
                           campo=campo, valor=valor, error=msg,
                           seccion=seccion, corregido=corregido)
        selfLocal.log(f"   ❌ {seccion} p.{page_num}: {len(errores)} error(es)")
    else:
        reporte.append({
            "ficha": ficha, "pagina": page_num, "digitador": digitador,
            "seccion": seccion, "campo": "", "valor": "",
            "error": "", "corregido": "No aplica", "estado": "OK",
        })
        selfLocal.log(f"   ✅ {seccion} p.{page_num}: sin errores")

# ── Paginador ─────────────────────────────────────────────────────────────────

def getPageButtonXpath(page_num: int) -> str:
    zero_based = page_num - 1
    tr_index   = (zero_based // PAGES_PER_ROW) + 1
    td_index   = (zero_based % PAGES_PER_ROW) * 2 + 1
    BASE = '//*[@id="main_body"]/div/div/main/div/div/div/div[2]/div[3]/div/center/table/tbody/tr/td[3]/table/tbody'
    return f'{BASE}/tr[{tr_index}]/td[{td_index}]/input'

def contarPaginas() -> int:
    total, tr_index = 0, 1
    BASE = '//*[@id="main_body"]/div/div/main/div/div/div/div[2]/div[3]/div/center/table/tbody/tr/td[3]/table/tbody'
    while True:
        encontro, td_index = False, 1
        while True:
            try:
                WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located(
                        (By.XPATH, f'{BASE}/tr[{tr_index}]/td[{td_index}]/input')
                    )
                )
                total += 1
                encontro = True
                td_index += 2
            except:
                break
        if not encontro:
            break
        tr_index += 1
    return max(total, 1)

def navegarAPagina(page_num: int) -> bool:
    try:
        btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, getPageButtonXpath(page_num)))
        )
        btn.click()
        time.sleep(0.7)
        return True
    except:
        return False
    
def _volverAlListado():
    try:
        WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH,
                '//*[@id="main_body"]/div/div/main/div/div/div/div[1]'
                '/div/div[2]/table/tbody/tr/td/form/button'))
        ).click()
    except:
        pass

def _haySeccion2(timeout: int = 3) -> bool:
    BASE = ('//*[@id="main_body"]/div/div/main/div/div/div/div[2]/div[3]'
            '/div/center/table/tbody/tr/td[3]/table/tbody')
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, f'{BASE}/tr[1]/td[1]/input'))
        )
        return True
    except:
        return False

# ── Flujo por ficha ───────────────────────────────────────────────────────────

def extraerUsuario() -> str:
    XPATH = "/html/body/div/div/main/div/div/div/div[2]/div[2]/div[1]/table/tbody/tr/td[9]"
    for _ in range(5):
        try:
            el = WebDriverWait(driver, 4).until(
                EC.presence_of_element_located((By.XPATH, XPATH))
            )
            texto = (el.text or "").strip()
            if texto:
                return texto
        except Exception:
            pass
        time.sleep(0.8)
    return ""

def procesarFicha(ficha: str, xpaths_entorno: dict):
    selfLocal.log(f"\n🗂️  Ficha {ficha}")

    solo_calidad      = selfLocal.solo_calidad_var.get()
    solo_comprobador  = selfLocal.solo_comprobador_var.get()

    # ── Buscar ficha ──────────────────────────────────────────────────────────
    el_filtro = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.ID, "valorFiltro"))
    )
    el_filtro.clear()
    el_filtro.send_keys(str(ficha))
    WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.ID, "BtnSearchRegs"))
    ).click()

    digitador = extraerUsuario()

    try:
        edit_btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, "btnEditReg"))
        )
    except:
        selfLocal.log(f"   ⚠️  Ficha {ficha} no encontrada.")
        reporte.append({
            "ficha": ficha, "pagina": "-", "digitador": digitador,
            "seccion": "-", "campo": "-", "valor": "-",
            "error": "Ficha no encontrada en el sistema",
            "corregido": "No aplica", "estado": "No encontrada",
        })
        return

    edit_btn.click()
    time.sleep(1)

    documentos_ficha = []

    # ── Sección 1 (solo si no es solo_comprobador) ────────────────────────────
    if not solo_comprobador:
        selfLocal.log("   📋 Validando sección 1...")
        errores_s1 = validarSeccion1(driver, ficha, 1, digitador, HELPERS)
        _registrarErroresPagina(ficha, 1, digitador, errores_s1, "Sección 1")

    # ── Entrar a sección 2 ────────────────────────────────────────────────────
    try:
        WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, xpaths_entorno["seccion_2"]))
        ).click()
        time.sleep(1)
    except Exception as e:
        selfLocal.log(f"   ❌ No se pudo entrar a sección 2: {e}")
        _volverAlListado()
        return

    # ── Verificar si hay sección 2 (paginación rápida) ───────────────────────
    if not _haySeccion2(timeout=3):
        espacio_fic = leerValueElemento("Espacio_fic")

        if espacio_fic == "04":
            selfLocal.log(f"   ℹ️  Ficha {ficha}: sin sección 2 (entorno comunitario — OK)")
            reporte.append({
                "ficha": ficha, "pagina": "-", "digitador": digitador,
                "seccion": "Sección 2", "campo": "", "valor": "",
                "error": "",
                "corregido": "No aplica",
                "estado": "OK — sin sección 2 (comunitario)",
            })
        else:
            selfLocal.log(f"   ❌ Ficha {ficha}: sin sección 2 (entorno {espacio_fic})")
            reporte.append({
                "ficha": ficha, "pagina": "-", "digitador": digitador,
                "seccion": "Sección 2",
                "campo": "Paginación",
                "valor": f"Espacio_fic={espacio_fic}",
                "error": f"La ficha no tiene sección 2 pero el entorno no es comunitario (Espacio_fic={espacio_fic})",
                "corregido": "No aplica",
                "estado": "Error",
            })

        _volverAlListado()
        return

    # ── Hay sección 2 — contar páginas y recorrerlas ──────────────────────────
    total_paginas = contarPaginas()
    selfLocal.log(f"   📄 Páginas en sección 2: {total_paginas}")

    for page_num in range(1, total_paginas + 1):
        if page_num > 1:
            selfLocal.log(f"   🔎 Navegando a página {page_num}...")
            if not navegarAPagina(page_num):
                selfLocal.log(f"   ❌ No se pudo navegar a página {page_num}.")
                continue

        if solo_comprobador:
            # Solo acumular documentos, sin validar campos
            _acumularDocumentos(driver, ficha, page_num, documentos_ficha)
        else:
            selfLocal.log(f"   📋 Validando sección 2, página {page_num}...")
            errores_s2 = validarSeccion2(driver, ficha, page_num, digitador,
                                         documentos_ficha, HELPERS)
            _registrarErroresPagina(ficha, page_num, digitador, errores_s2, "Sección 2")

    # ── Validación de documentos externos ─────────────────────────────────────
    if solo_calidad:
        selfLocal.log("   ℹ️  Modo 'Solo calidad': omitiendo validación de documentos externos.")
    elif documentos_ficha:
        modo = "'Solo comprobador'" if solo_comprobador else "normal"
        selfLocal.log(f"   🔍 Validando documentos externos (modo {modo})...")
        resultados_docs = validarDocumentos(driver, documentos_ficha, solo_comprobador=False)
        reporte_documentos.extend(resultados_docs)

    # ── Volver al listado ─────────────────────────────────────────────────────
    _volverAlListado()
    selfLocal.log(f"   🚪 Ficha {ficha} completada.")

# ── Navegación al entorno ─────────────────────────────────────────────────────

def enterEnvironmentAndFormat():
    entorno = selfLocal.entorno_var.get()

    if not entorno:
        selfLocal.log("❌ No se seleccionó ningún entorno")
        return None

    xpaths = selfLocal.entorno_xpath.get(entorno)
    selfLocal.log(f"🌐 Entrando al entorno: {entorno}")
    driver.find_element(By.XPATH, xpaths["entorno"]).click()
    time.sleep(1)
    driver.find_element(By.XPATH, xpaths["base"]).click()
    time.sleep(1)
    selfLocal.log(f"✅ Entorno '{entorno}' cargado")
    return xpaths


# ── Reporte Excel ─────────────────────────────────────────────────────────────

def _pedirRutaGuardado() -> str:
    """
    Abre un diálogo para que el usuario elija dónde guardar el reporte.
    Retorna la ruta completa con nombre de archivo, o una ruta por defecto
    en el escritorio si el usuario cancela.
    """
    nombre_sugerido = f"reporte_validacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)  # el diálogo aparece encima de todo

    ruta = filedialog.asksaveasfilename(
        title="Guardar reporte de validación",
        defaultextension=".xlsx",
        initialfile=nombre_sugerido,
        initialdir=os.path.expanduser("~/Desktop"),  # abre en el escritorio por defecto
        filetypes=[("Archivo Excel", "*.xlsx")],
    )

    root.destroy()

    if not ruta:
        # El usuario canceló → guardar en el escritorio con nombre automático
        ruta = os.path.join(os.path.expanduser("~/Desktop"), nombre_sugerido)
        selfLocal.log(f"   ℹ️  Diálogo cancelado. Se guardará en: {ruta}")

    return ruta


def exportarReporte():
    wb = Workbook()

    thin   = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    FILLS = {
        "OK":            PatternFill("solid", fgColor="E2EFDA"),
        "Error":         PatternFill("solid", fgColor="FDDCDC"),
        "Incompleto":    PatternFill("solid", fgColor="FFF2CC"),
        "No encontrada": PatternFill("solid", fgColor="F2F2F2"),
    }
    FILL_CORREGIDO = PatternFill("solid", fgColor="FFF2CC")

    headers    = ["Ficha", "Página", "Digitador", "Campo con error",
                  "Valor registrado", "Error detectado", "Corregido", "Estado"]
    col_widths = [16, 7, 18, 30, 35, 55, 10, 14]

    def _escribir_hoja(ws, filas):
        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill      = PatternFill("solid", fgColor="1F3864")
            cell.font      = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = center
            cell.border    = border
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        for row_idx, reg in enumerate(filas, 2):
            estado    = reg.get("estado", "OK")
            corregido = reg.get("corregido", "No")

            if estado == "Error" and corregido == "Sí":
                fill = FILL_CORREGIDO
            else:
                fill = FILLS.get(estado, FILLS["OK"])

            valores = [
                reg.get("ficha"),   reg.get("pagina"),  reg.get("digitador"),
                reg.get("campo"),   reg.get("valor"),   reg.get("error"),
                corregido,          estado,
            ]
            for col, val in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = center if col in (2, 7, 8) else left

    # ── Hoja resumen ──────────────────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "Resumen"
    _escribir_hoja(ws_all, reporte)

    # ── Una hoja por sección ──────────────────────────────────────────────────
    secciones = sorted({r.get("seccion", "") for r in reporte if r.get("seccion")})
    for sec in secciones:
        ws_sec = wb.create_sheet(sec)
        _escribir_hoja(ws_sec, [r for r in reporte if r.get("seccion") == sec])

    # ── Hoja validación documentos externos ───────────────────────────────────
    if reporte_documentos:
        ws2 = wb.create_sheet("Validación documentos")
        headers2    = ["Ficha", "Página", "Número doc", "Fuente",
                       "Consultado", "Coincide", "Detalle discrepancias"]
        col_widths2 = [16, 7, 16, 14, 11, 10, 65]

        for col, (h, w) in enumerate(zip(headers2, col_widths2), 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill      = PatternFill("solid", fgColor="1F3864")
            cell.font      = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = center
            cell.border    = border
            ws2.column_dimensions[cell.column_letter].width = w
        ws2.freeze_panes = "A2"

        for row_idx, res in enumerate(reporte_documentos, 2):
            coincide = res.get("coincide")
            fill2 = (PatternFill("solid", fgColor="E2EFDA") if coincide is True
                     else PatternFill("solid", fgColor="FDDCDC") if coincide is False
                     else PatternFill("solid", fgColor="FFF2CC"))
            valores2 = [
                res.get("ficha"),   res.get("pagina"),
                res.get("num_doc"), res.get("fuente"),
                "Sí" if res.get("consultado") else "No",
                "Sí" if coincide is True else ("No" if coincide is False else "N/A"),
                res.get("detalle", ""),
            ]
            for col, val in enumerate(valores2, 1):
                cell = ws2.cell(row=row_idx, column=col, value=val)
                cell.fill      = fill2
                cell.border    = border
                cell.alignment = center if col in (2, 5, 6) else left

    # ── Resumen numérico ──────────────────────────────────────────────────────
    total   = len(reporte)
    n_ok    = sum(1 for r in reporte if r["estado"] == "OK")
    n_error = sum(1 for r in reporte if r["estado"] == "Error")
    n_corr  = sum(1 for r in reporte if r.get("corregido") == "Sí")
    n_inc   = sum(1 for r in reporte if r["estado"] == "Incompleto")
    n_noenc = sum(1 for r in reporte if r["estado"] == "No encontrada")

    res_row = total + 3
    for i, (label, val, color) in enumerate([
        ("Total registros",         total,   None),
        ("Sin errores",             n_ok,    "E2EFDA"),
        ("Con error",               n_error, "FDDCDC"),
        ("Errores corregidos auto", n_corr,  "FFF2CC"),
        ("Datos incompletos",       n_inc,   "FFF2CC"),
        ("Fichas no encontradas",   n_noenc, "F2F2F2"),
    ]):
        c1 = ws_all.cell(row=res_row + i, column=1, value=label)
        c2 = ws_all.cell(row=res_row + i, column=2, value=val)
        c1.font = Font(bold=True)
        if color:
            c1.fill = PatternFill("solid", fgColor=color)
            c2.fill = PatternFill("solid", fgColor=color)

    # ── Guardar — preguntarle al usuario la ruta ──────────────────────────────
    ruta = _pedirRutaGuardado()

    try:
        wb.save(ruta)
        selfLocal.log(f"\n📋 Reporte guardado en: {ruta}")
        selfLocal.log(f"   ✅ OK:{n_ok} ❌ Error:{n_error} 🔧 Corregidos:{n_corr} "
                      f"⚠️ Inc:{n_inc} — No enc:{n_noenc}")
    except PermissionError:
        # El archivo puede estar abierto en Excel
        nombre_alt = _pedirRutaGuardado()
        try:
            wb.save(nombre_alt)
            selfLocal.log(f"\n📋 Reporte guardado en: {nombre_alt}")
        except Exception as e:
            selfLocal.log(f"\n❌ No se pudo guardar el reporte: {e}")
    except Exception as e:
        selfLocal.log(f"\n❌ Error al guardar el reporte: {e}")


# ── Punto de entrada ──────────────────────────────────────────────────────────

def ejecutarValidacion(ui_self):
    global selfLocal, driver, reporte, reporte_documentos
    selfLocal          = ui_self
    driver             = ui_self.driver
    reporte            = []
    reporte_documentos = []

    FICHAS = [f.strip() for f in selfLocal.fichas if f.strip()]
    if not FICHAS:
        selfLocal.log("❌ No hay fichas para procesar")
        return

    xpaths_entorno = enterEnvironmentAndFormat()
    if not xpaths_entorno:
        return

    total_fichas = len(FICHAS)
    selfLocal.log(f"📋 Total fichas a procesar: {total_fichas}")
    selfLocal.actualizar_progreso(0, total_fichas)
    validacion_iniciada = True

    try:
        for idx, ficha in enumerate(FICHAS, 1):
            if selfLocal.stop_event.is_set():
                selfLocal.log("⏹ Validación detenida por el usuario.")
                break
            try:
                procesarFicha(ficha, xpaths_entorno)
            except Exception as e:
                selfLocal.log(f"⚠️  Error inesperado en ficha {ficha}: {e}")
                reporte.append({
                    "ficha":     ficha,
                    "pagina":    "-",
                    "digitador": "-",
                    "campo":     "Error inesperado",
                    "valor":     "-",
                    "error":     str(e),
                    "estado":    "Error inesperado",
                    "edad":      "",
                    "fecha_nac": "",
                    "fecha_int": "",
                })
            selfLocal.actualizar_progreso(idx, total_fichas)

    except Exception as e:
        selfLocal.log(f"💥 Error crítico durante la validación: {e}")

    finally:
        if validacion_iniciada and (reporte or reporte_documentos):
            selfLocal.log("💾 Generando reporte con lo procesado hasta ahora...")
            exportarReporte()
        elif validacion_iniciada:
            selfLocal.log("⚠️  No hay datos que reportar (ninguna ficha fue procesada)")