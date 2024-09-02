# códigoInsert.py

# Código de ejemplo para crear un validador

import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import re
import shutil
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import pandas as pd
import tkinter.simpledialog as simpledialog
from colorama import init, Fore, Style
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import datetime
import json
import sys
import os

# Añadir la ruta al directorio que contiene 'funciones.py'
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

# Ahora puedes importar el módulo funciones
from funciones import validarVacias, preguntaDescarga

# Valor del entorno personalizado
entorno = {PLACEHOLDER}

try:
    with open('bases.json', 'r') as f:
        bases = json.load(f)
except FileNotFoundError:
    bases = {}

def mostrar_ventana_progreso(titulo, max_val):
    ventana = tk.Tk()
    ventana.title(titulo)
    label = tk.Label(ventana, text=titulo)
    label.pack(pady=10)
    progress = tk.Progressbar(ventana, orient="horizontal", length=300, mode="determinate")
    progress.pack(pady=20)
    return ventana, progress, label

def actualizar_barra_progreso(ventana, progress, valor):
    progress['value'] = valor
    ventana.update_idletasks()

# Validadores educativos
init()

colum = {"column": set(), "row": 0}
celTexto = {"ColumText": set()}
Genero = {"Genero": set()}
etniaVal = {"etniaVal": set()}
afiliacion = {"afiliacion": set()}
CeldasVacias = {"vacias": set()}
CeldasVacias_Condicional = {"vacias": set(), "row": 0}
placas = {"placas": set()}
Tel = {"Tel": set()}
Manzana = {"Manzana": set()}
rural = {"rural": set()}

def setBase(base):
    loadExcel()
    chooseBase(base)
    preguntaDescarga()

def loadExcel():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return
    global workbook
    workbook = openpyxl.load_workbook(file_path)
    global sheet
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    return headers, sheet

def chooseBase(base):
    # Obtenemos la función utilizando getattr
    funcion = globals().get(base)
    if callable(funcion):
        funcion()
    else:
        print(f"No existe una función llamada '{base}'.")  
#///////// espacio para insertar funciones de llamado por pagina/////////////
    
    

def SesionesCoelctivas():
    # Páginas del archivo Excel cargado
    num_paginas = len(workbook.sheetnames)
    print(f"El archivo Excel tiene {num_paginas} páginas.")

    # Primero, validar la página 1
    if num_paginas >= 1 and workbook.sheetnames[0] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]  # Acceder a la página 1
        print("Validando la página 1...")
        # validar_pagina1_sesiones(sheet)
        
        
#///////// espacio para insertar funciones de llamado por pagina/////////////

# Aquí puedes agregar más funciones para validadores según sea necesario.
