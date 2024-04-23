import datetime
from datetime import datetime
import math
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook, Workbook
import pandas as pd
from tkinter import filedialog, messagebox
import os
import re
from functions import pandas, openpyxl


pandas
##------------------------------------------------------------------------------------    
##---------------------CARGUE Y LECTURA DEL ARCHIVO EXCEL-----------------------------
##------------------------------------------------------------------------------------
def setBase(base):
    print(f"Validar >>>{base.upper()}<<<")
    loadFilesFromFolder()
    chooseBase(base)
    saveFile()
    
def loadFilesFromFolder():
    print("Cargando y consolidando archivos...")
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        merge_files(folder_selected)

def merge_files(folder_path):    
    # Crea un nuevo libro de Excel
    wb_combined = Workbook() 
    # Itera sobre todos los archivos en la carpeta seleccionada
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".csv"):
            sheet_name = os.path.splitext(file_name)[0]
            ws = wb_combined.create_sheet(title=sheet_name)
            df = pd.read_csv(os.path.join(folder_path, file_name), header=1, encoding='latin-1', delimiter=";")
    
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
                
    # Eliminar la hoja predeterminada creada al inicio
    wb_combined.remove(wb_combined['Sheet'])
    
    # Guarda el nuevo libro combinado
    global fileRoute
    fileRoute = folder_path+'/consolidado.xlsx'
    wb_combined.save(fileRoute)

    # Verifica si el archivo se guardó correctamente
    if os.path.isfile(fileRoute):
        loadExcel()
    else:
        print(f"ERROR: No se pudo guardar el archivo '{fileRoute}'.")
                
def loadExcel():
    global totalErrores
    totalErrores = 0
    if fileRoute:
        global workbook
        workbook = load_workbook(fileRoute)
    else:
        print("El archivo no se cargó")
        
##------------------------------------------------------------------------------------    
##---------------------------SWICTH PARA LAS BASES------------------------------------
##------------------------------------------------------------------------------------
def chooseBase(base):
    switch = {
        "sesiones_colectivas": sc,
        "hcb": hcb,
        "mascota_verde": mv,
        "persona_mayor": pm,
        "pci": pci,
        "ead": ead
    }
    execute_validator = switch.get(base)
    execute_validator()
    
##------------------------------------------------------------------------------------    
##--------------------------FUNCIONES PARA CADA BASE----------------------------------
##------------------------------------------------------------------------------------
def sc():
    for index, sheet_name in enumerate(workbook.sheetnames):
        global sheet
        sheet = workbook[sheet_name]
        data = sheet.values
        cols = next(data) # Obtener los encabezados de las columnas (ignorar la primera columna)
        global df
        df = pd.DataFrame(data, columns=cols)
        global df_modified
        df_modified = list_pages()
        print(f"-------------------Página {index+1}-------------------")
        if index == 0:
            totalErroresPg_1 = sc_pg1()
        
        if index == 1:
            totalErroresPg_2 = sc_pg2()
            
        if index == 2:
            totalErroresPg_3 = sc_pg3()
            print("----------------------------------------------")
            
    cantErrSc = totalErroresPg_1+totalErroresPg_2+totalErroresPg_3
    print(f">>TOTAL ERRORES EN SESIONES COLECTIVAS: {(cantErrSc)}") 
    
def hcb():
    for index, sheet_name in enumerate(workbook.sheetnames):
        global sheet
        sheet = workbook[sheet_name]
        data = sheet.values
        cols = next(data) # Obtener los encabezados de las columnas (ignorar la primera columna)
        global df
        df = pd.DataFrame(data, columns=cols)
        global df_modified
        df_modified = list_pages()
        print(f"-------------------Página {index+1}-------------------")
        if index == 0:
            totalErroresPg_1 = hcb_pg1()
            
        if index == 1:
            totalErroresPg_2 = hcb_pg2()
            print("----------------------------------------------")
            
    cantErrHcb = totalErroresPg_1+totalErroresPg_2
    print(f">>TOTAL ERRORES EN HCB: {(cantErrHcb)}")
    
def mv():
    for index, sheet_name in enumerate(workbook.sheetnames):
        global sheet
        sheet = workbook[sheet_name]
        data = sheet.values
        cols = next(data) # Obtener los encabezados de las columnas (ignorar la primera columna)
        global df
        df = pd.DataFrame(data, columns=cols)
        global df_modified
        df_modified = list_pages()
        print(f"-------------------Página {index+1}-------------------")
        if index == 0:
            totalErroresPg_2 = mv_pg1()
        if index == 1:
            totalErroresPg_2 = mv_pg2()
            print("----------------------------------------------")
            
    cantErrMv = totalErroresPg_2
    print(f">>TOTAL ERRORES EN MASCOTA VERDE: {(cantErrMv)}")

def pm():
    for index, sheet_name in enumerate(workbook.sheetnames):
        global sheet
        sheet = workbook[sheet_name]
        data = sheet.values
        cols = next(data) # Obtener los encabezados de las columnas (ignorar la primera columna)
        global df
        df = pd.DataFrame(data, columns=cols)
        global df_modified
        df_modified = list_pages()
        print(f"-------------------Página {index+1}-------------------")
        if index == 0:
            totalErroresPg_1 = pm_pg1()
        if index == 1:
            totalErroresPg_2 = pm_pg2()
            print("----------------------------------------------")
            
    cantErrPm = (totalErroresPg_1+totalErroresPg_2)
    print(f">>TOTAL ERRORES EN PERSONA MAYOR: {(cantErrPm)}")

def pci():
    for index, sheet_name in enumerate(workbook.sheetnames):
        global sheet
        sheet = workbook[sheet_name]
        data = sheet.values
        cols = next(data) # Obtener los encabezados de las columnas (ignorar la primera columna)
        global df
        df = pd.DataFrame(data, columns=cols)
        global df_modified
        df_modified = list_pages()
        print(f"-------------------Página {index+1}-------------------")
        if index == 0:
            totalErroresPg_1 = pci_pg1()
        if index == 1:
            totalErroresPg_2 = pci_pg2()
        if index == 2:
            totalErroresPg_3 = pci_pg3()
            print("----------------------------------------------")
            
    cantErrPm = (totalErroresPg_1+totalErroresPg_2+totalErroresPg_3)
    print(f">>TOTAL ERRORES EN PCI: {(cantErrPm)}")

def ead():
    for index, sheet_name in enumerate(workbook.sheetnames):
        global sheet
        sheet = workbook[sheet_name]
        data = sheet.values
        cols = next(data) # Obtener los encabezados de las columnas (ignorar la primera columna)
        global df
        df = pd.DataFrame(data, columns=cols)
        global df_modified
        df_modified = list_pages()
        print(f"-------------------Página {index+1}-------------------")
        if index == 0:
            totalErroresPg_1 = ead_pg1()
            print("----------------------------------------------")
            
    cantErrPm = (totalErroresPg_1)
    print(f">>TOTAL ERRORES EN ESCALA ABREVIADA: {(cantErrPm)}")
##------------------------------------------------------------------------------------    
##------------------------------GENERAL FUNCTIONS-------------------------------------
##------------------------------------------------------------------------------------

bgError = 'FFFF0000'
bgSecError = '005FFF'
def setBgError(index, columnName, color):    
    bgError = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell = sheet.cell(row=index+2, column=df_modified.columns.get_loc(columnName)+1)
    cell.fill = bgError
    
    bgErrorFicha = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    cellFicha = sheet.cell(row=index+2, column=df_modified.columns.get_loc('Ficha_fic')+1)
    cellFicha.fill = bgErrorFicha
    
    bgUsr = PatternFill(start_color='00ba4a', end_color='00ba4a', fill_type='solid')
    cellUsr = sheet.cell(row=index+2, column=df_modified.columns.get_loc('Red_fic')+1)
    cellUsr.fill = bgUsr
    cellUsr.font = Font(bold=True)

def getColumnNameByNumber(numberColumn):
    columnName = df_modified.iloc[:, numberColumn].name
    return columnName
    
def validar_telefono(*columnsNames):
    cantErroresTel = 0
    for columnName in columnsNames:
        if columnName in df_modified.columns:
            for index, fila in df_modified.iterrows():   
                cellTelefono = int(fila[columnName]) if pd.notna(fila[columnName]) else fila[columnName]
                if len(str(cellTelefono).strip()) not in [7, 10] and pd.notna(cellTelefono):
                    cantErroresTel += 1                
                    setBgError(index, columnName, bgError)
        else:
            print("No se encuentra la columna Teléfono")
        
    if cantErroresTel > 0:
        print(f"Teléfonos con longitud incorrecta: {cantErroresTel}")
    return cantErroresTel

def validate_email(*columnsNames):
    cantErrorsEmail = 0
    pattern = re.compile(r'^[\w\.-]+@[\w\.-]+\.\w+$')
    for columnName in columnsNames:
        if columnName not in df_modified.columns:
            raise ValueError(f'La columna {columnName} no se encuentra')
        
        fill_rows = df_modified[pd.notna(df_modified[columnName])]
        errors = fill_rows[~fill_rows[columnName].apply(lambda x: bool(pattern.search(str(x))))]
        
        cantErrorsEmail += len(errors)
        for index in errors.index:
            setBgError(index, columnName, bgError)
        
    if cantErrorsEmail > 0:
        print(f'Correos con formato incorrecto: {cantErrorsEmail}')
    return cantErrorsEmail

def validarNoManzana(columnNameManzana, columnNameNroManzana):
    totalErrApple = 0
    
    if columnNameManzana in df_modified.columns:
        for index, fila in df_modified.iterrows():
            cellManzana = fila[columnNameManzana]
            nroManzana = fila[columnNameNroManzana]
            if cellManzana == "SI" and pd.isna(nroManzana):                
                totalErrApple += 1
                setBgError(index, columnNameNroManzana,bgError)
    else:
        print("No se encuentra la columna manzana del cuidado")
        
    if totalErrApple > 0:
        print(f"Errores en manzana del cuidado: {totalErrApple}")
    return totalErrApple

def required_fields(arrayFields=[], type=1, cantNextColumn=1):
    totalEmptyFields = 0
    if type == 1:
        for field in arrayFields:
            if field in df_modified.columns:
                for j, fila in df_modified.iterrows():
                    cellField = fila[field]
                    if pd.isna(cellField):
                        totalEmptyFields += 1
                        setBgError(j, field, bgError)
            else:
                print(f'No se encontró la columna {field}')
    elif type == 2:
        for field in arrayFields:
            if field in df_modified.columns:
                column_index = df_modified.columns.get_loc(field) + cantNextColumn  # Obtener el índice de la columna actual y sumar 1
                if column_index < len(df_modified.columns):  # Verificar si el índice está dentro de los límites
                    next_field = df_modified.columns[column_index]  # Obtener el nombre de la siguiente columna
                    for j, fila in df_modified.iterrows():
                        cellField = fila[next_field]  # Acceder a la siguiente columna
                        if pd.isna(cellField):
                            totalEmptyFields += 1
                            setBgError(j, next_field, bgError) 
                else:
                    print(f'No hay siguiente columna después de {field}')
            else:
                print(f'No se encontró la columna {field}')
            
    if totalEmptyFields > 0:
        print(f"Campos obligatorios vacíos: {totalEmptyFields}")
    return totalEmptyFields

#Campos requeridos (validar página siguiente)
def required_fields_next_column(columnsNames = []):
    totalEmptyFields = 0
    
    for columnName in columnsNames:
        if columnName not in df_modified.columns:
            raise ValueError(f'La columna {columnName} no se encuentra')
        
        columnIndex = df_modified.columns.get_loc(columnName) + 1 #Indice de la columna principal más uno
        columnNameNext = df_modified.columns[columnIndex] if columnIndex < len(df_modified.columns) else None #Nombre de la siguiente columna
        
        errors = df_modified[pd.isna(df_modified[columnNameNext])]
        totalEmptyFields += len(errors)
        for index in errors.index:
            setBgError(index, columnNameNext, bgError)

    if totalEmptyFields > 0:
        print(f'Campos de alerta vacíos: {totalEmptyFields}')
    return totalEmptyFields

def difference_dates(date_interv, date2):   
    result = pd.to_datetime(date_interv) - pd.to_datetime(date2)
    return result

def comparar_fechas_vs_intervencion(columnNameDateInter, columnNameDate):
    totalErrDate = 0
    
    if columnNameDate not in df_modified.columns or columnNameDateInter not in df_modified.columns:
        raise ValueError(f'La(s) columna(s) {columnNameDate} y/o {columnNameDateInter} no se encontraron ')
    
    for index, fila in df_modified.iterrows():
        cellDateSesion = fila[columnNameDate]
        cellDateSesionInter = fila[columnNameDateInter]
        
        if pd.notna(cellDateSesion):
            if pd.to_datetime(cellDateSesion) < pd.to_datetime(cellDateSesionInter):
                setBgError(index, columnNameDate, bgError)
                totalErrDate += 1

    if totalErrDate > 0:
        print(f'Fechas menor que la de intervención: {totalErrDate}')
    return totalErrDate

def compare_dates(columnNameDateMin, columnNameDateMaj):
    totalErrDate = 0
    
    if columnNameDateMin not in df_modified.columns or columnNameDateMaj not in df_modified.columns:
        raise ValueError(f'La(s) columna(s) {columnNameDateMin} y/o {columnNameDateMaj} no se encontraron ')
    
    fill_rows = df_modified[pd.notna(df_modified[columnNameDateMin]) & pd.notna(df_modified[columnNameDateMaj])]
    
    for index, fila in fill_rows.iterrows():
        cellDateMaj = fila[columnNameDateMaj]
        cellDateMin = fila[columnNameDateMin]
        
        if pd.notna(cellDateMaj):
            if pd.to_datetime(cellDateMaj) < pd.to_datetime(cellDateMin):
                setBgError(index, columnNameDateMaj, bgError)
                totalErrDate += 1

    return totalErrDate

def fecha_mayor(columnNameDate):
    totalErrDate = 0
    fecha_actual = datetime.now().date()
    
    if columnNameDate not in df_modified.columns: 
        raise ValueError(f'No se encontró la columna {columnNameDate}')
        
    for index, fila in df_modified.iterrows():
        cellDate = fila[columnNameDate]
        if pd.notna(cellDate):
            if pd.to_datetime(cellDate).date() > fecha_actual:
                setBgError(index, columnNameDate, bgError)
                totalErrDate += 1
                
    if totalErrDate > 0:
        print(f'Fecha incoherente: {totalErrDate}')
    return totalErrDate

def calculate_age(birth_date, intervention_date):
    age = (pd.to_datetime(intervention_date) - pd.to_datetime(birth_date)).days // 365.25
    return math.ceil(age)

def type_institution(columnNameType, columnNameOther):
    totalErrTypeInst = 0
    if columnNameType not in df_modified.columns or columnNameOther not in df_modified.columns:
        raise ValueError(f"La columna {columnNameType} y/o {columnNameOther} no se encuentran")
    
    fill_rows = df_modified[pd.notna(df_modified[columnNameType]) & pd.notna(df_modified[columnNameOther])]
    
    pattern = re.compile(r'\ Otra\b', flags=re.IGNORECASE)# Expresión regular
    errors = fill_rows[(fill_rows[columnNameType].apply(lambda x: bool(pattern.search(str(x))))) & 
                       (fill_rows[columnNameOther].str.len() < 9) |
                       (fill_rows[columnNameOther].str.len() >= 9) & 
                        (~fill_rows[columnNameType].apply(lambda x: bool(pattern.search(str(x)))))]
    
    totalErrTypeInst = len(errors)
        
    for index in errors.index:
        setBgError(index, columnNameOther, bgSecError)
        setBgError(index, columnNameType, bgSecError)
        
    if totalErrTypeInst > 0:
        print(f'Errores en tipo de institución {totalErrTypeInst}')
        
    return totalErrTypeInst

def list_pages():
    for num in df['Ficha_fic'].unique():
        indices = df.index[df['Ficha_fic'] == num].tolist()
        
        for i, idx in enumerate(indices):
            cell = sheet.cell(row=idx+2, column=df.columns.get_loc('Red_fic')+1)
            cell.value = i + 1
    return df

#Validar que el campo contenga sólo números
def validate_only_number(columnsName=[]):
    totalErrNumber = 0
    pattern =re.compile(r'^[^0-9]+$')
    for columnName in columnsName:
        if columnName not in df_modified.columns:
            raise ValueError(f'La columna {columnName} no se encuentra')
        
        fill_rows = df_modified[pd.notna(df_modified[columnName])]
        errors = fill_rows[fill_rows[columnName].apply(lambda x: bool(pattern.search(str(x))))]
        
        totalErrNumber += len(errors)
        for index in errors.index:
            setBgError(index, columnName, bgError)
        
    if totalErrNumber > 0:
        print(f'Números con símbolos y/o caracteres: {totalErrNumber}')
    return totalErrNumber

#Validar que el campo contenga sólo texto
def validate_only_text(*columnsName):
    totalErrText = 0
    pattern =re.compile(r'^[^0-9.,:]+$')
    for columnName in columnsName:
        if columnName not in df_modified.columns:
            raise ValueError(f'La columna {columnName} no se encuentra')
        
        fill_rows = df_modified[pd.notna(df_modified[columnName])]
        errors = fill_rows[~fill_rows[columnName].apply(lambda x: bool(pattern.search(str(x))))]
        
        totalErrText += len(errors)
        for index in errors.index:
            setBgError(index, columnName, bgError)
        
    if totalErrText > 0:
        print(f'Texto mal escrito: {totalErrText}')
    return totalErrText

#Validar que el campo institución y barrio no contenga caracteres extraños
def validate_only_text_inst_barr(*columnsName):
    totalErrText = 0
    pattern =re.compile(r'^[^,:;|/()=$%&*-_]+$')
    
    for columnName in columnsName:
        if columnName not in df_modified.columns:
            raise ValueError(f'La columna {columnName} no se encuentra')
        
        fill_rows = df_modified[pd.notna(df_modified[columnName])]
        errors = fill_rows[fill_rows[columnName].apply(lambda x: bool(pattern.search(str(x))))]
        
        totalErrText += len(errors)
        for index in errors.index:
            setBgError(index, columnName, bgError)
        
    if totalErrText > 0:
        print(f'Texto mal escrito: {totalErrText}')
    return totalErrText

def validate_address(addressComponents):
    totalErrAddress = 0
    for component in addressComponents:
        if addressComponents[component] not in df_modified.columns:
            raise ValueError(f'La columna {addressComponents[component]} no se encuentra')
    
    fill_rows = df_modified[pd.notna(df_modified[addressComponents['columnNameZone']])]
    fill_rows[addressComponents['columnNameZone']] = fill_rows[addressComponents['columnNameZone']].astype(str)
    errors_urban = fill_rows[(fill_rows[addressComponents['columnNameZone']].str.startswith('1')) &
                            ((pd.isna(fill_rows[addressComponents['columnNameAx1']])) |
                            (pd.isna(fill_rows[addressComponents['columnNameNumber']])) |
                            (pd.isna(fill_rows[addressComponents['columnNameAx2']])) |
                            (pd.isna(fill_rows[addressComponents['columnNamePlate']])) |
                            (pd.notna(fill_rows[addressComponents['columnNameTrail']])) |
                            (pd.notna(fill_rows[addressComponents['columnNameX']])) |
                            (pd.notna(fill_rows[addressComponents['columnNameY']])))]
    
    errors_rural = fill_rows[(fill_rows[addressComponents['columnNameZone']].str.startswith('2')) &
                            ((pd.notna(fill_rows[addressComponents['columnNameAx1']])) |
                            (pd.notna(fill_rows[addressComponents['columnNameNumber']])) |
                            (pd.notna(fill_rows[addressComponents['columnNameAx2']])) |
                            (pd.notna(fill_rows[addressComponents['columnNamePlate']])) |
                            (pd.isna(fill_rows[addressComponents['columnNameTrail']])) |
                            (pd.isna(fill_rows[addressComponents['columnNameX']])) |
                            (pd.isna(fill_rows[addressComponents['columnNameY']])))]
    
    totalErrAddress = (len(errors_urban) + len(errors_rural))
    final_errors_urb = errors_urban[errors_urban.isna().any(axis=1)]
    
    for index in final_errors_urb.index:
        if final_errors_urb.loc[index].isna().any():
            setBgError(index, addressComponents['columnNameZone'], bgError)
    
    if totalErrAddress > 0:
        print(f'Errores en dirección: {totalErrAddress}')
    return totalErrAddress
##------------------------------------------------------------------------------------    
##-----------------------------POBLATIONAL FUNCTIONS----------------------------------
##------------------------------------------------------------------------------------

#Poblacional - longitud Nro documento
def len_num_doc(columNameTypeDoc, columnNameNumDoc):
    totalErrLenDoc = 0
    # Verificar la existencia de las columnas requeridas
    if columNameTypeDoc not in df_modified.columns or columnNameNumDoc not in df_modified.columns:
        raise ValueError(f'Las columnas {columNameTypeDoc} y/o {columnNameNumDoc} no se encuentran.')
    
    #Filtrar filas que no están vacías
    fill_rows = df_modified[pd.notna(df_modified[columNameTypeDoc]) & pd.notna(df_modified[columnNameNumDoc])]
    errors = fill_rows[(fill_rows[columNameTypeDoc].isin([61,60, '2- RC', '3- TI'])) & 
                            (fill_rows[columnNameNumDoc].astype(str).str.len() != 10)]
    
    totalErrLenDoc = len(errors)
    for index in errors.index:
        setBgError(index, columnNameNumDoc, bgError)
        
    if totalErrLenDoc > 0:
        print(f'Números de documento con longitud incorrecta: {totalErrLenDoc}')
    return totalErrLenDoc

#Poblacional edad y tipo de documento
def age_vs_typedoc(columnNameTypeDoc, columnNameAge):
    totalTypeDocErr = 0
    # Verificar la existencia de las columnas requeridas
    if columnNameTypeDoc not in df_modified.columns or columnNameAge not in df_modified.columns:
        raise ValueError(f'Las columnas {columnNameTypeDoc} y/o {columnNameAge} no se encuentran')
    
    #Filtrar filas que no están vacías
    fill_rows = df_modified[pd.notna(df_modified[columnNameTypeDoc]) & pd.notna(df_modified[columnNameAge])]
    errors = fill_rows[((fill_rows[columnNameAge] < 7) & 
                              (~fill_rows[columnNameTypeDoc].isin([60, 66, 64, 2482, 1638, 1640, 1639]))) |
                         ((fill_rows[columnNameAge] > 7) & (fill_rows[columnNameAge] < 18) &
                              (~fill_rows[columnNameTypeDoc].isin([61, 66, 64, 2482, 1640, 1639, 2482]))) |
                         (fill_rows[columnNameAge] >= 18) &
                              (~fill_rows[columnNameTypeDoc].isin([59, 62, 64, 65, 1637, 1638, 1640, 1639, 2482]))]
    #Cantidad de errores
    totalTypeDocErr = len(errors)
    for index in errors.index:
        setBgError(index, columnNameTypeDoc, bgSecError)
        setBgError(index, columnNameAge, bgSecError)
        
    if totalTypeDocErr > 0:
        print(f'Tipos de documento que no corresponden con la edad: {totalTypeDocErr}')
    return totalTypeDocErr

#Poblacional - Nacionalidad y tipo de documento
def nac_vs_typedoc(columnNameTypeDoc, columnNameNac, text_colombian_nation):    
    #Verificar existencia de las columnas
    if columnNameNac not in df_modified.columns or columnNameTypeDoc not in df_modified.columns:
        raise ValueError(f'Las columnas {columnNameTypeDoc} y/o {columnNameNac} no se encuentran')
    
    #Filtrar las filas no vacías
    fill_rows = df_modified[pd.notna(df_modified[columnNameTypeDoc]) & pd.notna(df_modified[columnNameNac])]
    errors = fill_rows[(fill_rows[columnNameTypeDoc].isin([59, 60, 61, '2- RC', '1- CC', '3- TI'])) &
                            (fill_rows[columnNameNac] != text_colombian_nation) | 
                       (~fill_rows[columnNameTypeDoc].isin([59, 60, 61, 66, 63, '2- RC', '1- CC', '3- TI'])) &
                            (fill_rows[columnNameNac] == text_colombian_nation)]
    
    totalNacErr = len(errors)
    for index in errors.index:
        setBgError(index, columnNameNac, bgSecError)
        setBgError(index, columnNameTypeDoc, bgSecError)
        
    if totalNacErr>0:
        print(f'Nacionalidad que no corresponde con el tipo de documento: {totalNacErr}')
    return totalNacErr

#Poblacional - Sexo y género
def gen_vs_sex(columnNameSex, columnNameGen):
    if columnNameGen not in df_modified.columns or columnNameSex not in df_modified.columns:
        raise ValueError(f"Las columnas {columnNameGen} y/o {columnNameSex} no se encuentran")
    
    fill_rows = df_modified[pd.notna(df_modified[columnNameGen]) & pd.notna(df_modified[columnNameSex])]
    errors = fill_rows[(fill_rows[columnNameSex] == '1- Hombre') & (fill_rows[columnNameGen] == '2- Femenino') |
                       (fill_rows[columnNameSex] == '2- Mujer') & (fill_rows[columnNameGen] == '1- Masculino')]
    
    totalGenErr = len(errors)
    for index in errors.index:
        setBgError(index, columnNameGen, bgSecError)
        setBgError(index, columnNameSex, bgSecError)
    
    if totalGenErr > 0:
        print(f'Errores en sexo y/o género: {totalGenErr}')
    return totalGenErr

#Poblacional - edad y estado civil
def age_vs_maritalStatus(columnNameInterventionDate, columnNameBirthDay, columnNameMarital):
    if columnNameBirthDay not in df_modified.columns or columnNameMarital not in df_modified.columns or columnNameInterventionDate not in df_modified.columns:
        raise ValueError(f"Las columnas {columnNameBirthDay}, {columnNameInterventionDate} y/o {columnNameMarital} no se encuentran")

    fill_rows = df_modified[pd.notna(df_modified[columnNameBirthDay]) & pd.notna(df_modified[columnNameMarital])]
    #fill_rows.loc[:,'Edad'] = fill_rows.apply(lambda row: calculate_age(row[columnNameBirthDay], row[columnNameInterventionDate]), axis=1)
    errors = fill_rows[(fill_rows['Edad'] < 14) & (fill_rows[columnNameMarital] != '6- No aplica') |
                       (fill_rows['Edad'] >= 14) & (fill_rows[columnNameMarital] == '6- No aplica')]
    
    totalErrMarital = len(errors)
    for index in errors.index:
        setBgError(index, columnNameMarital, bgError)
    
    if totalErrMarital > 0:
        print(f'Estado civil no concuerda con la edad: {totalErrMarital}')
    return totalErrMarital

#Poblacional - nacionalidad y población diferencial
def nac_vs_pdi(columnNameNac, columnNamePdi):
    if columnNameNac not in df_modified.columns or columnNamePdi not in df_modified.columns:
        raise ValueError(f'Las columnas {columnNameNac} y/o {columnNamePdi} no se encuentran')
    
    fill_rows = df_modified[pd.notna(df_modified[columnNameNac]) & pd.notna(df_modified[columnNamePdi])]
    
    pattern = re.compile(r'\bMigrante\b', flags=re.IGNORECASE)# Expresión regular
    errors = fill_rows[(fill_rows[columnNameNac] == 'Colombia') & (fill_rows[columnNamePdi].apply(lambda x: bool(pattern.search(str(x))))) |
                       (fill_rows[columnNameNac] != 'Colombia') & (~fill_rows[columnNamePdi].apply(lambda x: bool(pattern.search(str(x)))))]

    totalErrNacPdi = len(errors)
    
    for index in errors.index:
        setBgError(index, columnNamePdi, bgError)
    
    if totalErrNacPdi > 0:
        print(f'Errores en población diferencial: {totalErrNacPdi}')
    return totalErrNacPdi

#Poblacional - Habla español
def et_vs_lang(columnNameEt, columNameLang):
    if columnNameEt not in df_modified.columns or columNameLang not in df_modified.columns:
        raise ValueError(f'Las columnas {columnNameEt} y/o {columNameLang} no se encuentran')
    
    fill_rows = df_modified[pd.notna(df_modified[columnNameEt]) & pd.notna(df_modified[columNameLang])]
    errors = fill_rows[(fill_rows[columnNameEt] != '6- Ninguno') & (fill_rows[columNameLang] == -1)]
    totalErrEt = len(errors)
    
    for index in errors.index:
        setBgError(index, columNameLang, bgError)
    
    if totalErrEt > 0:
        print(f'Errores en habla español: {totalErrEt}')
    return totalErrEt

#Poblacional - EAPB
def afiliacion_eapb(columnNameTipo, columnNameEPS):
    cantErrAfiliacion = 0
    
    if columnNameEPS not in df_modified.columns or columnNameTipo not in df_modified.columns:
        raise ValueError(f'La(s) columna(s) {columnNameEPS} y/o {columnNameTipo}, no se encuentra(n)')
    
    fill_rows = df_modified[pd.notna(df_modified[columnNameEPS]) & pd.notna(df_modified[columnNameTipo])]
    errors = fill_rows[fill_rows[columnNameTipo].isin([0]) | fill_rows[columnNameEPS].isin([0]) |
                       ((fill_rows[columnNameTipo] == 135) & (fill_rows[columnNameEPS].str.strip() != 'NO ASEGURADO'))]
    
    cantErrAfiliacion += len(errors)
    for index in errors.index:
        setBgError(index, columnNameEPS, bgSecError)
        setBgError(index, columnNameTipo, bgSecError)
    
    if cantErrAfiliacion > 0:
        print(f'Errores en afiliación EAPB: {cantErrAfiliacion}')
    
    return cantErrAfiliacion

#Poblacional - validar ocupación
def ocupacion(columnNameBirthDay, columnNameInterventionDate, columnNameOccupation):
    cantErrOcupacion = 0
    if columnNameBirthDay not in df_modified.columns or columnNameInterventionDate not in df_modified.columns or columnNameOccupation not in df_modified.columns:
        raise ValueError(f"La(s) columna(s) {columnNameBirthDay}, {columnNameInterventionDate} y/o {columnNameOccupation} no se encuentra(n)")
    
    fill_rows = df_modified[pd.notna(df_modified[columnNameBirthDay]) & pd.notna(df_modified[columnNameOccupation])]
    pattern = re.compile(r'\ No Aplica\b', flags=re.IGNORECASE)# Expresión regular
    fill_rows['Edad'] = fill_rows.apply(lambda row: calculate_age(row[columnNameBirthDay], row[columnNameInterventionDate]), axis=1)
    errors = fill_rows[(fill_rows['Edad'] >= 18) & fill_rows[columnNameOccupation].apply(lambda x: bool(pattern.search(str(x)))) |
                       (fill_rows['Edad'] < 7) & (fill_rows[columnNameOccupation] != '12- No aplica | ')]
    
    cantErrOcupacion += len(errors)
    for index in errors.index:
        setBgError(index, columnNameOccupation, bgError)
        
    if cantErrOcupacion > 0:
        print(f'Errores en ocupación: {cantErrOcupacion}')
    return cantErrOcupacion

def discapacidad_categoria(columnNamePDI, columnNameCategoria):
    cantErrDisc = 0
    if columnNamePDI not in df_modified.columns or columnNameCategoria not in df_modified.columns:
        raise ValueError(f"La(s) columna(s) {columnNamePDI} y/o {columnNameCategoria} no se encuentra(n)")
    
    pattern = re.compile(r'\Discapacidad\b', flags=re.IGNORECASE)# Expresión regular
    fill_rows = df_modified[pd.notna(df_modified[columnNamePDI]) | pd.notna(df_modified[columnNamePDI])]
    errors = fill_rows[(fill_rows[columnNamePDI].apply(lambda x: bool(pattern.search(str(x)))) & pd.isna(fill_rows[columnNameCategoria])) |
                       (pd.notna(fill_rows[columnNameCategoria]) & ~fill_rows[columnNamePDI].apply(lambda x: bool(pattern.search(str(x)))))]
    
    cantErrDisc = len(errors)

    for index in errors.index:
        setBgError(index, columnNameCategoria, bgSecError)
        setBgError(index, columnNamePDI, bgSecError)

    if cantErrDisc > 0:
        print(f"Errores en categoría de la discapacidad: {cantErrDisc}")

    return cantErrDisc
#-----------------------------------SESIONES PÁGINA 1---------------------------------
def sc_pg1():
    requiredFieldsPg1 = ['.Nombre de la institución / Establecimiento / Equipo étnico.',
                   '.Zona.', '.Localidad.', '.UPZ/UPR.', '.Barrio.', '.Teléfono.',
                   '.Barrio priorizado.', '.Tipo de Institución.', '.Manzana de cuidado.']
    
    addressComponents = {
        'columnNameZone': '.Zona.',
        'columnNameAx1': '.Eje Principal.',
        'columnNameNumber':'.Número.',
        'columnNameAx2':'.Eje generador.',
        'columnNamePlate':'.Placa.',
        'columnNameTrail':'.Vereda.',
        'columnNameX':'.Coordenadas X.',
        'columnNameY':'.Coordenadas Y.'
    }
    
    columnsNames_only_numbers = ['.Número.', '.Eje generador.', '.Placa.', '.Teléfono.']
    catnErroresPg_1 = (required_fields(requiredFieldsPg1) + validarNoManzana(requiredFieldsPg1[8], '.Nro Manzana.') + validar_telefono(requiredFieldsPg1[5])
                       +validate_only_text_inst_barr(requiredFieldsPg1[0], requiredFieldsPg1[4])
                       +validate_address(addressComponents)+type_institution(requiredFieldsPg1[7], '.Otra. ¿Cual? (Tipo de Institución).')+validate_only_number(columnsNames_only_numbers))
    if catnErroresPg_1 == 0:
        print('Sin errores en la primera página')
    return catnErroresPg_1
#-----------------------------------SESIONES PÁGINA 2---------------------------------
def sc_pg2():
    requiredFieldsPg2 = ['.Componente.', '.Línea operativa.', '.Dimensión.', 
                         '.Temática.', '.Número sesión.', '.Fecha.', '.Nombre profesional 1.']
    cantErroresPg_2 = (required_fields(requiredFieldsPg2)+comparar_fechas_vs_intervencion('Fecha_intervencion', '.Fecha.')+validate_only_number(['.Número sesión.']))
    if cantErroresPg_2 == 0:
        print('Sin errores en la segunda página')
    return cantErroresPg_2

#-----------------------------------SESIONES PÁGINA 3---------------------------------
def sc_pg3():
    reqFieldsPg3 = ['..OMS..', '..FINDRISC..', '..EPOC..', '.Sesiones.', 'IdTipoDocumento',
                         'Documento', 'PrimerNombre','PrimerApellido', 'IdNacionalidad', 'IdSexo',
                         'IdGenero', 'IdEstadoCivil', 'FechaNacimiento', 'IdEtnia', 'PoblacionDiferencialInclusion']
    
    cantErroresPg_3 = (required_fields(reqFieldsPg3)+len_num_doc(reqFieldsPg3[4], reqFieldsPg3[5])
                       +age_vs_typedoc(reqFieldsPg3[4], 'Edad')+nac_vs_typedoc(reqFieldsPg3[4], reqFieldsPg3[8], 'Colombia')
                       +gen_vs_sex(reqFieldsPg3[9], reqFieldsPg3[10])+age_vs_maritalStatus('Fecha_intervencion',reqFieldsPg3[12],reqFieldsPg3[11])
                       +nac_vs_pdi(reqFieldsPg3[8], reqFieldsPg3[14])+et_vs_lang(reqFieldsPg3[13], 'HablaEspaniol')
                       +validate_only_text(reqFieldsPg3[6], reqFieldsPg3[7], 'SegundoNombre', 'SegundoApellido')
                       +fecha_mayor(reqFieldsPg3[12]))
    if cantErroresPg_3 == 0:
        print('Sin errores en la tercera página')
    return cantErroresPg_3

#-------------------------------------HCB PÁGINA 1-----------------------------------
def hcb_pg1():
    reqFieldsPg1 = ['.Zona.', '.Localidad.', '.UPZ/UPR.', '.Barrio.', '.Manzana de cuidado.', '.Barrio priorizado.', 
                    '.Estrato.', '.Nombre de la Asociación de madres comunitarias:.', '.Nombre del HCB.',
                    '.Teléfono 1.', '.¿Ha participado previamente en la estrategia AIEPI Comunitario?.',
                    '.¿Ha implementado la estrategia Mi Mascota Verde?.', '.¿Cuenta con huerta casera?.',
                    '.¿Cuenta con espacios de reunión periódica con los padres de familia?.', '.Lugar de Funcionamiento.',
                    '.¿De que servicios dispone?.']
    
    reqFieldsPg1Sec2 = ['.HCB en un lugar seguro (sin: remoción en masa, inundaciones - ronda hídrica, avalanchas).',
                        '.Paredes y techos sin grietas, huecos, humedades.','.Adecuado manejo de combustibles (sólidos, líquidos, gaseosos).',
                        '.Las áreas habitacionales de la vivienda están separadas entre sí (baño, cocinas y habitaciones).',
                        '.Preparación de alimentos con leña.','.La vivienda tiene iluminación y ventilación adecuada.',
                        '.Se fuma en la vivienda.','.Las condiciones físicas y locativas del baño son adecuadas.']
    
    addressComponents = {
        'columnNameZone': '.Zona.',
        'columnNameAx1': '..Tipo de vía..',
        'columnNameNumber':'..Número..',
        'columnNameAx2':'..Número..',
        'columnNamePlate':'..Placa..',
        'columnNameTrail':'.Vereda.',
        'columnNameX':'..Coordenadas X..',
        'columnNameY':'..Coordenadas Y..'
    }
    
    columnsNames_only_numbers = ['..Número..', '..Placa..', '.Estrato.', '..Perros..', '..Gatos..', '..Otros..', 
                           '..Perros...1','..Gatos...1', '..Perros...2', '..Gatos...2', '..Caracterización...56',
                           '..Evaluación...56', '..Caracterización...57', '..Evaluación...57', '..Caracterización...58',
                           '..Evaluación...58', '..Caracterización...59','..Evaluación...59','..Caracterización...60',
                           '..Evaluación...60', '..Caracterización...61', '..Evaluación...61','..Caracterización...62',
                           '..Evaluación...62','..Caracterización...63', '..Evaluación...63']
    
    cantErrorsPg_1 = (required_fields(reqFieldsPg1)+required_fields(reqFieldsPg1Sec2, 2, 1)+validar_telefono(reqFieldsPg1[9])
                      +validate_address(addressComponents)+validate_only_number(columnsNames_only_numbers))
    if cantErrorsPg_1 == 0:
        print("Sin errores en la primera página")
    return cantErrorsPg_1

#-------------------------------------HCB PÁGINA 2-----------------------------------
def hcb_pg2():
    #Campos requeridos
    reqFieldsPg2 = ['IdTipoDocumento','Documento', 'PrimerNombre','PrimerApellido', 'IdNacionalidad', 'IdSexo',
                    'IdGenero', 'IdEstadoCivil', 'FechaNacimiento', 'IdEtnia', 'PoblacionDiferencialInclusion',
                    'IdAfiliacionSGSSS', 'NombreEAPB', 'IdNivelEducativo', 'Ocupacion']
    #Campos de alertas
    fieldsAlerts = ['.Enfermedad Transmisible y ETV.', '.Condición crónica.', '.Alerta nutricional.', '.Alertas psicosociales.',
                    '.Alerta Salud Bucal.', '.Alerta infancia.', '.Alertas en mujeres.', '.Alertas discapacidad - Limitaciones para la actividad.']
    
    cantErroresPg_2 = (required_fields(reqFieldsPg2)+len_num_doc(reqFieldsPg2[0],reqFieldsPg2[1])
                       +age_vs_typedoc(reqFieldsPg2[0], 'Edad')+nac_vs_typedoc(reqFieldsPg2[0], reqFieldsPg2[4], 'Colombia')
                       +gen_vs_sex(reqFieldsPg2[5],reqFieldsPg2[6])+age_vs_maritalStatus('Fecha_intervencion',reqFieldsPg2[8],reqFieldsPg2[7])
                       +nac_vs_pdi(reqFieldsPg2[4], reqFieldsPg2[10])+et_vs_lang(reqFieldsPg2[9], 'HablaEspaniol')
                       +validate_alerts_hcb(fieldsAlerts)+tamizajes_vs_peso_hcb()+salud_bucal_hcb()
                       +afiliacion_eapb(reqFieldsPg2[11], reqFieldsPg2[12])+validate_only_text(reqFieldsPg2[2], reqFieldsPg2[3], 'SegundoNombre', 'SegundoApellido')
                       +sb_clasificacion_hcb()+ocupacion(reqFieldsPg2[8], 'Fecha_intervencion', reqFieldsPg2[14])
                       +(fecha_mayor(reqFieldsPg2[8])))
    if cantErroresPg_2 == 0:
        print("Sin errores en la segunda página")
    return cantErroresPg_2

#Formato de las alertas
def validate_alerts_hcb(columnsNames=[]):
    pattern = re.compile(r'^[^,]*\d(?:,\d+)*\d?[^\d,]*$')
    cantErrAlerts = 0
    for columnName in columnsNames:
        if columnName not in df_modified.columns:
            raise ValueError(f"No se encuentra la columna: {columnName}")
        
        #Obtener nombre de la siguiente columna
        columnIndexAlert = df_modified.columns.get_loc(columnName) + 1 #Indice de la columna principal más uno
        columnNameAlert = df_modified.columns[columnIndexAlert] if columnIndexAlert < len(df_modified.columns) else None #Nombre de la siguiente columna
        fill_rows = df_modified[pd.notna(df_modified[columnNameAlert])]
        for index, value in fill_rows[columnNameAlert].items():
            value_str = "{:.0f}".format(value) if not isinstance(value, str) else value
            if not pattern.match(str(value_str)):    
                cantErrAlerts += 1
                setBgError(index, columnNameAlert, bgError)
                
    if cantErrAlerts > 0:
        print(f'Errores en alertas: {cantErrAlerts}')
    return cantErrAlerts

#Validar campos de tamizajes con datos de peso, talla e IMC
def tamizajes_vs_peso_hcb():
    cantErrTam = 0
    colunmNameTamizajes = ['..Tamizaje OMS..','..Tamizaje FINDRISC..','..Tamizaje EPOC..']
    columnNamePti = ['..Peso Caracterización<br><br>..', '..Talla Caracterización<br><br>..','..Desvia. Están./IMC/Percentil Fen. Caracterización..']
    
    def find_errors(columnsNamesMain = [], columnNameSecond=[]):
        for columnName in columnsNamesMain:
            if columnName not in df_modified.columns:
                raise ValueError(f"La columna {columnName} no se encuentra")
            
            fill_rows = df_modified[pd.notna(df_modified[columnName])]
            errors = fill_rows[(pd.notna(fill_rows[columnName]) & (pd.isna(fill_rows[columnNameSecond[0]]) |
                                                                pd.isna(fill_rows[columnNameSecond[1]]) |
                                                                pd.isna(fill_rows[columnNameSecond[2]])))]
            for index in errors.index:
                setBgError(index, columnNameSecond[0], bgError)
                setBgError(index, columnNameSecond[1], bgError)
                setBgError(index, columnNameSecond[2], bgError)
                
            return len(errors)
        
    cantErrTam = find_errors(colunmNameTamizajes,columnNamePti) + find_errors(columnNamePti,colunmNameTamizajes)
    
    if cantErrTam > 0:
        print(f"Errores en tamizajes: {cantErrTam}")
    return cantErrTam

#Validar campo de priorizado salud bucal vs alerta salud bucal
def salud_bucal_hcb():
    cantErrSb = 0
    columnsNames = ['..Caracterización - Clasificación..', '..Caracterización - Alertas..', '..Priorizado para acompañamiento familiar..']
    
    for columnName in columnsNames:
        if columnName not in df_modified.columns:
            raise ValueError(f'La columna {columnName} no se encuentra')
        
        fill_rows = df_modified[pd.notna(df_modified[columnName])]
        if columnName != '..Priorizado para acompañamiento familiar..':
            errors = fill_rows[(pd.notna(fill_rows[columnName]) & pd.isna(fill_rows[columnsNames[2]])) |
                               (pd.notna(fill_rows[columnsNames[2]]) & pd.isna(fill_rows[columnName]))]
            cantErrSb += len(errors)
            for index in errors.index:
                setBgError(index, columnsNames[0], bgSecError)
                setBgError(index, columnsNames[1], bgSecError)
                setBgError(index, columnsNames[2], bgSecError)
                
        elif columnName == '..Priorizado para acompañamiento familiar..':
            errors = fill_rows[~fill_rows[columnsNames].isin(['SI', 'NO']).any(axis=1)]
            cantErrSb += len(errors)
            for index in errors.index:
                setBgError(index, columnName, bgSecError)
    
    if cantErrSb > 0:
        print(f'Errores en salud bucal: {cantErrSb}')
    return cantErrSb

#Validar formato clasificación
def sb_clasificacion_hcb():
    cantErrClas = 0
    columnsNames = ['..Caracterización - Clasificación..', '..Evaluación - Clasificación..']

    for columnName in columnsNames:
        if columnName not in df_modified.columns:
            raise ValueError(f'La columna {columnName} no se encuentra')
        
        fill_rows = df_modified[pd.notna(df_modified[columnName])]
        errors = fill_rows[~fill_rows[columnName].astype(str).str.endswith('%')]
        
        cantErrClas += len(errors)
        
        for index in errors.index:
            setBgError(index, columnName, bgError)
            
    if cantErrClas > 0:
        print(f'Campo de clasificación a los que les falta signo "%" {cantErrClas}')
    return cantErrClas

#------------------------------MASCOTA VERDE PÁGINA 1---------------------------------
def mv_pg1():
    reqFieldsPg1 = ['.NOMBRE DEL HCB.', '.NOMBRE DE LA MADRE COMUNITARIA.', '.Zona.', '.Localidad.', '.UPZ/UPR.', '.Barrio.',
                    '.Manzana del cuidado.', '.Barrio priorizado.', '.Teléfono.', '.Nombre profesional 1.', '.Nombre profesional 2.']
    
    addressComponents = {
        'columnNameZone': '.Zona.',
        'columnNameAx1': '..Tipo de vía..',
        'columnNameNumber':'..Número..',
        'columnNameAx2':'..Número..',
        'columnNamePlate':'..Placa..',
        'columnNameTrail':'.Vereda.',
        'columnNameX':'..Coordenadas X..',
        'columnNameY':'..Coordenadas Y..'
    }
    
    cantErroresPg_1 = (required_fields(reqFieldsPg1)+validar_telefono(reqFieldsPg1[8])+validate_address(addressComponents)
                       +validate_only_text_inst_barr('.NOMBRE DEL HCB.', '.NOMBRE DE LA MADRE COMUNITARIA.', '.Barrio.', '.Vereda.')
                       +validate_only_text('..Letra..', '.Nombre profesional 1.', '.Nombre profesional 2.'))
    if cantErroresPg_1 == 0:
        print("Sin errores en la primera página")
    return cantErroresPg_1

#------------------------------MASCOTA VERDE PÁGINA 2---------------------------------
def mv_pg2():
    reqFieldsPg2 = ['IdTipoDocumento','Documento', 'PrimerNombre','PrimerApellido', 'IdNacionalidad', 'IdSexo',
                    'IdGenero', 'IdEstadoCivil', 'FechaNacimiento', 'IdEtnia', 'PoblacionDiferencialInclusion']
    
    columnsNamesNumSes = ['.Numero.', '.Numero..1', '.Numero..2', '.Numero..3', '.Numero..4', '.Numero..5']

    columnsNamesSesions = {
        "fields_s1":['.Numero.', '.Fecha.', '.Nombre profesional 1.'],
        "fields_s2":['.Numero..1','.Fecha..1', '..Temática...1', '..Alertas Psicosociales. Separados por coma...1', '..Nivel de crecimiento...1', 
                   '..Mantenimiento...1', '..Estado de la mascota...1', '.Nombre profesional 1..1'],
        "fields_s3":['.Numero..2', '.Fecha..2', '..Temática...2', '..Alertas Psicosociales. Separados por coma...2', '..Nivel de crecimiento...2',
                     '..Mantenimiento...2', '..Estado de la mascota...2', '.Nombre profesional 1..2']
    }

    cantErroresPg_2 = (required_fields(reqFieldsPg2)+len_num_doc(reqFieldsPg2[0],reqFieldsPg2[1])
                       +age_vs_typedoc(reqFieldsPg2[0], 'Edad')+nac_vs_typedoc(reqFieldsPg2[0], reqFieldsPg2[4], 'Colombia')
                       +gen_vs_sex(reqFieldsPg2[5],reqFieldsPg2[6])+age_vs_maritalStatus('Fecha_intervencion', reqFieldsPg2[8],reqFieldsPg2[7])
                       +nac_vs_pdi(reqFieldsPg2[4], reqFieldsPg2[10])+et_vs_lang(reqFieldsPg2[9], 'HablaEspaniol')
                       +validate_only_text('.Nombre de la mascota.', '.Nombre profesional 1.', '.Nombre profesional 1..1')+validateNumSesion_mv(columnsNamesNumSes)
                       +comparar_fechas_vs_intervencion('Fecha_intervencion','.Fecha.')+validateNameProfesional_mv()+validateDates_mv()+validate_pet_name(columnsNamesSesions))
    if cantErroresPg_2 == 0:
        print("Sin errores en la segunda página")
    return cantErroresPg_2

def validateNameProfesional_mv():
    cantErrProf = 0
    #Números de columna de profesionales
    columnsNumberProf = [24, 25, 34, 35, 44, 45, 54, 55]
    columnNamesProf = {
        "columnNameP1S1":'',
        "columnNameP2S1":'',
        "columnNameP1S2":'',
        "columnNameP2S2":'',
        "columnNameP1S3":'',
        'columnNameP2S3':'',
        "columnNameP1S4":'',
        "columnNameP2S4":''
    }
    lista = list(columnNamesProf.items()) #Lista de tuplas

    def findError(columnNameP1, columnNameP2):
        foundErrors = 0
        if columnNameP1 not in df_modified.columns or columnNameP2 not in df_modified.columns:
            raise ValueError(f'La(s) columna(s) {columnNameP1} y/o {columnNameP2} no se encuentra(n)')

        errors = df_modified[(pd.notna(columnNameP1) & (df_modified[columnNameP1] == df_modified[columnNameP2])) |
                             (pd.notna(columnNameP2) & (df_modified[columnNameP2] == df_modified[columnNameP1]))]
        foundErrors += len(errors)

        for index in errors.index:
            setBgError(index, columnNameP1, bgSecError)
            setBgError(index, columnNameP2, bgSecError)
        
        return foundErrors
    
    for index, number in enumerate(columnsNumberProf):
        key, value = lista[index]
        columnNamesProf[key] = getColumnNameByNumber(number)

    cantErrProf = (findError(columnNamesProf['columnNameP1S1'], columnNamesProf['columnNameP2S1'])+
                findError(columnNamesProf['columnNameP1S2'], columnNamesProf['columnNameP2S2'])+
                findError(columnNamesProf['columnNameP1S3'], columnNamesProf['columnNameP2S3'])+
                findError(columnNamesProf['columnNameP1S4'], columnNamesProf['columnNameP2S4']))
        
    if cantErrProf > 0:
        print(f'Errores en nombre de profesional: {cantErrProf}')
    return cantErrProf

def validateNumSesion_mv(columnsNames):
    cantErrNumSes = 0

    for index, columnName in enumerate(columnsNames):
        if columnName not in df_modified.columns:
            raise ValueError(f'La columna {columnName} no se encuentra')

        fill_rows = df_modified[pd.notna(df_modified[columnName])]
        errors = fill_rows[fill_rows[columnName] != index+1]

        cantErrNumSes += len(errors)
        for index in errors.index:
            setBgError(index, columnName, bgError)

    if cantErrNumSes > 0:
        print(f"Número de sesión incorrecto: {cantErrNumSes}")
    return cantErrNumSes
    
def validateDates_mv():
    totalErrDate = 0
    columnsNamesDates = ['.Fecha.', '.Fecha..1', '.Fecha..2', '.Fecha..3', '.Fecha..4', '.Fecha..5']

    for index, columnName in enumerate(columnsNamesDates):
        if columnName not in df_modified.columns:
            raise ValueError(f"La columna {columnName} no existe")
        
        if index != 0:
            totalErrDate += compare_dates(columnsNamesDates[index-1], columnsNamesDates[index])
    
    if totalErrDate > 0:
        print(f'Fecha incoherente: {totalErrDate}')

    return totalErrDate

#Verificar que si tiene nombre de mascota tenga asistencia a la sesión 2 en adelante
def validate_pet_name(columnsNamesSesions):
    cantErrPetName = 0
    columnNamePet = '.Nombre de la mascota.'

    if columnNamePet not in df_modified.columns:
        raise ValueError(f"La columna {columnNamePet} no existe")

    for key, columns in columnsNamesSesions.items():
        if key == 'fields_s1':
            continue
        print(columns)
        for index, columnName in enumerate(columns):
            if index != 3:
                if columnName not in df_modified.columns:
                    raise ValueError(f"La columna {columnName} no existe")
            
                errors = df_modified[pd.notna(df_modified[columnNamePet]) & pd.isna(df_modified[columnName])]
                cantErrPetName += len(errors)

                for idx in errors.index:
                    setBgError(idx, columnName, bgSecError)
                    setBgError(idx, columnNamePet, bgSecError)

    # Imprime si hay errores
    if cantErrPetName > 0:
        print(f'Error en nombre de mascota/nro sesión: {cantErrPetName}')

    return cantErrPetName

def validateDataSesions_mv():
    cantErrSes = 0
    
#------------------------------PERSONA MAYOR PÁGINA 1---------------------------------
def pm_pg1():
    reqFieldsPg1 = ['.Nombre de la institución.', '.Zona.', '.Localidad.', '.UPZ/UPR.', '.Barrio priorizado.', '.Manzana del cuidado.',
                    '.Teléfono1.', '.Correo1.', '.Primer nombre.', '.Primer apellido.', '.Tipo documento.', '.Documento.', '.Total de personas adultas en la institución(Hombres).',
                    '.Total de personas adultas en la institución(Mujeres).', '.Total de personas mayores en la institución(Hombres).', '.Total de personas mayores en la institución(Mujeres).',
                    '.Número de personas que requieren controles de salud periódicos para el manejo y seguimiento a diagnósticos y tratamiento.', '.Cuantas de ellas asisten a los servicios de salud.',
                    '..Permanecen solos en casa (Abandono de la red familiar)..', '..Necesitan cuidados especiales..', '..Pacientes crónicos..', '..Situación económica de la familia..',
                    '..Personas en condición de vulnerabilidad (Habitante de calle, VCA)..', '..Voluntad propia..']
    
    reqFieldspg1_next = ['.¿Cuenta con plataforma estratégica (visión, misión, objetivos/principios, valores, políticas)?.', '.¿Cuenta con proceso de administración de información?.',
                         '.¿Maneja formato de ingreso al usuario?.', '.¿Maneja formato de egreso del usuario?.', '.Realiza seguimiento periódico de historias clínicas.',
                         '.¿Cuenta con registro de actividades y acciones realizadas con los usuarios?.', '.¿Cuenta con registro de las actividades realizadas con los familiares de las personas institucionalizadas?.',
                         '.¿Cuenta con un programa de atención integral para el Desarrollo Humano?.', '.¿Existen planes de atención individuales orientados a fortalecer capacidades y habilidades de las personas mayores?.',
                         '.¿Implementa prácticas promocionales para la salud y realiza registro de las mismas?.', '.¿Cuenta con la guía de buen trato?.', '.¿Desarrolla actividades de promoción del buen trato?.',
                         '.¿Desarrolla actividades de prevención del maltrato?.', '.¿Desarrollan actividades colectivas (participación, encuentros)?.', '.¿Desarrollan formación del talento humano en bienestar de ellos y de las personas mayores?.',
                         '.¿El personal de enfermería se encuentra capacitado para prestar servicios integrales y de calidad acorde a las condiciones de salud de la población?.',
                         '.¿Cuenta con protocolo de aislamiento?.', '.¿En la institución se dispone de un lugar específico de aislamiento?.', '.¿Se cuenta con protocolo de limpieza y desinfección al ingreso de la institución?.',
                         '.¿Durante el desarrollo de actividades en la institución, se conserva el distanciamiento social?.', '.¿Implementan técnica de lavado de manos por lo menos cada tres horas?.',
                         '.¿Todas las personas que se encuentran en la institución utilizan tapabocas?.', '.¿Se cuenta con ruta sanitaria activa en la institución?.']
    
    addressComponents = {
        'columnNameZone': '.Zona.',
        'columnNameAx1': '.Eje Principal.',
        'columnNameNumber':'.Número.',
        'columnNameAx2':'.Eje generador.',
        'columnNamePlate':'.Placa.',
        'columnNameTrail':'.Vereda.',
        'columnNameX':'.Coordenadas X.',
        'columnNameY':'.Coordenadas Y.'
    }

    columnsNames_only_numbers = ['.Documento.', '.Teléfono1.', '.Teléfono2.', '.Placa.', '.Eje generador.', '.Número.']

    cantErroresPg_1 = (required_fields(reqFieldsPg1)+required_fields_next_column(reqFieldspg1_next)+validar_telefono(reqFieldsPg1[6], '.Teléfono2.')+validarNoManzana(reqFieldsPg1[5],'.Numero de manzana.')+
                       +validate_address(addressComponents)+validate_email(reqFieldsPg1[7], '.Correo1.')+validate_only_text(reqFieldsPg1[8], reqFieldsPg1[9], '.Segundo nombre.', '.Segundo apellido.')
                       +validate_only_text_inst_barr(reqFieldsPg1[0], '.Barrio.', '.Vereda.')+validate_only_number(columnsNames_only_numbers)+validate_type_doc_pm('.Tipo documento.')
                       +comparar_fechas_vs_intervencion('Fecha_intervencion','.Fecha evaluación.'))

    if cantErroresPg_1 == 0:
        print("Sin errores en la primera página")
    return cantErroresPg_1

def validate_evaluation_fields(columnNameDate, numColumns, columnsNamesEva = []):
    totalErrorsEva = 0
    if columnNameDate not in df_modified.columns or numColumns not in df_modified.columns:
        raise ValueError(f'La columna {columnName} y/o {columnNameDate} no se encuentra')
    
    for columnName in columnsNamesEva:
        if columnName not in df_modified.columns:
            raise ValueError(f'La columna {columnName} no se encuentra')
        
        fill_rows = df_modified[pd.notna(df_modified[columnName])]
        errors = fill_rows[(pd.notna(fill_rows[columnName]) & pd.isna(fill_rows[columnNameDate])) |
                           pd.notna(fill_rows[columnNameDate]) & pd.isna(fill_rows[columnName])]
        totalErrorsEva += len(errors)
        for index in errors.index:
            setBgError(index, columnName, bgSecError)
            setBgError(index, columnNameDate, bgSecError)

    if totalErrorsEva > 0:
        print(f"Campos de evaluación sin la fecha: {totalErrorsEva}")
    return totalErrorsEva

def validate_type_doc_pm(columnName):
    totalErrorsDoc = 0
    if columnName not in df_modified.columns:
        raise ValueError(f'La columna {columnName} no se encuentra')

    fill_rows = df_modified[pd.notna(df_modified[columnName])]
    errors = fill_rows[fill_rows[columnName].isin(['2- RC', '3- TI', '8- Menor sin ID.', 60, 61, 66])]

    totalErrorsDoc += len(errors)
    for index in errors.index:
        setBgError(index, columnName, bgError)

    if totalErrorsDoc > 0:
        print(f'Tipos de documento que no corresponden con persona mayor: {totalErrorsDoc}')

    return totalErrorsDoc
#------------------------------PERSONA MAYOR PÁGINA 2---------------------------------
def pm_pg2():
    reqFieldsPg2 = ['.Persona en abandono.', 'IdTipoDocumento', 'Documento', 'PrimerNombre',
                    'PrimerApellido', 'IdNacionalidad', 'IdSexo', 'IdGenero', 'IdEstadoCivil',
                    'FechaNacimiento', 'IdEtnia', 'IdAfiliacionSGSSS', 'NombreEAPB', 'IdNivelEducativo',
                    'PoblacionDiferencialInclusion']
    
    reqFieldspg2_next = ['.Condición crónica.', '.Enfermedad transmisible y ETV.', '.Alerta nutricional.', '.Alerta Psicosociales.', '.Alerta salud oral.']

    cantErroresPg_2 = (required_fields(reqFieldsPg2)+age_vs_typedoc(reqFieldsPg2[1], 'Edad')+nac_vs_typedoc(reqFieldsPg2[1],reqFieldsPg2[5], 'Colombia')
                       +gen_vs_sex(reqFieldsPg2[6], reqFieldsPg2[7])+age_vs_maritalStatus('Fecha_intervencion',reqFieldsPg2[9],reqFieldsPg2[8])
                       +nac_vs_pdi(reqFieldsPg2[5], reqFieldsPg2[14])+et_vs_lang(reqFieldsPg2[10], 'HablaEspaniol')
                       +discapacidad_categoria(reqFieldsPg2[14], 'CategoriasDiscapacidad')+required_fields_next_column(reqFieldspg2_next)
                       +validate_only_text(reqFieldsPg2[3], 'SegundoNombre', reqFieldsPg2[4], 'SegundoApellido')+validate_type_doc_pm(reqFieldsPg2[1])
                       +fecha_mayor(reqFieldsPg2[9]))
    if cantErroresPg_2 == 0:
        print("Sin errores en la segunda página")
    return cantErroresPg_2

#------------------------PLAN DE CUIDADO INSTITUCIONAL PÁGINA 1---------------------------
def pci_pg1():
    reqFieldsPg1 = ['.Nombre institución.', '.Zona.', '.Localidad.', '.UPZ/UPR.', '.Barrio.', '.Teléfono1.',
                    '.Manzana del cuidado.', '.Barrio Priorizado.', '.Tipo institución.']
    
    addressComponents = {
        'columnNameZone': '.Zona.',
        'columnNameAx1': '.Eje Principal.',
        'columnNameNumber':'.Número.',
        'columnNameAx2':'.Eje generador.',
        'columnNamePlate':'.Placa.',
        'columnNameTrail':'.Vereda.',
        'columnNameX':'.Coordenadas X.',
        'columnNameY':'.Coordenadas Y.'
    }
    
    columnsNames_only_numbers = [reqFieldsPg1[5], '.Teléfono2.', '.Número.', '.Eje generador.', '.Placa.']
    
    cantErroresPg_1 = (required_fields(reqFieldsPg1)+validate_address(addressComponents)+validar_telefono(reqFieldsPg1[5], '.Teléfono2.')
                       +validarNoManzana(reqFieldsPg1[6], '.Nro Manzana.')+type_institution(reqFieldsPg1[8], '.Tipo institución	. Otra.')
                       +validate_only_text_inst_barr(reqFieldsPg1[0], reqFieldsPg1[4])+validate_only_number(columnsNames_only_numbers))
    
    if cantErroresPg_1 == 0:
        print("Sin errores en la primera página")
    return cantErroresPg_1

#------------------------PLAN DE CUIDADO INSTITUCIONAL PÁGINA 2---------------------------
def pci_pg2():
    reqFieldsPg2 = ['.Sesión.','.Línea operativa.', '.Dimensión.', '.Temática.']
    
    cantErroresPg_2 = (required_fields(reqFieldsPg2)+validate_only_number([reqFieldsPg2[0], '.Nro. participantes.'])
                       +comparar_fechas_vs_intervencion('Fecha_intervencion', '.Fecha.')+fecha_mayor('.Fecha.'))
    
    if cantErroresPg_2 == 0:
        print("Sin errores en la segunda página")
    return cantErroresPg_2
#------------------------PLAN DE CUIDADO INSTITUCIONAL PÁGINA 3---------------------------
def pci_pg3():
    reqFieldsPg3 = ['.Nro. sesión.', '.Línea operativa.', '.Fecha.', '.Descripción.']
    
    cantErroresPg_3 = (required_fields(reqFieldsPg3)+comparar_fechas_vs_intervencion('Fecha_intervencion', '.Fecha.')
                       +fecha_mayor('.Fecha.')+validate_only_number([reqFieldsPg3[0], reqFieldsPg3[1]]))
    
    if cantErroresPg_3 == 0:
        print("Sin errores en la segunda página")
    return cantErroresPg_3
##---------------------------------ESCALA ABREVIADA-----------------------------------
def ead_pg1():
    reqFieldsPg1 = ['.Fecha Caracterización.', '.Entorno.', '.Nombre del HCB/Jardín.', '.Localidad.', '.1er. NOMBRE.', '.1er. APELLIDO.',
                    '.Tipo de documento.', '.N° Documento:.', '.Nacionalidad.', '.Población diferencial y de inclusión.', '.Nombre Digitador.']
    
    reqFieldsPg1_next_1 = ['.Fecha de caracterización.', '.Fecha de nacimiento.', '.Edad del niño (a).', '.Total Acumulado al inicio.', '.Número de Items correctos.',
                           '.Total (PD).', '.total puntuación típica PT.', '.Puntuación típica PT.', '.Edad (en meses/dias).', '.Nivel de desarrollo del niño o niña - Caracterización..']
    
    cantErroresPg_1 = (required_fields(reqFieldsPg1)+required_fields(reqFieldsPg1_next_1, 2, 1)+validate_only_text_inst_barr('.Nombre del HCB/Jardín.')
                       +validate_only_text('.1er. NOMBRE.', '.2do. NOMBRE.', '.1er. APELLIDO.', '.2do. APELLIDO.', '.Nombre Digitador.')
                       +nac_vs_typedoc(reqFieldsPg1[6],reqFieldsPg1[8],'COL')+len_num_doc(reqFieldsPg1[6], reqFieldsPg1[7])+comparar_fechas_vs_intervencion('Fecha_intervencion', '.Fecha Caracterización.'))

    return cantErroresPg_1

##------------------------------------------------------------------------------------
##------------------------------------------------------------------------------------

def saveFile():
    response_save = messagebox.askquestion("Guardar archivo", "¿Guardar el archivo generado?")
    if response_save == "yes":
        cadenaGuardar = "Guardando archivo..."
        print(cadenaGuardar)
        
        try:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            file_path_modificado = file_path.replace('.xlsx', '_errores.xlsx')
            workbook.save(file_path_modificado)
            
            print("¡El archivo ha sido guardado correctamente!")
            # Preguntar al usuario si desea abrir el archivo guardado
            open_file = messagebox.askquestion("Abrir Archivo", "¿Desea abrir el archivo guardado?")
            if open_file == 'yes':
                os.startfile(file_path_modificado)# Abre el archivo guardado
                print(f"El archivo quedó guardado en la ruta {file_path_modificado}")
        except Exception as e:
            print("Error", f"No se pudo guardar el archivo: {str(e)}")
    else:
            print("Tu archivo no se guardará")
            
#En prueba para los tres puntos cargando
def cargandoSave(cadena):
    if cadena:
        # Borra el último carácter usando slicing
        nueva_cadena = cadena[:-3]
        return nueva_cadena
    else:
        return cadena