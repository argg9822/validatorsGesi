"""
validadores/beneficiarios.py
Validador de beneficiarios del sistema GESI.

Interfaz estándar:
    validate(file_path: str, progress_cb=None) -> dict
        progress_cb(pct: int, message: str)
        Retorna: {"ok": bool, "message": str, "errors": list, "output_file": str|None}
"""

import re
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl", "--quiet"])
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font


# ── Reglas de validación configurables ───────────────────────────────────────
COLUMNAS_REQUERIDAS = [
    "TIPO_DOC", "NUM_DOC", "PRIMER_NOMBRE", "PRIMER_APELLIDO",
    "FECHA_NACIMIENTO", "SEXO", "MUNICIPIO"
]

TIPOS_DOC_VALIDOS = {"CC", "TI", "RC", "CE", "NIT", "PA", "PE"}
SEXOS_VALIDOS     = {"M", "F", "MASCULINO", "FEMENINO"}
LONG_DOC_MAX      = 15


def _progress(cb, pct: int, msg: str):
    if cb:
        cb(pct, msg)


def validate(file_path: str, progress_cb=None) -> dict:
    errors = []
    _progress(progress_cb, 10, "Leyendo archivo Excel...")

    try:
        df = pd.read_excel(file_path, dtype=str)
    except Exception as e:
        return {"ok": False, "message": f"Error leyendo archivo: {e}", "errors": [], "output_file": None}

    # Normalizar columnas
    df.columns = [c.strip().upper().replace(" ", "_") for c in df.columns]

    _progress(progress_cb, 20, "Verificando columnas requeridas...")

    # ── Columnas faltantes ────────────────────────────────────────────────────
    missing = [c for c in COLUMNAS_REQUERIDAS if c not in df.columns]
    if missing:
        return {
            "ok": False,
            "message": f"Columnas faltantes: {', '.join(missing)}",
            "errors": [{"fila": 0, "columna": c, "error": "Columna faltante"} for c in missing],
            "output_file": None
        }

    total_rows = len(df)
    _progress(progress_cb, 30, f"Validando {total_rows} registros...")

    for idx, row in df.iterrows():
        fila = idx + 2  # Excel empieza en fila 2

        # Tipo de documento
        tipo_doc = str(row.get("TIPO_DOC", "")).strip().upper()
        if not tipo_doc or tipo_doc == "NAN":
            errors.append({"fila": fila, "columna": "TIPO_DOC", "error": "Campo vacío"})
        elif tipo_doc not in TIPOS_DOC_VALIDOS:
            errors.append({"fila": fila, "columna": "TIPO_DOC",
                           "error": f"Tipo inválido: '{tipo_doc}'. Válidos: {TIPOS_DOC_VALIDOS}"})

        # Número de documento
        num_doc = str(row.get("NUM_DOC", "")).strip()
        if not num_doc or num_doc == "nan":
            errors.append({"fila": fila, "columna": "NUM_DOC", "error": "Campo vacío"})
        elif not num_doc.isdigit():
            errors.append({"fila": fila, "columna": "NUM_DOC",
                           "error": f"Debe contener solo dígitos: '{num_doc}'"})
        elif len(num_doc) > LONG_DOC_MAX:
            errors.append({"fila": fila, "columna": "NUM_DOC",
                           "error": f"Supera {LONG_DOC_MAX} dígitos"})

        # Nombres vacíos
        for campo in ["PRIMER_NOMBRE", "PRIMER_APELLIDO"]:
            val = str(row.get(campo, "")).strip()
            if not val or val.lower() == "nan":
                errors.append({"fila": fila, "columna": campo, "error": "Campo vacío"})

        # Fecha de nacimiento
        fecha_str = str(row.get("FECHA_NACIMIENTO", "")).strip()
        if not fecha_str or fecha_str.lower() == "nan":
            errors.append({"fila": fila, "columna": "FECHA_NACIMIENTO", "error": "Campo vacío"})
        else:
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
                try:
                    dt = datetime.strptime(fecha_str, fmt)
                    if dt.year < 1900 or dt > datetime.now():
                        errors.append({"fila": fila, "columna": "FECHA_NACIMIENTO",
                                       "error": f"Fecha fuera de rango: {fecha_str}"})
                    break
                except ValueError:
                    continue
            else:
                errors.append({"fila": fila, "columna": "FECHA_NACIMIENTO",
                               "error": f"Formato inválido: '{fecha_str}'"})

        # Sexo
        sexo = str(row.get("SEXO", "")).strip().upper()
        if not sexo or sexo == "NAN":
            errors.append({"fila": fila, "columna": "SEXO", "error": "Campo vacío"})
        elif sexo not in SEXOS_VALIDOS:
            errors.append({"fila": fila, "columna": "SEXO",
                           "error": f"Valor inválido: '{sexo}'"})

        # Progreso dinámico
        if total_rows > 0 and (idx % max(1, total_rows // 10)) == 0:
            pct = 30 + int((idx / total_rows) * 50)
            _progress(progress_cb, pct, f"Fila {fila} de {total_rows + 1}...")

    _progress(progress_cb, 82, "Generando reporte de errores...")
    output_file = _generate_report(file_path, df, errors)

    _progress(progress_cb, 100, "¡Validación completada!")
    n = len(errors)
    return {
        "ok":          n == 0,
        "message":     "Sin errores ✅" if n == 0 else f"Se encontraron {n} error(es). Reporte guardado.",
        "errors":      errors,
        "output_file": output_file
    }


def _generate_report(original_path: str, df: pd.DataFrame, errors: list) -> str:
    """Genera un Excel con las filas con error marcadas en rojo."""
    if not errors:
        return None

    output_dir = Path(original_path).parent / "output"
    output_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"reporte_beneficiarios_{ts}.xlsx"

    error_map: dict[int, list] = {}
    for e in errors:
        error_map.setdefault(e["fila"], []).append(e)

    wb = openpyxl.load_workbook(original_path)
    ws = wb.active

    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font   = Font(color="9C0006", bold=True)

    # Agregar columna de errores
    err_col = ws.max_column + 1
    ws.cell(row=1, column=err_col, value="ERRORES_VALIDACION").font = Font(bold=True)

    for fila_excel, row_errors in error_map.items():
        msgs = " | ".join(f"{e['columna']}: {e['error']}" for e in row_errors)
        for col in range(1, ws.max_column):
            cell = ws.cell(row=fila_excel, column=col)
            cell.fill = red_fill
            cell.font = red_font
        ws.cell(row=fila_excel, column=err_col, value=msgs)

    wb.save(out_path)
    return str(out_path)