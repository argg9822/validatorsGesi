import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, numbers
from openpyxl.utils import get_column_letter
import pandas as pd
import re
import shutil
from colorama import init, Fore, Style
import os
import datetime
import time



def mostrar_ventana_progreso(titulo, max_val):
    # Crear una nueva ventana para mostrar el progreso
    ventana = tk.Tk()
    ventana.title(titulo)

    # Crear y colocar una etiqueta para mostrar el texto de la función en ejecución
    label = tk.Label(ventana, text=titulo)
    label.pack(pady=10)

    # Crear y colocar la barra de progreso
    progress = ttk.Progressbar(ventana, orient="horizontal", length=300, mode="determinate")
    progress.pack(pady=20)

    return ventana, progress, label

def actualizar_barra_progreso(ventana, progress, valor):
    progress['value'] = valor
    ventana.update_idletasks()

#validadores educativo 
init()
# Declarar el objeto colum inicialmente para almacenar variables para las funciones
colum = {"column": set(), "row": 0}
celTexto = {"ColumText": set()}
Genero = {"Genero": set()}
etniaVal = {"etniaVal": set()}
afiliacion = {"afiliacion": set()}
CeldasVacias = {"vacias": set()}
CeldasVacias_Condicional = {"vacias": set(), "row": 0}

placas = {"placas": set()}
Tel = {"Tel": set()}
Manzana = {"Manzana": set()}
rural = {"rural": set()}

    
def setBase(base):
    loadExcel()
    chooseBase(base)
    preguntaDescarga()
    
    
# llamar o cargar archivo excel
def loadExcel():
    # Abrir el archivo Excel
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return
    
    global workbook
    workbook = openpyxl.load_workbook(file_path)
    global sheet
    sheet = workbook.active
    
    # Leer los encabezados de la primera fila
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # Retornar los encabezados y la hoja de cálculo
    return headers, sheet
    
def chooseBase(base):
    switch = {
        "sesiones_colectivas": SesionesCoelctivas,
        "prevencion_embarazo": PrevencionEmbarazo,
        "higiene_manos": higieneManos,
        "pretest": pretest,
        "jornadas": jornadas, 
        "autocuidado": autocuidado,
        "mascota_verde": mascota_verde,
        "salud_mental": salud_mental,
        "higiene_bucal": higiene_bucal
    }
    execute_validator = switch.get(base)
    execute_validator()
    
def mascota_verde():
    # Páginas del archivo Excel cargado
    num_paginas = len(workbook.sheetnames)
    print(f"Se validaron {num_paginas} páginas.")
    # Primero, validar la página 1
    if num_paginas >= 1 and workbook.sheetnames[0] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]  # Acceder a la página 1
        print("Página 1...")
        mascota_pag1(sheet)
    
    if num_paginas >= 1 and workbook.sheetnames[1] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[1]]  # Acceder a la página 1
        print("Página 2...")
        mascota_pag2(sheet)
def higiene_bucal():
    num_paginas = len(workbook.sheetnames)
    print(f"El archivo Excel tiene {num_paginas} páginas.")
    # Primero, validar la página 1
    if num_paginas >= 1 and workbook.sheetnames[0] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]  # Acceder a la página 1
        print("Página 1...")
        hb_pag1(sheet)
        
    if num_paginas >= 1 and workbook.sheetnames[1] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[1]]  # Acceder a la página 1
        print("Validando la página 2...")
        hb_pag2(sheet)
    
def salud_mental():
    num_paginas = len(workbook.sheetnames)
    print(f"El archivo Excel tiene {num_paginas} páginas.")
    # Primero, validar la página 1
    if num_paginas >= 1 and workbook.sheetnames[0] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]  # Acceder a la página 1
        print("Validando la página 1...")
        saludmental_pag1(sheet)
    
def SesionesCoelctivas():
    # Páginas del archivo Excel cargado
    num_paginas = len(workbook.sheetnames)
    print(f"El archivo Excel tiene {num_paginas} páginas.")

    # Primero, validar la página 1
    if num_paginas >= 1 and workbook.sheetnames[0] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]  # Acceder a la página 1
        print("Validando la página 1...")
        validar_pagina1_sesiones(sheet)

    # Luego, validar la página 2 si existe
    if num_paginas >= 2 and workbook.sheetnames[1] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[1]]  # Acceder a la página 2
        print("Validando la página 2...")
        validar_pagina2_sesiones(sheet)
    
    # Luego, validar la página 3 si existe
    if num_paginas >= 3 and workbook.sheetnames[2] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[2]]  # Acceder a la página 2
        print("Validando la página 3...")
        validar_pagina3_sesiones(sheet)

def PrevencionEmbarazo(): 
    # Páginas del archivo Excel cargado
    num_paginas = len(workbook.sheetnames)
    print(f"El archivo Excel tiene {num_paginas} páginas.")
    # Primero, validar la página 1
    if num_paginas >= 1 and workbook.sheetnames[0] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]  # Acceder a la página 1
        print("Validando la página 1...")
        prevencionPag1(sheet) 
    
    if num_paginas >= 2 and workbook.sheetnames[1] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[1]]  # Acceder a la página 1
        print("Validando la página 2...")
        prevencionPag2(sheet)

def higieneManos():
    # Páginas del archivo Excel cargado
    num_paginas = len(workbook.sheetnames)
    print(f"El archivo Excel tiene {num_paginas} páginas.")
    # Primero, validar la página 1
    if num_paginas >= 1 and workbook.sheetnames[0] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]  # Acceder a la página 1
        print("Validando la página 1...")
        hm_pag1(sheet) 
        
    if num_paginas >= 2 and workbook.sheetnames[1] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[1]]  # Acceder a la página 1
        print("Validando la página 2...")
        hm_pag2(sheet) 
        
    if num_paginas >= 3 and workbook.sheetnames[2] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[2]]  # Acceder a la página 1
        print("Validando la página 3...")
        hm_pag3(sheet) 
        
def pretest():
    # Páginas del archivo Excel cargado
    num_paginas = len(workbook.sheetnames)
    print(f"El archivo Excel tiene {num_paginas} páginas.")
    # Primero, validar la página 1
    if num_paginas >= 1 and workbook.sheetnames[0] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]  # Acceder a la página 1
        print("Validando la página 1...")
        pretest_Pag1(sheet) 
        
    if num_paginas >= 1 and workbook.sheetnames[1] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[1]]  # Acceder a la página 1
        print("Validando la página 2...")
        pretest_Pag2(sheet) 
    
def jornadas():
    # Páginas del archivo Excel cargado
    num_paginas = len(workbook.sheetnames)
    print(f"El archivo Excel tiene {num_paginas} páginas.")
    # Primero, validar la página 1
    if num_paginas >= 1 and workbook.sheetnames[0] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]  # Acceder a la página 1
        print("Validando la página 1...")
        jornadas_Pag1(sheet) 
        
    if num_paginas >= 1 and workbook.sheetnames[1] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[1]]  # Acceder a la página 1
        print("Validando la página 2...")
        jornadas_Pag2(sheet) 

def autocuidado():
    # Páginas del archivo Excel cargado
    num_paginas = len(workbook.sheetnames)
    print(f"El archivo Excel tiene {num_paginas} páginas.")
    # Primero, validar la página 1
    if num_paginas >= 1 and workbook.sheetnames[0] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]  # Acceder a la página 1
        print("Validando la página 1...")
        Auto_Pag1(sheet) 
        
    if num_paginas >= 1 and workbook.sheetnames[1] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[1]]  # Acceder a la página 1
        print("Validando la página 2...")
        Auto_Pag2(sheet) 

#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#/////////////////////////////////sESIONES COLECTIVAS////////////////////////////////////////////////////////////////
#///////////////////////////////////////////////////////////////////////////////////////////////////////////////////  
def validar_pagina1_sesiones(sheet):
    # Mostrar ventana de progreso
    
    regex = re.compile("^[a-zA-ZÑñáéíóúÁÉÍÓÚ\s]+$")
    patternTel = re.compile(r'^\d{7}(\d{3})?$')
    
    try:
        remplazarComillas(sheet)
        ultima_fila = sheet.max_row
        celdas_pintadas_rojo = 0

        ventana, progress, label = mostrar_ventana_progreso(f"Validando pagina1_sesiones...{ultima_fila}", ultima_fila - 1)
        ventana.update()  # Refrescar la ventana principal
        
        for i in range(2, ultima_fila + 1):
            
            # Tipo institución
            if len(sheet.cell(i, 8).value.strip()) > 0 and len(sheet.cell(i, 9).value.strip()) > 0:
                celdas_pintadas_rojo += 1
                colum["column"] = {8, 9, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            # Nombre institución
            if not sheet.cell(row=i, column=10).value:
                celdas_pintadas_rojo += 1
                colum["column"] = {10, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            # Barrio
            if not sheet.cell(row=i, column=21).value or not regex.match(sheet.cell(row=i, column=21).value):
                celdas_pintadas_rojo += 1
                colum["column"] = {21, 2}
                colum["row"] = i
                pintar(colum, sheet)
            
            # verificar si es barrio priorizado
            if sheet.cell(i, 24).value == "SI" and not len(sheet.cell(i, 25).value.strip()) > 0:
                celdas_pintadas_rojo += 1
                colum["column"] = {24, 25, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            # Verifica la condición para el cuarto conjunto de celdas (teléfono)
            telefono = str(sheet.cell(i, 44).value)
            if not patternTel.match(telefono):
                celdas_pintadas_rojo += 1
                colum["column"] = {44, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if not sheet.cell(i, 10).value:
                celdas_pintadas_rojo += 1
                colum["column"] = {10, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            CeldasVacias["vacias"] = {19}
            celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
                         
            placas["placas"] = {26, 31, 35}
            celdas_pintadas_rojo += numeroDirecciones(sheet, placas)  # numeros de direcciones
            
            actualizar_barra_progreso(ventana, progress, i * 100 / ultima_fila)
            
            if progress['value'] == 100:
                ventana.after(100, lambda: ventana.destroy())  # Cerrar la ventana después de 100 ms
                break
            
        label.config(text=f"Total errores encontrados: {celdas_pintadas_rojo}. de {ultima_fila}")
        ventana.update()  
        
        print(f"Total errores encontrados {celdas_pintadas_rojo}. de {ultima_fila}")
        

    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")   
        
        
def validar_pagina2_sesiones(sheet):
    regex = re.compile("^[a-zA-ZÑñáéíóúÁÉÍÓÚ\s]+$")
    patternTel = re.compile(r'^\d{7}(\d{3})?$')
    
    try:
        remplazarComillas(sheet)
        ultima_fila = sheet.max_row
        celdas_pintadas_rojo = 0
        
        ventana, progress, label = mostrar_ventana_progreso("Validando pagina2_sesiones...", ultima_fila - 1)
        ventana.update()  # Refrescar la ventana principal

        # Obtener la fecha actual
        fechaActual = datetime.datetime.now()

        # Convertir la fecha actual a un formato de fecha (si solo necesitas la fecha)
        fechaActual = fechaActual.date()
        
        for i in range(2, ultima_fila + 1):
            # Tipo institución
            if not len(sheet.cell(i, 8).value.strip()) > 0 :
                celdas_pintadas_rojo += 1
                colum["column"] = {8, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            # ingresart mas validadores para la pagina dos
                
            if not sheet.cell(row=i, column=14).value or not regex.match(sheet.cell(row=i, column=14).value):
                celdas_pintadas_rojo += 1
                colum["column"] = {14, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if not sheet.cell(row=i, column=16).value or not regex.match(sheet.cell(row=i, column=16).value):
                celdas_pintadas_rojo += 1
                colum["column"] = {16, 2}
                colum["row"] = i
                pintar(colum, sheet)
            
            if not sheet.cell(row=i, column=18).value or not regex.match(sheet.cell(row=i, column=18).value):
                celdas_pintadas_rojo += 1
                colum["column"] = {18, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if sheet.cell(i, 13).value < sheet.cell(i, 3).value : 
                celdas_pintadas_rojo += 1
                colum["column"] = {13, 3, 2}
                colum["row"] = i
                pintar(colum, sheet)
            
             # Obtener el valor de la celda y convertirlo a datetime.date si es un string
            
            fecha_celda_texto = sheet.cell(i, 13).value

            # Convertir el texto de la celda a objeto datetime si es un string
            try:
                fecha_celda = datetime.datetime.strptime(fecha_celda_texto, "%Y/%m/%d").date()
            except ValueError:
                # Manejar el caso en que la conversión no sea posible
                #print("La fecha en la celda {} no se pudo convertir. Revisar el formato.".format((i, 13)))
                continue

            # Comparar la fecha en la celda con la fecha actual
            if fecha_celda > fechaActual:
                celdas_pintadas_rojo += 1
                colum["column"] = {13, 3, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            actualizar_barra_progreso(ventana, progress, i * 100 / ultima_fila)
            
            if progress['value'] == 90:
                ventana.after(10, lambda: ventana.destroy())  # Cerrar la ventana después de 10 ms
                break   
        
        label.config(text=f"Total errores encontrados: {celdas_pintadas_rojo}. de {ultima_fila}")
        ventana.update() 
           
         # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}. de {ultima_fila}")

    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")
        
        
def validar_pagina3_sesiones(sheet):
    regex = re.compile("^[a-zA-ZÑñáéíóúÁÉÍÓÚ\s]+$")
    adulto_menor_sin_id = re.compile(r'^\d{2,5}[A-Za-z]{2,5}\d{5,6}$')

    NumeroDocumento = re.compile("^\d{10}$")
    
    try:
        remplazarComillas(sheet)  
        ultima_fila = sheet.max_row
        celdas_pintadas_rojo = 0
        
        ventana, progress, label = mostrar_ventana_progreso("Validando sexo sesiones...", ultima_fila - 1)
        ventana.update()  # Refrescar la ventana principal
            
        for i in range(2, ultima_fila + 1):
            # Tipo institución
            if sheet.cell(i, 11).value == "2- Mujer" and sheet.cell(i, 12).value != "2- Femenino" :
                celdas_pintadas_rojo += 1
                colum["column"] = {11, 12, 2}
                colum["row"] = i
                pintar(colum, sheet)
            
            if sheet.cell(i, 11).value == "1- Hombre" and sheet.cell(i, 12).value != "1- Masculino" :
                celdas_pintadas_rojo += 1
                colum["column"] = {11, 12, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if sheet.cell(i, 11).value == "3- Intersexual" and sheet.cell(i, 12).value != "3- Transgénero" :
                celdas_pintadas_rojo += 1
                colum["column"] = {11, 12, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            FechaIntervencion = sheet.cell(i, 3).value
            FechaNacimiento = sheet.cell(i, 14).value
            FechaNacimiento = FechaNacimiento.replace('/', '-')  # Reemplazar '/' por '-'
            FechaNacimiento_format = FechaNacimiento.replace('`', '')  
            FechaIntervencion_format = FechaIntervencion.replace('`', '') 
            edad = calcular_edad(FechaNacimiento_format, FechaIntervencion_format)
            
            if edad >= 0 and edad <= 6 :
                tipodocumento = "2- RC"
                Nacionalidad = "COL"
                
            if edad >= 7 and edad <= 17 :
                tipodocumento = "3- TI"
                Nacionalidad = "COL"
                
            if edad >= 18:
                tipodocumento = "1- CC"
                Nacionalidad = "COL"
                
            if (sheet.cell(i, 10).value != tipodocumento and sheet.cell(i, 10).value != "8- Menor sin ID." and \
                sheet.cell(i, 10).value != "7- Adulto sin ID.") and sheet.cell(i, 16).value == Nacionalidad :
                celdas_pintadas_rojo += 1
                colum["column"] = {10, 16, 14, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
          
            if sheet.cell(i,10).value == tipodocumento and sheet.cell(i,16).value != Nacionalidad:
                celdas_pintadas_rojo += 1
                colum["column"] = {10, 16, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if edad > 100:
                celdas_pintadas_rojo += 1
                colum["column"] = {14, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            # estado civil
            if edad <= 13 and sheet.cell(i,13).value != "6- No aplica":
                celdas_pintadas_rojo += 1
                colum["column"] = {13, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            #verificar poblacion
            if edad <= 14 and sheet.cell(i,19).value != "Estudiante":
                celdas_pintadas_rojo += 1
                colum["column"] = {14, 19, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            #verificar poblacion
            if edad <= 17 and sheet.cell(i,19).value == "Docente":
                celdas_pintadas_rojo += 1
                colum["column"] = {14, 19, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            numeroDocumento = sheet.cell(i, 9).value
            
            # Verifica si el número de documento cumple con el patrón y satisface las condiciones adicionales
            if (not NumeroDocumento.match(numeroDocumento) and 
                sheet.cell(i, 10).value not in ["8- Menor sin ID.", "7- Adulto sin ID.", "13- PPT Permiso por Protección Temporal", "5- NUIP"] and 
                edad < 35):
                celdas_pintadas_rojo += 1
                colum["column"] = {9, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if len(numeroDocumento) < 5: 
                celdas_pintadas_rojo += 1
                colum["column"] = {9, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if sheet.cell(i, 10).value  in ["8- Menor sin ID.", "7- Adulto sin ID."] and not adulto_menor_sin_id.match(numeroDocumento) :
                celdas_pintadas_rojo += 1
                colum["column"] = {10, 9, 2}
                colum["row"] = i
                pintar(colum, sheet)
               
            if sheet.cell(i, 16).value == "COL" and \
                sheet.cell(i, 10).value not in ["2- RC", "3- TI", "1- CC", "8- Menor sin ID.", "7- Adulto sin ID."]:
                    celdas_pintadas_rojo += 1
                    colum["column"] = {16, 10, 2}
                    colum["row"] = i
                    pintar(colum, sheet)
                    
            if sheet.cell(i, 16).value != "COL" and \
                sheet.cell(i, 10).value in ["2- RC", "3- TI", "1- CC"]:
                    celdas_pintadas_rojo += 1
                    colum["column"] = {16, 10, 2}
                    colum["row"] = i
                    pintar(colum, sheet)
                    
            if not sheet.cell(row=i, column=8).value or not regex.match(sheet.cell(row=i, column=8).value):  
                celdas_pintadas_rojo += 1
                colum["column"] = {8, 2}
                colum["row"] = i
                pintar(colum, sheet)   
                
            actualizar_barra_progreso(ventana, progress, i * 100 / ultima_fila)
            
            if progress['value'] == 100:
                ventana.after(10, lambda: ventana.destroy())  # Cerrar la ventana después de 100 ms
                break 
            
        label.config(text=f"Total errores encontrados: {celdas_pintadas_rojo}. de {ultima_fila}")
        ventana.update()  
                   
        # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}. de {ultima_fila}")

    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}") 
        
        
#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#/////////////////////////////////PREVENCION EMBARAZO////////////////////////////////////////////////////////////////
#///////////////////////////////////////////////////////////////////////////////////////////////////////////////////  
def prevencionPag1(sheet):
    
    NumeroDocumento = re.compile("^\d{10}$")
    try:
        remplazarComillas(sheet)  
        ultima_fila = sheet.max_row
        celdas_pintadas_rojo = 0
        
        
        
        #validador de campos por la edad 
        for i in range(2, ultima_fila + 1):
            FechaIntervencion = sheet.cell(i, 3).value
            FechaNacimiento = sheet.cell(i, 18).value
            FechaNacimiento = FechaNacimiento.replace('/', '-')  # Reemplazar '/' por '-'
            FechaNacimiento_format = FechaNacimiento.replace('`', '')  
            FechaIntervencion_format = FechaIntervencion.replace('`', '')             
            edad = calcular_edad(FechaNacimiento_format, FechaIntervencion_format)
            
            if edad >= 0 and edad <= 6 :
                tipodocumento = "2- RC"
                Nacionalidad = "Colombia"
                
            if edad >= 7 and edad <= 17 :
                tipodocumento = "3- TI"
                Nacionalidad = "Colombia"
                
            if edad >= 18:
                tipodocumento = "1- CC"
                Nacionalidad = "Colombia"
                
            if (sheet.cell(i, 8).value != tipodocumento and sheet.cell(i, 8).value != "8- Menor sin ID." and \
                sheet.cell(i, 8).value != "7- Adulto sin ID.") and sheet.cell(i, 14).value == Nacionalidad :
                celdas_pintadas_rojo += 1
                colum["column"] = {8, 18, 14, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if sheet.cell(i,8).value == tipodocumento and sheet.cell(i,14).value != Nacionalidad:
                celdas_pintadas_rojo += 1
                colum["column"] = {8, 14, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if edad > 100:
                celdas_pintadas_rojo += 1
                colum["column"] = {18, 19, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            # estado civil
            if edad <= 13 and sheet.cell(i,17).value != "6- No aplica":
                celdas_pintadas_rojo += 1
                colum["column"] = {17,18, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            numeroDocumento = sheet.cell(i, 9).value
            # Verifica si el número de documento cumple con el patrón y satisface las condiciones adicionales
            if (not NumeroDocumento.match(numeroDocumento) and 
                sheet.cell(i, 8).value not in ["8- Menor sin ID.", "7- Adulto sin ID.", "13- PPT Permiso por Protección Temporal", "5- NUIP"] and 
                edad < 35):
                celdas_pintadas_rojo += 1
                colum["column"] = {9, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if len(numeroDocumento) < 5: 
                celdas_pintadas_rojo += 1
                colum["column"] = {9, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
        celTexto["ColumText"] = {10, 11, 12, 13, 59, 131, 133}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
        
        Genero["Genero"]= {15, 16}
        celdas_pintadas_rojo += validadorsexoGenero(sheet, Genero) #nueva vigencia ok
        
        etniaVal["etnia"]= {21, 22}
        celdas_pintadas_rojo += Valetnia(sheet, etniaVal)# nueva vigencia ok
        
        afiliacion["afiliacion"]= {24, 25}
        celdas_pintadas_rojo += Valiafiliacion(sheet, afiliacion) # nueva vigencia ok
        
        CeldasVacias["vacias"] = {41, 10 , 12, 9, 18,  24, 81, 86, 87, 88, 113, 116 ,118, 119, 120, 123, 124 , 126, 128, 130 , 131, 132, 133  }
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)#columnas requeridas lista nueva vigencia
        
        placas["placas"] = {63,72 }
        celdas_pintadas_rojo += numeroDirecciones(sheet, placas)#columnas requeridas
        
        
        #validar si Gestante 
        for i in range(2, ultima_fila + 1):
            if sheet.cell(i, 90).value == "SI" : # validacion si es gestante y campos
                if sheet.cell(i,92).value == " ":
                    celdas_pintadas_rojo += 1
                    colum["column"] = {90, 91, 2}
                    colum["row"] = i
                    pintar(colum, sheet)
                                    
                for col_num in {97, 98, 99, 100, 101, 121}:
                    cell_value = sheet.cell(row=i, column=col_num).value
                    if cell_value == " ":
                        celdas_pintadas_rojo += 1   # campos obligatorios si es gestante
                        colum["column"] = {col_num, 2}
                        colum["row"] = i
                        pintar(colum, sheet)
                
                for col_num in {102, 104, 105, 106, 107, 108, 109, 110, 111 }:
                    cell_value = sheet.cell(row=i, column=col_num).value
                    if cell_value != " ":
                        celdas_pintadas_rojo += 1   # verifica si campos que nos son requeridos en gestante tienen algun dato
                        colum["column"] = {col_num, 2}
                        colum["row"] = i
                        pintar(colum, sheet)
                
            elif sheet.cell(i, 93).value == "SI" :
                if sheet.cell(i,94).value == " ":
                    celdas_pintadas_rojo += 1
                    colum["column"] = {93, 94, 2}
                    colum["row"] = i
                    pintar(colum, sheet)
                                    
                for col_num in {98, 102, 104, 105, 106, 107, 108, 109, 110, 111, 121}:
                    cell_value = sheet.cell(row=i, column=col_num).value
                    if cell_value == " ":
                        celdas_pintadas_rojo += 1   # verificar datos de lacatntes requerido
                        colum["row"] = i
                        pintar(colum, sheet)
                
                for col_num in {96, 97, 99, 100, 101}:
                    cell_value = sheet.cell(row=i, column=col_num).value
                    if cell_value != " ":
                        celdas_pintadas_rojo += 1   # verificar si los campos no requeridos estan llenos diferentes a lactantye
                        colum["column"] = {col_num, 2}
                        colum["row"] = i
                        pintar(colum, sheet)
                        
            
        
                
       # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")

    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}") 
def prevencionPag2(sheet):
    try:
        remplazarComillas(sheet)  
        ultima_fila = sheet.max_row
        celdas_pintadas_rojo = 0

        # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")

    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")

#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#/////////////////////////////////HIGIENE DE MANOS///////////////////////////////////////////////////////////////////
#///////////////////////////////////////////////////////////////////////////////////////////////////////////////////    
def hm_pag1(sheet):
    try:
        remplazarComillas(sheet)  
        ultima_fila = sheet.max_row
        celdas_pintadas_rojo = 0
        #validar celdas vacias 
        CeldasVacias["vacias"] = {11 ,13, 21, 27, 33, 37}
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)#columnas requeridas
        #validacion de texto que no contenga caracteres especiales 
        celTexto["ColumText"] = {13, 21,}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
        
        # VALIDACION SI ES RURAL O URBANA
        for i in range(2, ultima_fila +1):
            if sheet.cell(i,16).value == "1- Urbana":
                #numeros de direccion
                placas["placas"] = {27, 33, 37}
                celdas_pintadas_rojo += numeroDirecciones(sheet, placas)#columnas requeridas
            else:
                rural["rural"] = {43, 45, 46}
                celdas_pintadas_rojo += Val_Rural(sheet, rural)#columnas requeridas
                
        # validar telefonos
        Tel["Tel"] = {47, 48}
        celdas_pintadas_rojo += ValidarTel(sheet, Tel)#columnas requeridas telefono
        # validar manzana del cuidado 
        Manzana["Manzana"] = {23, 24}
        celdas_pintadas_rojo += manzanaPriorizada(sheet, Manzana)#columnas requeridas telefono
        
        # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")

    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")
def hm_pag2(sheet):
    try:
        remplazarComillas(sheet) 
        
        celdas_pintadas_rojo = 0 
        #validar celdas vacias 
        CeldasVacias["vacias"] = {8, 10, 13, 17 }
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        #validacion de texto que no contenga caracteres especiales 
        celTexto["ColumText"] = {8, 9, 10, 11}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
        
        Var_edad = { # seleccionar la columna donde se encuentra cada campo 
            "F_Intervencion": 3,
            "F_nacimiento": 16,
            "T_Doc": 12,
            "Nac": 14,
            "No_doc": 13,
            "est_civil": 0, # colocar 0 si no tiene la columna de estado civil 
            "Nacionalidad": "Colombia",
            "vinculo_Jefe": 0, # colocar 0 si no existe el campo 
            "poblacion": 18# cambiar si es necesariio ya que puede solo aparecer "COl"
        }

        celdas_pintadas_rojo += Docuemento(sheet, Var_edad)
        
        # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")     
def hm_pag3(sheet):
    try:
        remplazarComillas(sheet) 
        celdas_pintadas_rojo = 0 
        
        # CeldasVacias["vacias"] = {8, 10}
        # celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        celTexto["ColumText"] = {12,13}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
        
        
         # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")

#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#//////////////////////Formato pre test y post test docentes jardines infantiles/////////////////////////////////////
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 

def pretest_Pag1(sheet):
    
    try:
        remplazarComillas(sheet) 
        celdas_pintadas_rojo = 0 
        ultima_fila = sheet.max_row
        CeldasVacias["vacias"] = {12, 13}
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        celTexto["ColumText"] = {12, 13}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
        
        NumeroDocumento = re.compile("^\d{8}$|^\d{10}$")
        for i in range(2, ultima_fila + 1):
            numeroDocumento = sheet.cell(i, 10).value
            # Verifica si el número de documento cumple con el patrón y satisface las condiciones adicionales
            if not NumeroDocumento.match(numeroDocumento) and sheet.cell(i, 9).value not in ["8- Menor sin ID.", "7- Adulto sin ID.", "13- PPT Permiso por Protección Temporal", "5- NUIP"]:
                celdas_pintadas_rojo += 1
                colum["column"] = {10, 2}
                colum["row"] = i
                pintar(colum, sheet)
        
        
         # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")
    
def pretest_Pag2(sheet):
    try:
        remplazarComillas(sheet) 
        celdas_pintadas_rojo = 0 
        ultima_fila = sheet.max_row
        CeldasVacias["vacias"] = {8, 9, 10, 11, 12, 13, 14, 15, 16, 17}
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        celTexto["ColumText"] = {19, 20}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
        
        NumeroDocumento = re.compile("^\d{8}$|^\d{10}$")
        for i in range(2, ultima_fila + 1):
            numeroDocumento = sheet.cell(i, 21).value
            # Verifica si el número de documento cumple con el patrón y satisface las condiciones adicionales
            if not NumeroDocumento.match(numeroDocumento):
                celdas_pintadas_rojo += 1
                colum["column"] = {21, 2}
                colum["row"] = i
                pintar(colum, sheet)
        
                
         # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")
 
 
#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#///////////////////////////////////////////////JORNADAS/////////////////////////////////////////////////////////////
#///////////////////////////////////////////////////////////////////////////////////////////////////////////////////  
def jornadas_Pag1(sheet):
    try:
        remplazarComillas(sheet) 
        patternTel = re.compile(r'^\d{7}(\d{3})?$')
        celdas_pintadas_rojo = 0 
        ultima_fila = sheet.max_row
        
        CeldasVacias["vacias"] = {8, 9, 11, 12, 14, 16, 17, 18, 1, 20, 21, 23, 30, 34, 43, 45}
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        
        celTexto["ColumText"] = {19}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
        
           # VALIDACION SI ES RURAL O URBANA
        for i in range(2, ultima_fila +1):
            if sheet.cell(i,14).value == "1- Urbana":
                #numeros de direccion
                placas["placas"] = {24, 30, 34}
                celdas_pintadas_rojo += numeroDirecciones(sheet, placas)#columnas requeridas
            else:
                rural["rural"] = {40, 41, 42}
                celdas_pintadas_rojo += Val_Rural(sheet, rural)#columnas requeridas
        
            # Verifica la condición para el cuarto conjunto de celdas (teléfono)
            telefono = str(sheet.cell(i, 43).value)
            if not patternTel.match(telefono):
                celdas_pintadas_rojo += 1
                colum["column"] = {43, 2}
                colum["row"] = i
                pintar(colum, sheet)
        
         # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")

def jornadas_Pag2(sheet):
    try:
        remplazarComillas(sheet) 
        celdas_pintadas_rojo = 0 
        ultima_fila = sheet.max_row
        CeldasVacias["vacias"] = {8, 9, 10, 14, 21, 22, 23, 25}
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        celTexto["ColumText"] = {14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
        
        for i in range(2, ultima_fila + 1):
            if sheet.cell(i, 8).value < sheet.cell(i,3).value: 
                celdas_pintadas_rojo += 1
                colum["column"] = {8, 3, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            cell_value = sheet.cell(row=i, column=12).value
            if isinstance(cell_value, str) and cell_value.isdigit():
                sheet.cell(row=i, column=12).value = float(cell_value)
                
            cell_value = sheet.cell(row=i, column=12).value
            if isinstance(cell_value, (int, float)) and cell_value < 20:
                celdas_pintadas_rojo += 1   # Luego, verifica si el valor convertido es mayor a 250 y aplica el formato de relleno rojo si es necesario
                colum["column"] = {12, 2}
                colum["row"] = i
                pintar(colum, sheet) 
        
        
         # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")



#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#///////////////////////////////////////////////AUTOCUIDADO//////////////////////////////////////////////////////////
#///////////////////////////////////////////////////////////////////////////////////////////////////////////////////  
def Auto_Pag1(sheet):
    try:
        remplazarComillas(sheet) 
        celdas_pintadas_rojo = 0 
        ultima_fila = sheet.max_row
        CeldasVacias["vacias"] = {8, 10, 13}
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        celTexto["ColumText"] = {8, 9, 10, 11}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
        
        NumeroDocumento = re.compile("^\d{8}$|^\d{10}$")
        for i in range(2, ultima_fila + 1):
            numeroDocumento = sheet.cell(i, 13).value
            # Verifica si el número de documento cumple con el patrón y satisface las condiciones adicionales
            if not NumeroDocumento.match(numeroDocumento) and sheet.cell(i, 12).value not in ["8- Menor sin ID.", "7- Adulto sin ID.", "13- PPT Permiso por Protección Temporal", "5- NUIP"]:
                celdas_pintadas_rojo += 1
                colum["column"] = {13, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
         # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")
def Auto_Pag2(sheet):
    try:
        remplazarComillas(sheet) 
        celdas_pintadas_rojo = 0 
        ultima_fila = sheet.max_row
        CeldasVacias["vacias"] = {8, 9, 11, 12, 13, 15, 16, 17, 18, 20, 21, 23, 24, 26, 27, 28, 29}
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        
        
        NumeroDocumento = re.compile("^\d{8}$|^\d{10}$")
        adulto_menor_sin_id = re.compile(r'^\d{2,4}[A-Za-z]{2,5}\d{5,6}$')
        for i in range(2, ultima_fila + 1):
            numeroDocumento = sheet.cell(i, 8).value
            # Verifica si el número de documento cumple con el patrón y satisface las condiciones adicionales
            if not NumeroDocumento.match(numeroDocumento):
                if not adulto_menor_sin_id.match(numeroDocumento):
                    celdas_pintadas_rojo += 1
                    colum["column"] = {8, 2}
                    colum["row"] = i
                    pintar(colum, sheet)
                
            if sheet.cell(i,9).value < sheet.cell(i,3).value or sheet.cell(i,9).value > sheet.cell(i,3).value :
                celdas_pintadas_rojo += 1
                colum["column"] = {9, 3, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
                
         # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")
   
#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#///////////////////////////////////////////MASCOTA VERDE Y YO//////////////////////////////////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////
def mascota_pag1(sheet):
    try:
        remplazarComillas(sheet) 
        celdas_pintadas_rojo = 0 
        ultima_fila = sheet.max_row
        CeldasVacias["vacias"] = {8, 9, 10, 11}
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        
        celTexto["ColumText"] = {8, 11}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
        
        adulto_menor_sin_id = re.compile(r'^\d{2,4}[A-Za-z]{2,5}\d{5,6}$')
        
        NumeroDocumento = re.compile("^\d{8}$|^\d{10}$")
        for i in range(2, ultima_fila + 1):
            numeroDocumento = sheet.cell(i, 10).value
            # Verifica si el número de documento cumple con el patrón y satisface las condiciones adicionales
            if not NumeroDocumento.match(numeroDocumento) and sheet.cell(i, 9).value not in ["8- Menor sin ID.", "7- Adulto sin ID.", "13- PPT Permiso por Protección Temporal", "5- NUIP"]:
                celdas_pintadas_rojo += 1
                colum["column"] = {10, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if sheet.cell(i, 9).value  in ["8- Menor sin ID.", "7- Adulto sin ID."] and not adulto_menor_sin_id.match(numeroDocumento) :
                celdas_pintadas_rojo += 1
                colum["column"] = {10, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
                
                
         # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")

def mascota_pag2(sheet):
    
    try:
        remplazarComillas(sheet) 
        celdas_pintadas_rojo = 0 
        ultima_fila = sheet.max_row
        CeldasVacias["vacias"] = {9, 10, 11, 12,  15, 16}# obligatorios
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        
        for i in range(2, ultima_fila + 1):
            if sheet.cell(i,9).value < sheet.cell(i,3).value:
                celdas_pintadas_rojo += 1
                colum["column"] = {9, 3,2}
                colum["row"] = i
                pintar(colum, sheet)
                
            
        for i in range(2, ultima_fila + 1):
            if sheet.cell(i, 18).value != " " :# valida si la fecha de la sesion 2 esta llena       
                CeldasVacias_Condicional["vacias"] = {23, 24, 27, 28} # obligatorios por condicion 
                CeldasVacias_Condicional["row"] = i
                celdas_pintadas_rojo += vaciasXcondicion(sheet, CeldasVacias_Condicional)
                
                if sheet.cell(i,18).value < sheet.cell(i,9).value:
                    celdas_pintadas_rojo += 1
                    colum["column"] = {18, 9, 2}
                    colum["row"] = i
                    pintar(colum, sheet)
                
                
        for i in range(2, ultima_fila + 1):
            if sheet.cell(i, 30).value != " " :# valida si la fecha de la sesion 2 esta llena       
                CeldasVacias_Condicional["vacias"] = {34, 35, 38, 39} # obligatorios por condicion 
                CeldasVacias_Condicional["row"] = i
                celdas_pintadas_rojo += vaciasXcondicion(sheet, CeldasVacias_Condicional)
                
        
        celTexto["ColumText"] = {8, 11, 24, 27, 28, 35 ,38, 39}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
   
         # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")

#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#////////////////////////////////////////////////SALUD MENTAL////////////////////////////////////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////

def saludmental_pag1(sheet):
    try:
        remplazarComillas(sheet) 
        patternTel = re.compile(r'^\d{7}(\d{3})?$')
        celdas_pintadas_rojo = 0 
        ultima_fila = sheet.max_row
        CeldasVacias["vacias"] = {9, 10, 11, 13, 14, 15, 18, 19, 22, 24, 60 ,51 }
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        
        celTexto["ColumText"] = {22, 23, 24, 25, 136, 138}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
        
        Genero["Genero"]= {30, 31}
        celdas_pintadas_rojo += validadorsexoGenero(sheet, Genero)
        
        for i in range(2, ultima_fila + 1):
            if sheet.cell(i,10).value == "Universidades" or sheet.cell(i,10).value == "Jardines" :
                if sheet.cell(i,17).value != " " :
                    colum["column"] = {10, 17, 2}
                    colum["row"] = i
                    pintar(colum, sheet)
            else:
                if sheet.cell(i,17).value == " " :
                    colum["column"] = {17, 2}
                    colum["row"] = i
                    pintar(colum, sheet)
                
                
                    
        for i in range(2, ultima_fila + 1):
            if sheet.cell(i, 42).value == "5- No asegurado" and not "no asegurado" in str(sheet.cell(i, 43).value).lower():
                celdas_pintadas_rojo += 1
                colum["column"] = {42, 43, 2}
                colum["row"] = i
                pintar(colum, sheet)
            else:
                cantidad = len(sheet.cell(i,43).value)    
                if cantidad < 3 : 
                    celdas_pintadas_rojo += 1
                    colum["column"] = {43, 2}
                    colum["row"] = i
                    pintar(colum, sheet)
                    
                
        Var_edad = { # seleccionar la columna donde se encuentra cada campo 
            "F_Intervencion": 3,
            "F_nacimiento": 37,
            "T_Doc": 27,
            "Nac": 29,
            "No_doc": 28,
            "est_civil": 0, # colocar 0 si no tiene la columna de estado civil 
            "Nacionalidad": "50", # cambiar si es necesariio ya que puede solo aparecer "COl"
            "vinculo_Jefe": 0,
            "vinculo_Jefe": 0, # colocar 0 si no existe el campo 
            "poblacion": 0
        }
        
        celdas_pintadas_rojo += Docuemento(sheet, Var_edad)
        
        # VALIDACION SI ES RURAL O URBANA
        for i in range(2, ultima_fila +1):
            if sheet.cell(i,47).value == "1- Urbana":
                #numeros de direccion
                placas["placas"] = {61, 67, 71}
                celdas_pintadas_rojo += numeroDirecciones(sheet, placas)#columnas requeridas
            else:
                rural["rural"] = {43, 45, 46}
                celdas_pintadas_rojo += Val_Rural(sheet, rural)#columnas requeridas
                
            if sheet.cell(i, 101).value < sheet.cell(i, 3).value:
                celdas_pintadas_rojo += 1
                colum["column"] = {101, 3, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if sheet.cell(i, 104).value != " ":
                if sheet.cell(i, 104).value < sheet.cell(i, 101).value:
                    celdas_pintadas_rojo += 1
                    colum["column"] = {101, 3, 2}
                    colum["row"] = i
                    pintar(colum, sheet)
            
            if sheet.cell(i, 106).value != " ":
                if sheet.cell(i, 106).value < sheet.cell(i, 104).value:
                    celdas_pintadas_rojo += 1
                    colum["column"] = {101, 3, 2}
                    colum["row"] = i
                    pintar(colum, sheet)
                

       # Verifica la condición para el cuarto conjunto de celdas (teléfono)
            telefono = str(sheet.cell(i, 84).value)
            if not patternTel.match(telefono) or telefono == " " :
                celdas_pintadas_rojo += 1
                colum["column"] = {84, 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            telefono2 = str(sheet.cell(i, 85).value)
            if not patternTel.match(telefono2) and telefono2 != " ":
                celdas_pintadas_rojo += 1
                colum["column"] = {85, 2}
                colum["row"] = i
                pintar(colum, sheet)
         
        CeldasVacias["vacias"] = {88, 89, 90, 91, 92, 93, 94, 95, 96, 98, 110, 112, 115 }
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
              
         # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
        
        # VALIDAR CURSO 
        for i in range(2, ultima_fila +1):
            if sheet.cell(i,10).value == "Colegios" and sheet.cell(i,10).value == "No aplica":
                celdas_pintadas_rojo += 1
                colum["column"] = {85, 2}
                colum["row"] = i
                pintar(colum, sheet)
        
    except Exception as e:
        print("Error", f"Se produjo un error en salud mental en la parte : {str(e)}")
    
#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#//////////////////////////////////////////////HIGIENE BUCAL////////////////////////////////////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////////////////////
def hb_pag1(sheet):
    try:
        remplazarComillas(sheet) 
        patternTel = re.compile(r'^\d{7}(\d{3})?$')
        
        celdas_pintadas_rojo = 0 
        ultima_fila = sheet.max_row
        CeldasVacias["vacias"] = {10, 11, 12, 13, 14, 29,23 }
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        celTexto["ColumText"] = {10, 11, 23}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
       
       
         #VALIDACION SI ES RURAL O URBANA
        for i in range(2, ultima_fila +1):
            if sheet.cell(i,17).value == "1- Urbana":
                #numeros de direccion
                placas["placas"] = {30, 36, 40}
                celdas_pintadas_rojo += numeroDirecciones(sheet, placas)#columnas requeridas
            else:
                rural["rural"] = {48, 49, 50}
                celdas_pintadas_rojo += Val_Rural(sheet, rural)#columnas requeridas
        
        # Verifica la condición para el cuarto conjunto de celdas (teléfono)
            telefono = str(sheet.cell(i, 51).value)
            if not patternTel.match(telefono) or telefono == " " :
                celdas_pintadas_rojo += 1
                colum["column"] = {51, 2}
                colum["row"] = i
                pintar(colum, sheet)
         # Mostrar la cantidad de celdas pintadas de rojo
        
        Manzana["Manzana"] = {26, 27}
        celdas_pintadas_rojo += manzanaPriorizada(sheet, Manzana)
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")
    
    
def hb_pag2(sheet):
    
    try:
        remplazarComillas(sheet) 
        celdas_pintadas_rojo = 0 
        ultima_fila = sheet.max_row
        CeldasVacias["vacias"] = {9, 11, 14, 15, 41}
        celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)
        celTexto["ColumText"] = {9, 10, 11, 12}      
        celdas_pintadas_rojo += validarCeldasTexto(sheet, celTexto)
        
        Var_edad = { # seleccionar la columna donde se encuentra cada campo 
            "F_Intervencion": 3,
            "F_nacimiento": 19,
            "T_Doc": 14,
            "Nac": 16,
            "No_doc": 15,
            "est_civil": 0, # colocar 0 si no tiene la columna de estado civil 
            "Nacionalidad": "50", # cambiar si es necesariio ya que puede solo aparecer "COl"
            "vinculo_Jefe": 25, # colocar 0 si no existe el campo 
            "poblacion": 24
        }
        
        celdas_pintadas_rojo += Docuemento(sheet, Var_edad)
        
        Genero["Genero"]= {17, 18}
        celdas_pintadas_rojo += validadorsexoGenero(sheet, Genero)
        
        afiliacion["afiliacion"]= {27, 28}
        celdas_pintadas_rojo += Valiafiliacion(sheet, afiliacion)
        
        
         # Mostrar la cantidad de celdas pintadas de rojo
        print(f"Total errores encontrados {celdas_pintadas_rojo}.")
        
    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")
    
    

#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#//////////////////////////////////////////FUNCIONES A UTILIZAR//////////////////////////////////////////////////////
#///////////////////////////////////////////////////////////////////////////////////////////////////////////////////

def Docuemento(sheet, Var_edad ):
    
    
    ultima_fila = sheet.max_row
    celdas_pintadas_rojo = 0 
    NumeroDocumento = re.compile("^\d{10}$")
        #validador de campos por la edad 
    adulto_menor_sin_id = re.compile(r'^\d{2,4}[A-Za-z]{2,5}\d{5,6}$')
    
    for i in range(2, ultima_fila + 1):
        FechaIntervencion = sheet.cell(i, Var_edad["F_Intervencion"]).value
        
        
        FechaNacimiento = sheet.cell(i, Var_edad["F_nacimiento"]).value
       
        FechaNacimiento = FechaNacimiento.replace('/', '-')  # Reemplazar '/' por '-'
        FechaNacimiento_format = FechaNacimiento.replace('`', '')  
        FechaIntervencion_format = FechaIntervencion.replace('`', '')             
        edad = calcular_edad(FechaNacimiento_format, FechaIntervencion_format)
        
        if edad >= 0 and edad <= 6 :
            tipodocumento = "2- RC"
            Nacionalidad = Var_edad["Nacionalidad"]
            
        if edad >= 7 and edad <= 17 :
            tipodocumento = "3- TI"
            Nacionalidad = Var_edad["Nacionalidad"]
            
        if edad >= 18:
            tipodocumento = "1- CC"
            Nacionalidad = Var_edad["Nacionalidad"]
        
        if (sheet.cell(i, Var_edad["T_Doc"]).value != tipodocumento and sheet.cell(i, Var_edad["T_Doc"]).value != "8- Menor sin ID." and \
            sheet.cell(i, Var_edad["T_Doc"]).value != "7- Adulto sin ID.") and sheet.cell(i, Var_edad["Nac"]).value == Nacionalidad :
            celdas_pintadas_rojo += 1
            colum["column"] = {Var_edad["T_Doc"], Var_edad["F_nacimiento"], 2}
            colum["row"] = i
            pintar(colum, sheet)
            
        if sheet.cell(i,Var_edad["T_Doc"]).value == tipodocumento and sheet.cell(i, Var_edad["Nac"]).value != Nacionalidad:
            celdas_pintadas_rojo += 1
            colum["column"] = {Var_edad["T_Doc"], Var_edad["Nac"], 2}
            colum["row"] = i
            pintar(colum, sheet)
            
        if edad > 100:
            celdas_pintadas_rojo += 1
            colum["column"] = {Var_edad["T_Doc"], 2}
            colum["row"] = i
            pintar(colum, sheet)
            
        # estado civil
        if Var_edad["est_civil"] != 0:
            if edad <= 13 and sheet.cell(i,Var_edad["est_civil"]).value != "6- No aplica":
                celdas_pintadas_rojo += 1
                colum["column"] = {Var_edad["est_civil"], 2}
                colum["row"] = i
                pintar(colum, sheet)
                
        if Var_edad["vinculo_Jefe"] != 0:
            if sheet.cell(i,Var_edad["vinculo_Jefe"]).value != "3- Hijo(a)" and edad < 15:
                celdas_pintadas_rojo += 1
                colum["column"] = {Var_edad["vinculo_Jefe"], 2}
                colum["row"] = i
                pintar(colum, sheet)
                
        if Var_edad["poblacion"] != 0:
            if  sheet.cell(i, Var_edad["Nac"]).value != Nacionalidad and sheet.cell(i,Var_edad["poblacion"]).value != "13- Migrante" :
                celdas_pintadas_rojo += 1
                colum["column"] = {Var_edad["poblacion"], Var_edad["Nac"], 2}
                colum["row"] = i
                pintar(colum, sheet)
            
            
        numeroDocumento = sheet.cell(i, Var_edad["No_doc"]).value
        # Verifica si el número de documento cumple con el patrón y satisface las condiciones adicionales
        if (not NumeroDocumento.match(numeroDocumento) and 
            sheet.cell(i, Var_edad["T_Doc"]).value not in ["8- Menor sin ID.", "7- Adulto sin ID.", "13- PPT Permiso por Protección Temporal", "5- NUIP", "11- Permiso Especial de permanencia"] and 
            edad < 35):
            celdas_pintadas_rojo += 1
            colum["column"] = {Var_edad["No_doc"], 2}
            colum["row"] = i
            pintar(colum, sheet)
            
        if len(numeroDocumento) < 5: 
            celdas_pintadas_rojo += 1
            colum["column"] = {Var_edad["No_doc"], 2}
            colum["row"] = i
            pintar(colum, sheet)
            
        if sheet.cell(i, Var_edad["T_Doc"]).value in ["8- Menor sin ID.", "7- Adulto sin ID."] and not adulto_menor_sin_id.match(numeroDocumento) :
                celdas_pintadas_rojo += 1
                colum["column"] = {10, 2}
                colum["row"] = i
                pintar(colum, sheet)
            
            
    return celdas_pintadas_rojo     

def Val_Rural(sheet, rural):
    celdas_pintadas_rojo = 0
    ultima_fila = sheet.max_row
    columns = list(rural["rural"])
    print(columns)
    CeldasVacias["vacias"] = {columns}
    celdas_pintadas_rojo += validarVacias(sheet, CeldasVacias)#columnas requeridas
    
    #/////////////////////////// ingresar si las cordenadas son correctas 
     
    return celdas_pintadas_rojo
    
def manzanaPriorizada(sheet, Manzana):
    celdas_pintadas_rojo = 0
    ultima_fila = sheet.max_row
    columns = list(Manzana["Manzana"])
    print(columns)
    for i in range(2, ultima_fila + 1):
        cell_value_0 = sheet.cell(i, columns[0]).value
        
        cell_value_1 = sheet.cell(i, columns[1]).value
        
        # Verificar las condiciones combinadas
        if (cell_value_1 == "Si" and cell_value_0 == " ") or (cell_value_1 == "No" and cell_value_0 != " "):
            celdas_pintadas_rojo += 1
            colum["column"] = {columns[0], columns[1], 2}
            colum["row"] = i
            pintar(colum, sheet)
    
    return celdas_pintadas_rojo
    
def ValidarTel(sheet, Tel):
    patternTel = re.compile(r'^\d{7}(\d{3})?$')
    celdas_pintadas_rojo = 0
    ultima_fila = sheet.max_row
    Num_celTexto = len(Tel["Tel"])
    columns = list(Tel["Tel"])
    
    for a in range(Num_celTexto):
        for i in range(2, ultima_fila + 1):
            # Verifica la condición para el cuarto conjunto de celdas (teléfono)
            if a == 0:
                telefono = str(sheet.cell(i, columns[1]).value)
                if  sheet.cell(i, columns[1]).value == " " or not patternTel.match(telefono) or (telefono and (telefono.isdigit() and int(telefono) == 0 or all(char == '0' for char in telefono))):
                    celdas_pintadas_rojo += 1
                    colum["column"] = {columns[1], 2}
                    colum["row"] = i
                    pintar(colum, sheet)
            
            if a == 1 :
                telefono = str(sheet.cell(i, columns[0]).value)
                if sheet.cell(i, columns[0]).value != " " and not patternTel.match(telefono) or (telefono and (telefono.isdigit() and int(telefono) == 0 or all(char == '0' for char in telefono))):
                    celdas_pintadas_rojo += 1
                    colum["column"] = {columns[0], 2}
                    colum["row"] = i
                    pintar(colum, sheet)
        
    return celdas_pintadas_rojo
                
def numeroDirecciones(sheet, placas):
    celdas_pintadas_rojo = 0
    ultima_fila = sheet.max_row
    
    columns = list(placas["placas"])
    
    for i in range(2, ultima_fila + 1):
        # Verifica cada columna en el conjunto de columnas especificadas
        for col_num in columns:
            cell_value = sheet.cell(row=i, column=col_num).value
            if isinstance(cell_value, str) and cell_value.isdigit():
                sheet.cell(row=i, column=col_num).value = float(cell_value)
                
        for col_num in columns:
            cell_value = sheet.cell(row=i, column=col_num).value
            if isinstance(cell_value, (int, float)) and cell_value > 250:
                celdas_pintadas_rojo += 1   # Luego, verifica si el valor convertido es mayor a 250 y aplica el formato de relleno rojo si es necesario
                colum["column"] = {col_num, 2}
                colum["row"] = i
                pintar(colum, sheet) 
            
    return celdas_pintadas_rojo     

def vaciasXcondicion(sheet, vacias):
    celdas_pintadas_rojo = 0
    number_colum = len(vacias["vacias"])
    columnData = list(vacias["vacias"])
    row = (vacias["row"])
    
    for a in range(number_colum):
       
        if sheet.cell(row, columnData[a]).value == " ":
            celdas_pintadas_rojo += 1
            colum["column"] = {columnData[a], 2}
            colum["row"] = row
            pintar(colum, sheet)
    
    return  celdas_pintadas_rojo
    
    
def validarVacias(sheet, CeldasVacias):
    celdas_pintadas_rojo = 0
    ultima_fila = sheet.max_row
    Num_celTexto = len(CeldasVacias["vacias"])
    columns = list(CeldasVacias["vacias"])
    for a in range(Num_celTexto):
        for i in range(2, ultima_fila + 1):
            if sheet.cell(row=i, column=columns[a]).value == " " :  
                    celdas_pintadas_rojo += 1
                    colum["column"] = {columns[a], 2}
                    colum["row"] = i
                    pintar(colum, sheet) 
    return  celdas_pintadas_rojo
    
def Valiafiliacion(sheet, afiliacion):
    celdas_pintadas_rojo = 0
    ultima_fila = sheet.max_row
    columns = list(afiliacion["afiliacion"])
    print(columns)
    for i in range(2, ultima_fila + 1):
            # Tipo institución
            if sheet.cell(i, columns[0]).value == "5- No asegurado" and not "no asegurado" in str(sheet.cell(i, columns[1]).value).lower():
                celdas_pintadas_rojo += 1
                colum["column"] = {columns[0], columns[1], 2}
                colum["row"] = i
                pintar(colum, sheet)
           
    return  celdas_pintadas_rojo 
       
def Valetnia(sheet, etnia):
    celdas_pintadas_rojo = 0
    ultima_fila = sheet.max_row
    columns = list(etnia["etnia"])
    for i in range(2, ultima_fila + 1):
            # Tipo institución
            if sheet.cell(i, columns[0]).value != "6- Ninguno" and sheet.cell(i, columns[1]).value == "-1" :
                celdas_pintadas_rojo += 1
                colum["column"] = {columns[0], columns[1], 2}
                colum["row"] = i
                pintar(colum, sheet)
            elif sheet.cell(i, columns[0]).value == "6- Ninguno" and sheet.cell(i, columns[1]).value != "-1" :
                celdas_pintadas_rojo += 1
                colum["column"] = {columns[0], columns[1], 2}
                colum["row"] = i
                pintar(colum, sheet)
    
    return  celdas_pintadas_rojo  
    
def validadorsexoGenero(sheet, Genero):
    celdas_pintadas_rojo = 0
    ultima_fila = sheet.max_row
    columns = list(Genero["Genero"])
        
    for i in range(2, ultima_fila + 1):
            # Tipo institución
            if sheet.cell(i, columns[-1]).value == "2- Mujer" and sheet.cell(i, columns[0]).value != "2- Femenino" :
                celdas_pintadas_rojo += 1
                colum["column"] = {columns[-1], columns[0], 2}
                colum["row"] = i
                pintar(colum, sheet)
            
            if sheet.cell(i, columns[-1]).value == "1- Hombre" and sheet.cell(i, columns[0]).value != "1- Masculino" :
                celdas_pintadas_rojo += 1
                colum["column"] = {columns[0], columns[1], 2}
                colum["row"] = i
                pintar(colum, sheet)
                
            if sheet.cell(i, columns[-1]).value == "3- Intersexual" and sheet.cell(i, columns[0]).value != "3- Transgénero" :
                celdas_pintadas_rojo += 1
                colum["column"] = {columns[-1], columns[0], 2}
                colum["row"] = i
                pintar(colum, sheet)
                
    return  celdas_pintadas_rojo      
                             
def validarCeldasTexto(sheet, celTexto):
    celdas_pintadas_rojo = 0
    regex = re.compile("^[a-zA-ZÑñáéíóúÁÉÍÓÚ\s]+$")
    ultima_fila = sheet.max_row
    Num_celTexto = len(celTexto["ColumText"])
    columns = list(celTexto["ColumText"])
    for a in range(Num_celTexto):
        for i in range(2, ultima_fila + 1):
            if sheet.cell(row=i, column=columns[a]).value and not regex.match(sheet.cell(row=i, column=columns[a]).value):  
                    celdas_pintadas_rojo += 1
                    colum["column"] = {columns[a], 2}
                    colum["row"] = i
                    pintar(colum, sheet) 
    return  celdas_pintadas_rojo

# funcio para remplazar comillas
def remplazarComillas(sheet):
    try :
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and '`' in cell.value:
                    # Elimina las comillas
                    cell.value = cell.value.replace('`', '')

                    # Verifica si el valor es un número
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = numbers.FORMAT_NUMBER
                    # Verifica si el valor es una fecha
                    elif isinstance(cell.value, datetime.date):
                        cell.number_format = numbers.FORMAT_DATE_XLSX15
                    # Verifica si el valor es texto
                    else:
                        cell.number_format = numbers.FORMAT_TEXT      
    except Exception as e:
        print("Error", f"Se produjo un error de comillas: {str(e)}")
        
 
              
# Función para calcular la edad
def calcular_edad(fecha_nacimiento, fecha_intervencion):
    try :
        nacimiento = datetime.datetime.strptime(fecha_nacimiento, "%Y-%m-%d")
        intervencion = datetime.datetime.strptime(fecha_intervencion, "%Y-%m-%d")
        edad = intervencion.year - nacimiento.year - ((intervencion.month, intervencion.day) < (nacimiento.month, nacimiento.day))
        return edad
    except :
        edad = -50    
        return edad

#funcion para pintar celdas 
def pintar(colum, sheet):
    colorRed = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    number_colum = len(colum["column"])
    columns = list(colum["column"])
    for i in range(number_colum):
        sheet.cell(row=colum["row"], column=columns[i]).fill = colorRed
  
def saveFile():
    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        file_path_modificado = file_path.replace('.xlsx', '_errores.xlsx')
        
        # Guardar el libro de trabajo original con los cambios realizados
        #workbook.save(file_path)
        # Guardar el archivo modificado con el nombre específico para errores
        workbook.save(file_path_modificado)
        print("Archivo guardado", "El archivo ha sido guardado correctamente.")
        # Preguntar al usuario si desea abrir el archivo guardado
        open_file = messagebox.askquestion("Abrir Archivo", "¿Desea abrir el archivo guardado?")
        if open_file == 'yes':
            os.startfile(file_path_modificado)  # Abre el archivo guardado
    except Exception as e:
        print("Error", f"No se pudo guardar el archivo: {str(e)}")

def preguntaDescarga():
    try:
        respuesta = messagebox.askquestion("Abrir Archivo", "¿Guardar el archivo generado?")
        if respuesta == "yes":
            cadenaGuardar = "Guardando archivo..."
            print(cadenaGuardar)
            saveFile()
        else:
            print("Tu archivo no será descargado")
    except Exception as e:
        print(f"No se pudo guardar el archivo: {str(e)}")
        
