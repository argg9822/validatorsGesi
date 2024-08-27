import datetime
from datetime import datetime
import math
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook, Workbook
import pandas as pd
from tkinter import filedialog, messagebox
import os
import re

##------------------------------------------------------------------------------------    
##----------------------------LECTURA DEL ARCHIVO-------------------------------------
##------------------------------------------------------------------------------------

def setBase(base):
    print(f"Validar >>>{base.upper()}<<<")
    loadFile()
    chooseBase(base)
    saveFile(base)

def loadFile():
    global workbook    

    print("Cargando y validando archivo...")

    fileRoute = filedialog.askopenfilename(
        title="Selecciona un archivo de Excel",
        filetypes=[("Archivos de Excel", "*.xlsx *.xls")]
    )

    if fileRoute:
        if fileRoute:
            workbook = load_workbook(fileRoute)
        else:
            print("El archivo no se cargó")

##------------------------------------------------------------------------------------    
##---------------------------SWICTH PARA LAS BASES------------------------------------
##------------------------------------------------------------------------------------
def chooseBase(base):
    switch = {
        "sesiones_colectivas": sc,
        "cuidarte": cuidarte,
        "mujeres": mujeres
    }
    execute_validator = switch.get(base)
    execute_validator()

##------------------------------------------------------------------------------------
##------------------------------GENERAL FUNCTIONS-------------------------------------
##------------------------------------------------------------------------------------
bgError = 'FFFF0000'
bgSecError = '005FFF'

def set_bg_error(index, columnName, color):    
    bgError = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell = sheet.cell(row=index+3, column=df.columns.get_loc(columnName)+1)
    cell.fill = bgError
    
    bgErrorFicha = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    cellFicha = sheet.cell(row=index+3, column=df.columns.get_loc('Ficha_fic')+1)
    cellFicha.fill = bgErrorFicha
    
    bgUsr = PatternFill(start_color='00ba4a', end_color='00ba4a', fill_type='solid')
    cellUsr = sheet.cell(row=index+3, column=df.columns.get_loc('Red_fic')+1)
    cellUsr.fill = bgUsr
    cellUsr.font = Font(bold=True)

def list_pages(df):
    if 'Ficha_fic' in df.columns:
        for num in df['Ficha_fic'].unique():
            indices = df.index[df['Ficha_fic'] == num].tolist()
            
            for i, idx in enumerate(indices):
                cell = sheet.cell(row=idx+3, column=df.columns.get_loc('Red_fic')+1)
                cell.value = i + 1
    else:
        print("No se encontró la columna Ficha_fic")
    return df

def clean_dataframe(df):
    # Función para limpiar cada celda del DataFrame
    def clean_cell(cell):
        if isinstance(cell, str):
            # Si es una cadena, elimina los caracteres `
            return cell.strip('`')
        elif pd.notnull(cell):
            # Si no es una cadena pero no es nulo, convierte a cadena y elimina los caracteres `
            return str(cell).strip('`')
        else:
            # Si es nulo, lo devuelve tal cual
            return cell

    # Aplicar la limpieza a todo el DataFrame
    cleaned_df = df.applymap(clean_cell)
    
    # Limpiar los nombres de las columnas
    cleaned_df.columns = [clean_cell(col) for col in df.columns]

    return cleaned_df

def required_fields(columnsNames = [], nextColumn = 0):    
    totalEmptyFields = 0
    for columnName in columnsNames:
        if columnName not in df.columns:
            raise ValueError(f'La columna {columnName} no se encuentra')
        
        if nextColumn == 1:
            currentIndex = df.columns.get_loc(columnName)

            if currentIndex + 1 >= len(df.columns):
                raise ValueError(f'No hay una columna después de {columnName}')
            
            #Obtener el nombre de la siguiente columna
            columnName = df.columns[currentIndex + 1]
        
        for index, fila in df.iterrows():
            cellField = fila[columnName].str.strip()

            if pd.isna(cellField):
                totalEmptyFields += 1
                set_bg_error(index, columnName, bgError)

        df['columnName'] = df[columnName].str.strip()
        empty_fields = df[pd.isna(df[columnName])]
        totalEmptyFields += len(empty_fields)

        for index in empty_fields.index:
            set_bg_error(index, columnName, bgError)

    if totalEmptyFields > 0:
        print(f"Campos obligatorios vacíos: {totalEmptyFields}")

    return totalEmptyFields

def validate_only_text(*columnsName):
    totalErrors = 0
    pattern =re.compile(r'^[^0-9.,:]+$')
    for columnName in columnsName:
        if columnName not in df.columns:
            raise ValueError(f'La columna {columnName} no se encuentra')
        
        fill_rows = df[pd.notna(df[columnName])]
        errors = fill_rows[~fill_rows[columnName].apply(lambda x: bool(pattern.search(str(x))))]
        
        totalErrors += len(errors)
        for index in errors.index:
            set_bg_error(index, columnName, bgError)
        
    if totalErrors > 0:
        print(f'Texto mal escrito: {totalErrors}')

    return totalErrors

def fecha_mayor(columnNameDate):
    totalErrDate = 0
    fecha_actual = datetime.now().date()
    
    if columnNameDate not in df.columns: 
        raise ValueError(f'No se encontró la columna {columnNameDate}')
    
    for index, fila in df.iterrows():
        cellDate = fila[columnNameDate]
        if pd.notna(cellDate):
            try:
                fecha_celda = pd.to_datetime(cellDate).date()
                if fecha_celda > fecha_actual:
                    set_bg_error(index, columnNameDate, bgError)
                    totalErrDate += 1
            except ValueError:
                set_bg_error(index, columnNameDate, bgError)
                totalErrDate += 1

    if totalErrDate > 0:
        print(f'Fecha incoherente: {totalErrDate}')
    return totalErrDate

##------------------------------------------------------------------------------------    
##--------------------------FUNCIONES PARA CADA BASE----------------------------------
##------------------------------------------------------------------------------------

#SESIONES COLECTIVAS
def sc():
    global df
    global sheet    

    errorCountPg1 = 0
    errorCountPg3 = 0

    for index, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]

        data = list(sheet.values)
        cols = data[1]
        
        data = data[2:]

        init_df = pd.DataFrame(data, columns=cols)
        
        clean_df = clean_dataframe(init_df)

        df = list_pages(clean_df)

        print(f"-------------------Página {index+1}-------------------")
        if index == 0:
            errorCountPg1 = sc_pg_1()
        
        if index == 1:
            errorCountPg2 = sc_pg_2()
            
        if index == 2:
            errorCountPg3 = sc_pg_3()
            print("----------------------------------------------")
            
    errorCountSc = errorCountPg1+errorCountPg2+errorCountPg3
    print(f">>TOTAL ERRORES EN SESIONES COLECTIVAS: {(errorCountSc)}") 

def sc_pg_1():    
    countErrorsPg1 = 0
    requiredFields = ['LUGAR DE LA ACTIVIDAD', 'ZONA', 'LOCALIDAD', 'UPZ/UPR', 'BARRIO', 'BARRIO PRIORIZADO',
                   'MANZANA DE CUIDADO', 'TELÉFONO']
    
    countErrorsPg1 = (required_fields(requiredFields))

    if countErrorsPg1 > 0:
        print(f"Errores en la página 1: {countErrorsPg1}")
    else:
        print("Sin errores en la página 1")

    return countErrorsPg1

def sc_pg_2():
    countErrorsPg2 = 0
    requiredFields = ['COMPONENTE','PROCESO', 'TEMA', 'FECHA', 'NOMBRE PROFESIONAL 1']
    
    countErrorsPg2 = (required_fields(requiredFields)+validate_only_text('NOMBRE PROFESIONAL 1', 'NOMBRE PROFESIONAL 2')+fecha_mayor('FECHA'))
    
    if countErrorsPg2 > 0:
        print(f"Errores en la página 2: {countErrorsPg2}")
    
    return countErrorsPg2

def sc_pg_3():
    countErrorsPg3 = 0
    # requiredFields = ['PRIMER NOMBRE', 'PRIMER APELLIDO', 'TIPO DOCUMENTO', 'NÚMERO DOCUMENTO', 'SEXO', 'GENERO',
    #                'ESTADO CIVIL', 'ETNIA', 'NACIONALIDAD', 'POBLACIÓN DIFERENCIAL Y DE INCLUSIÓN', 'OCUPACIÓN']
    requiredFields = ['Sub-Sección => Individuo']
    
    #countErrorsPg3 = (required_fields(requiredFields)+validate_only_text(requiredFields[0], 'SEGUNDO NOMBRE', requiredFields[1], 'SEGUNDO APELLIDO'))
    countErrorsPg3 = (required_fields(requiredFields))

    if countErrorsPg3 > 0:
        print(f"Errores en la página 3: {countErrorsPg3}")

    return countErrorsPg3

#CUIDARTE
def cuidarte():
    global df
    global sheet

    errorCount = 0

    for index, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]

        data = list(sheet.values)

        # Obtener los encabezados de la fila 2 (índice 1)
        cols = data[1]
        
        # Obtener los datos a partir de la fila 3 (índice 2)
        data = data[2:]

        init_df = pd.DataFrame(data, columns=cols)
        
        clean_df = clean_dataframe(init_df)

        df = list_pages(clean_df)

        print(f"-------------------Página {index+1}-------------------")
        if index == 0:
            errorCount = cuidarte_pg_1()

            print("----------------------------------------------")
            
    print(f">>TOTAL ERRORES EN CUIDARTE: {(errorCount)}")


def cuidarte_pg_1():
    countErrorsPg1 = 0
    requiredFields = ['NOMBRES Y APELLIDOS COMPLETOS', 'DIMENSIÓN EN SALUD', 'NACIONALIDAD', 'PUNTAJE1', 'PUNTAJE2', 'PUNTAJE3']
    
    countErrorsPg1 = (required_fields(requiredFields)+validate_only_text('NOMBRES Y APELLIDOS COMPLETOS'))
    
    if countErrorsPg1 > 0:
        print(f"Errores en la página 1: {countErrorsPg1}")
    
    return countErrorsPg1

#MUJERES
def mujeres():
    global df
    global sheet

    errorCount = 0

    for index, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]

        data = list(sheet.values)

        # Obtener los encabezados de la fila 2 (índice 1)
        cols = data[1]
        
        # Obtener los datos a partir de la fila 3 (índice 2)
        data = data[2:]

        init_df = pd.DataFrame(data, columns=cols)
        
        clean_df = clean_dataframe(init_df)

        df = list_pages(clean_df)

        print(f"-------------------Página {index+1}-------------------")
        if index == 0:
            errorCount = mujeres_pg_1()

            print("----------------------------------------------")
            
    print(f">>TOTAL ERRORES EN CE MUJERES: {(errorCount)}")

def mujeres_pg_1():
    countErrorsPg1 = 0
    requiredFields = ['NACIONALIDAD', 'TIPO DE DOCUMENTO', 'NUMERO DE DOCUMENTO', 'NOMBRE COMPLETO', 'MANZANA DEL CUIDADO', 'POBLACIÓN DIFERENCIAL Y DE INCLUSIÓN',
                      '1. ¿EN QUE NIVEL CONSIDERA QUE LA INFORMACIÓN PORPORCIONADA POR LOS CENTROS DE ESCUCHA MUJERESALUD LE APORTA A CONOCER LOS DERECHOS EN SALUD PLENA?']
    requiredFieldsNext = ['1. ¿EN QUÉ NIVEL RECONOCE LOS DERECHOS EN SALUD PLENA?', '2. ¿EN QUÉ NIVEL IDENTIFICA LOS DIFERENTES TIPOS DE VIOLENCIA BASADAS EN GÉNERO Y LOS CANALES DE ATENCIÓN?',
                          '1. ¿EN QUE NIVEL CONSIDERA QUE LA INFORMACIÓN PORPORCIONADA POR LOS CENTROS DE ESCUCHA MUJERESALUD LE APORTA A CONOCER LOS DERECHOS EN SALUD PLENA?',
                          '2. ¿EN QUE NIVEL LA INFORMACIÓN ADQUIRIDA LE PERMITE AFRONTAR UNA SITUACIÓN DE VIOLENCIA?']
    
    countErrorsPg1 = (required_fields(requiredFields)+required_fields(requiredFieldsNext, 1)+validate_only_text('NOMBRE COMPLETO'))
    
    if countErrorsPg1 > 0:
        print(f"Errores en la página 1: {countErrorsPg1}")
    
    return countErrorsPg1

##------------------------------------------------------------------------------------    
##-------------------------------GUARDAR ARCHIVO--------------------------------------
##------------------------------------------------------------------------------------

def saveFile(base):
    response_save = messagebox.askquestion("Guardar archivo", "¿Guardar el archivo generado?")
    if response_save == "yes":
        cadenaGuardar = "Guardando archivo..."
        print(cadenaGuardar)
        
        try:
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            file_path_modificado = file_path.replace('.xlsx', f'{base}_errores.xlsx')
            workbook.save(file_path_modificado)
            
            print("¡El archivo ha sido guardado correctamente!")
            # Preguntar al usuario si desea abrir el archivo guardado
            open_file = messagebox.askquestion("Abrir Archivo", "¿Desea abrir el archivo guardado?")
            if open_file == 'yes':
                os.startfile(file_path_modificado)# Abre el archivo guardado
                print(f"El archivo quedó guardado en la ruta {file_path_modificado}")
        except Exception as e:
            print("Error", f"No se pudo guardar el archivo: {str(e)}")
    else:
            print("Tu archivo no se guardará")