# Código de ejemplo para crear un validador

import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import re
import shutil
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import pandas as pd
import tkinter.simpledialog as simpledialog
from colorama import init, Fore, Style
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import datetime

import tkinter as tk
from tkinter import ttk
import re
import time


def mostrar_ventana_progreso(titulo, max_val):
    # Crear una nueva ventana para mostrar el progreso
    ventana = tk.Tk()
    ventana.title(titulo)

    # Crear y colocar una etiqueta para mostrar el texto de la función en ejecución
    label = tk.Label(ventana, text=titulo)
    label.pack(pady=10)

    # Crear y colocar la barra de progreso
    progress = ttk.Progressbar(ventana, orient="horizontal", length=300, mode="determinate")
    progress.pack(pady=20)

    return ventana, progress, label

def actualizar_barra_progreso(ventana, progress, valor):
    progress['value'] = valor
    ventana.update_idletasks()

#validadores educativo 
init()
# Declarar el objeto colum inicialmente para almacenar variables para las funciones
colum = {"column": set(), "row": 0}
celTexto = {"ColumText": set()}
Genero = {"Genero": set()}
etniaVal = {"etniaVal": set()}
afiliacion = {"afiliacion": set()}
CeldasVacias = {"vacias": set()}
CeldasVacias_Condicional = {"vacias": set(), "row": 0}

placas = {"placas": set()}
Tel = {"Tel": set()}
Manzana = {"Manzana": set()}
rural = {"rural": set()}

    
def setBase(base):
    loadExcel()
    chooseBase(base)
    preguntaDescarga()
    
    
# llamar o cargar archivo excel
def loadExcel():
    # Abrir el archivo Excel
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return
    
    global workbook
    workbook = openpyxl.load_workbook(file_path)
    global sheet
    sheet = workbook.active
    
    # Leer los encabezados de la primera fila
    headers = []
    for cell in sheet[1]:
        headers.append(cell.value)
    
    # Retornar los encabezados y la hoja de cálculo
    return headers, sheet
    
def chooseBase(base):
    switch = {
        "sesiones_colectivas": SesionesCoelctivas,  #ejemplo de insercion de cada base para validar 
    }
    execute_validator = switch.get(base)
    execute_validator()
    
def SesionesCoelctivas(): # funcion para determinar la cantdad de paginas a validar 
    # Páginas del archivo Excel cargado
    num_paginas = len(workbook.sheetnames)
    print(f"El archivo Excel tiene {num_paginas} páginas.")

    # Primero, validar la página 1
    if num_paginas >= 1 and workbook.sheetnames[0] in workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]  # Acceder a la página 1
        print("Validando la página 1...")
        validar_pagina1_sesiones(sheet)

#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#/////////////////////////////////funciones////////////////////////////////////////////////////////////////
#///////////////////////////////////////////////////////////////////////////////////////////////////////////////////  
def validar_pagina1_sesiones(sheet):
    # Mostrar ventana de progreso
    
    regex = re.compile("^[a-zA-ZÑñáéíóúÁÉÍÓÚ\s]+$")
    patternTel = re.compile(r'^\d{7}(\d{3})?$')
    
    try:
        remplazarComillas(sheet) #ejecuta funcion para quitar comillas 
        ultima_fila = sheet.max_row
        celdas_pintadas_rojo = 0

        ventana, progress, label = mostrar_ventana_progreso(f"Validando pagina1_sesiones...{ultima_fila}", ultima_fila - 1)
        ventana.update()  # Refrescar la ventana principal
        
        for i in range(2, ultima_fila + 1):
            
            # Tipo institución
            if len(sheet.cell(i, 8).value.strip()) > 0 and len(sheet.cell(i, 9).value.strip()) > 0: # condicion para validar las celdas 
                celdas_pintadas_rojo += 1
                colum["column"] = {8, 9, 2}
                colum["row"] = i
                pintar(colum, sheet)# función para pintar las celdas establecidas 
                
            #////////////////////////////// Codigo para actualizar progreso de validacion NO QUITAR  ////////////////////////////////////////// 
            actualizar_barra_progreso(ventana, progress, i * 100 / ultima_fila)
            if progress['value'] == 100:
                ventana.after(100, lambda: ventana.destroy())  # Cerrar la ventana después de 100 ms
                break
            #/////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
            
        #//////////////////////////////// Actualizacion de vntana NO QUITAR ///////////////////////////////////////////////////////////
        label.config(text=f"Total errores encontrados: {celdas_pintadas_rojo}. de {ultima_fila}")
        ventana.update()  
        print(f"Total errores encontrados {celdas_pintadas_rojo}. de {ultima_fila}")
        #////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////

    except Exception as e:
        print("Error", f"Se produjo un error: {str(e)}")   
        
#////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
#///////////////////////////////// FIN funciones////////////////////////////////////////////////////////////////
#/////////////////////////////////////////////////////////////////////////////////////////////////////////////////// 


# funcio para remplazar comillas
def remplazarComillas(sheet):
    try :
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and '`' in cell.value:
                    # Elimina las comillas
                    cell.value = cell.value.replace('`', '')

                    # Verifica si el valor es un número
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = numbers.FORMAT_NUMBER
                    # Verifica si el valor es una fecha
                    elif isinstance(cell.value, datetime.date):
                        cell.number_format = numbers.FORMAT_DATE_XLSX15
                    # Verifica si el valor es texto
                    else:
                        cell.number_format = numbers.FORMAT_TEXT      
    except Exception as e:
        print("Error", f"Se produjo un error de comillas: {str(e)}")      
         
#funcion para pintar celdas 
def pintar(colum, sheet):
    colorRed = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    number_colum = len(colum["column"])
    columns = list(colum["column"])
    for i in range(number_colum):
        sheet.cell(row=colum["row"], column=columns[i]).fill = colorRed
  
def saveFile():
    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        file_path_modificado = file_path.replace('.xlsx', '_errores.xlsx')
        
        # Guardar el libro de trabajo original con los cambios realizados
        #workbook.save(file_path)
        # Guardar el archivo modificado con el nombre específico para errores
        workbook.save(file_path_modificado)
        print("Archivo guardado", "El archivo ha sido guardado correctamente.")
        # Preguntar al usuario si desea abrir el archivo guardado
        open_file = messagebox.askquestion("Abrir Archivo", "¿Desea abrir el archivo guardado?")
        if open_file == 'yes':
            os.startfile(file_path_modificado)  # Abre el archivo guardado
    except Exception as e:
        print("Error", f"No se pudo guardar el archivo: {str(e)}")

def preguntaDescarga():
    try:
        respuesta = messagebox.askquestion("Abrir Archivo", "¿Guardar el archivo generado?")
        if respuesta == "yes":
            cadenaGuardar = "Guardando archivo..."
            print(cadenaGuardar)
            saveFile()
        else:
            print("Tu archivo no será descargado")
    except Exception as e:
        print(f"No se pudo guardar el archivo: {str(e)}")
