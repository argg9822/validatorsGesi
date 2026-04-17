import threading
import time
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook

# Selenium e integraciones
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import customtkinter as ctk
from tkinter import filedialog, messagebox

# --- Configuración de Estilo ---
COLORS = {
    "bg_dark":      "#0D1117",
    "bg_card":      "#161B22",
    "bg_input":     "#010409",
    "accent":       "#238636",
    "accent_hover": "#2EA043",
    "text_main":    "#E6EDF3",
    "text_dim":     "#8B949E",
    "border":       "#30363D",
    "blue_btn":     "#1F6FEB",
    "red_btn":      "#DA3633"
}

class GesiApp(ctk.CTkToplevel):
    def __init__(self, master=None):
        super().__init__(master)

        self.title("GESI Pro - Automatización de Fichas")
        self.geometry("1100x850")
        self.configure(fg_color=COLORS["bg_dark"])
        
        self.after(100, self.lift)
        self.focus_force()
        
        # Variables de estado
        self.nombres = None
        self.driver = None
        self.captcha_listo = threading.Event()
        
        # Variables de métricas
        self.start_time = None
        self.total_registros = 0
        self.registros_procesados = 0

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._build_sidebar()
        self._build_main_panel()

    def _build_sidebar(self):
        self.sidebar = ctk.CTkFrame(self, width=300, fg_color=COLORS["bg_card"], corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        
        ctk.CTkLabel(self.sidebar, text="🔑 ACCESO GESI", 
                     font=ctk.CTkFont("Segoe UI", 20, "bold")).pack(pady=(30, 20))

        self.user_var = ctk.StringVar()
        self.pass_var = ctk.StringVar()
        
        ctk.CTkLabel(self.sidebar, text="Usuario", text_color=COLORS["text_dim"]).pack(anchor="w", padx=25)
        self.entry_user = ctk.CTkEntry(self.sidebar, textvariable=self.user_var, height=35, fg_color=COLORS["bg_input"])
        self.entry_user.pack(fill="x", padx=25, pady=(0, 10))

        ctk.CTkLabel(self.sidebar, text="Contraseña", text_color=COLORS["text_dim"]).pack(anchor="w", padx=25)
        self.entry_pass = ctk.CTkEntry(self.sidebar, textvariable=self.pass_var, show="*", height=35, fg_color=COLORS["bg_input"])
        self.entry_pass.pack(fill="x", padx=25, pady=(0, 20))

        self.btn_captcha = ctk.CTkButton(self.sidebar, text="Confirmar Captcha ✅", 
                                        fg_color=COLORS["blue_btn"], state="disabled",
                                        command=lambda: self.captcha_listo.set())
        self.btn_captcha.pack(fill="x", padx=25, pady=15)

    def _build_main_panel(self):
        self.main_view = ctk.CTkFrame(self, fg_color="transparent")
        self.main_view.grid(row=0, column=1, sticky="nsew", padx=25, pady=25)

        # --- Dashboard de Progreso ---
        self.dash_frame = ctk.CTkFrame(self.main_view, fg_color=COLORS["bg_card"], corner_radius=15)
        self.dash_frame.pack(fill="x", pady=(0, 20), ipady=10)

        # Labels de métricas
        self.metrics_container = ctk.CTkFrame(self.dash_frame, fg_color="transparent")
        self.metrics_container.pack(fill="x", padx=20, pady=10)
        
        self.lbl_progreso = ctk.CTkLabel(self.metrics_container, text="Progreso: 0%", font=("Segoe UI", 14, "bold"))
        self.lbl_progreso.grid(row=0, column=0, padx=20)
        
        self.lbl_restante = ctk.CTkLabel(self.metrics_container, text="Faltan: --:--:--", text_color="#ffcc00")
        self.lbl_restante.grid(row=0, column=1, padx=20)

        self.lbl_contador = ctk.CTkLabel(self.metrics_container, text="Procesados: 0 / 0")
        self.lbl_contador.grid(row=0, column=2, padx=20)

        # Barra de progreso
        self.progress_bar = ctk.CTkProgressBar(self.dash_frame, orientation="horizontal", progress_color=COLORS["accent"])
        self.progress_bar.set(0)
        self.progress_bar.pack(fill="x", padx=30, pady=(10, 20))

        # --- Log y Botones ---
        self.btn_file = ctk.CTkButton(self.main_view, text="📂 SELECCIONAR EXCEL", 
                                     fg_color=COLORS["border"], height=40, 
                                     command=self.seleccionar_archivo)
        self.btn_file.pack(fill="x", pady=5)

        self.txt_log = ctk.CTkTextbox(self.main_view, fg_color="#010409", border_width=1,
                                     font=ctk.CTkFont("Consolas", 12), text_color="#76e1fe")
        self.txt_log.pack(fill="both", expand=True, pady=10)
        self.txt_log.configure(state="disabled")

        self.btn_action = ctk.CTkButton(self.main_view, text="🚀 INICIAR PROCESO", 
                                       fg_color=COLORS["accent"], height=60,
                                       font=ctk.CTkFont("Segoe UI", 18, "bold"),
                                       command=self.start_thread)
        self.btn_action.pack(fill="x")

    def update_dashboard(self, procesados, total, tiempo_restante):
        porcentaje = procesados / total if total > 0 else 0
        self.progress_bar.set(porcentaje)
        self.lbl_progreso.configure(text=f"Progreso: {int(porcentaje*100)}%")
        self.lbl_contador.configure(text=f"Procesados: {procesados} / {total}")
        self.lbl_restante.configure(text=f"Faltan: {tiempo_restante}")

    def log(self, message):
        t = datetime.now().strftime("%H:%M:%S")
        self.txt_log.configure(state="normal")
        self.txt_log.insert("end", f"[{t}] {message}\n")
        self.txt_log.see("end")
        self.txt_log.configure(state="disabled")
        self.update_idletasks()

    def seleccionar_archivo(self):
        archivo = filedialog.askopenfilename(title="Seleccionar archivo", filetypes=[("Excel", "*.xlsx;*.xls")])
        if archivo:
            try:
                wb = load_workbook(archivo)
                self.nombres = wb.active 
                self.total_registros = sum(1 for row in self.nombres.iter_rows(min_row=2, max_col=1) if row[0].value is not None)
                self.log(f"✅ Excel cargado: {self.total_registros} registros encontrados.")
                self.lbl_contador.configure(text=f"Esperando inicio: 0 / {self.total_registros}")
                self.btn_file.configure(text=f"📦 {Path(archivo).name}")
            except Exception as e:
                self.log(f"❌ Error Excel: {e}")

    def wait_for_element(self, by, value, timeout=15):
        return WebDriverWait(self.driver, timeout).until(EC.presence_of_element_located((by, value)))

    def start_thread(self):
        if not self.nombres: return messagebox.showwarning("Atención", "Carga el Excel primero.")
        if not self.user_var.get(): return messagebox.showwarning("Atención", "Faltan credenciales.")
        threading.Thread(target=self.hc_crear, daemon=True).start()

    def hc_crear(self):
        self.btn_action.configure(state="disabled", text="EJECUTANDO...")
        try:
            options = webdriver.ChromeOptions()
            options.add_argument('--start-maximized')
            self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
            self.driver.get("https://gesiapps.saludcapital.gov.co/GESI_sistemas/login")
            
            self.log("🔑 Enviando credenciales...")
            self.wait_for_element(By.ID, 'usuario').send_keys(self.user_var.get())
            self.wait_for_element(By.ID, 'password').send_keys(self.pass_var.get())
            
            codigo = self.wait_for_element(By.ID, 'tokenAcceso').get_attribute('value')
            self.wait_for_element(By.ID, 'valorCodigoSeguridad').send_keys(codigo)
            
            self.log("⌛ Esperando resolución de CAPTCHA...")
            self.btn_captcha.configure(state="normal")
            self.captcha_listo.wait()
            self.captcha_listo.clear()
            self.btn_captcha.configure(state="disabled")
            
            self.next_step()
            
        except Exception as e:
            self.log(f"🛑 ERROR: {e}")
        finally:
            self.btn_action.configure(state="normal", text="🚀 REINICIAR AUTOMATIZACIÓN")

    def next_step(self):
        try:
            self.log("🚀 Iniciando navegación al módulo...")
            self.wait_for_element(By.XPATH, '/html/body/section/div/div/form/div/div/div[7]/div/div/button').click() 
            time.sleep(3)
            self.wait_for_element(By.XPATH, '/html/body/div/div/nav/div/div[4]/ul/li[6]').click() # Entorno GESI
            time.sleep(1)
            self.wait_for_element(By.XPATH, '/html/body/div/div/nav/div/div[4]/ul/li[6]/div/ul/form[15]/li/a').click() #base de entrono
            time.sleep(2)

            self.start_time = time.time()
            self.registros_procesados = 0

            for row in self.nombres.iter_rows(min_row=2, max_col=1, values_only=True):
                dato_excel = row[0]
                if dato_excel is None: continue 

                self.registros_procesados += 1
                
                # 1. Búsqueda
                input_busqueda = self.wait_for_element(By.ID, 'valorFiltro')
                input_busqueda.clear()
                input_busqueda.send_keys(str(dato_excel))
                self.wait_for_element(By.ID, 'BtnSearchRegs').click()
                time.sleep(2)
                
                # Entrar detalle
                self.wait_for_element(By.XPATH, '/html/body/div/div/main/div/div/div/div[2]/div[2]/div[1]/table/tbody/tr/td[1]/form/div/button/a/i').click()
                time.sleep(2)

                # 2. Tab y Actualizar
                self.wait_for_element(By.ID, 'valorControl16197').send_keys(Keys.TAB)
                self.wait_for_element(By.ID, 'botonActualizarInformacion').click()
                time.sleep(1)
                
                # Modal confirmar
                try:
                    self.wait_for_element(By.XPATH, '/html/body/div[2]/div/div[3]/button[1]', timeout=3).click()
                except:
                    pass
                time.sleep(1)

                # 3. Regresar
                self.wait_for_element(By.XPATH, '/html/body/div/div/main/div/div/div/div[1]/div/div[2]/table/tbody/tr/td/form/button').click()
                
                # --- MÉTRICAS ACTUALIZADAS ---
                ahora = time.time()
                transcurrido = ahora - self.start_time
                # Usamos un promedio basado en lo procesado
                promedio = transcurrido / self.registros_procesados
                restantes = self.total_registros - self.registros_procesados
                est_segundos = promedio * restantes
                
                t_restante_str = str(timedelta(seconds=int(est_segundos)))
                
                # Actualizar Dashboard UI
                self.update_dashboard(self.registros_procesados, self.total_registros, t_restante_str)
                self.log(f"Ficha {dato_excel} corregida.")
                time.sleep(1)

            self.log("🏁 PROCESO FINALIZADO CON ÉXITO")
            messagebox.showinfo("Completado", "Todos los registros han sido procesados.")

        except Exception as e:
            self.log(f"❌ Error en ciclo: {e}")

def main():
    root = ctk.CTk()
    root.withdraw() 
    GesiApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()