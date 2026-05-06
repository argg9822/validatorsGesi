import threading
import time
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Selenium e integraciones
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

import customtkinter as ctk
from tkinter import filedialog, messagebox

# ─────────────────────────────────────────────
#  PALETA  —  Negro & Rojo Élite
# ─────────────────────────────────────────────
COLORS = {
    "bg_void":      "#080808",   # negro puro de fondo
    "bg_panel":     "#0F0F0F",   # panel lateral
    "bg_card":      "#141414",   # tarjetas / cajas
    "bg_input":     "#0A0A0A",   # fondo de inputs
    "border_dim":   "#1E1E1E",   # bordes sutiles
    "border_glow":  "#8B0000",   # borde rojo oscuro
    "accent":       "#C0392B",   # rojo principal
    "accent_light": "#E74C3C",   # rojo brillante (hover)
    "accent_dim":   "#6B1A14",   # rojo muy oscuro (estado deshabilitado)
    "gold":         "#BF9B30",   # dorado para acentos premium
    "text_main":    "#F0F0F0",   # blanco principal
    "text_dim":     "#555555",   # gris texto secundario
    "text_mid":     "#888888",   # gris intermedio
    "console_txt":  "#FF6B6B",   # rojo consola
    "blue_btn":     "#1A3A5C",   # azul oscuro para captcha
    "blue_hover":   "#1F6FEB",   # azul hover
}

FONT_TITLE  = ctk.CTkFont("Georgia", 20, "bold")
FONT_LABEL  = ctk.CTkFont("Consolas", 11)
FONT_SMALL  = ctk.CTkFont("Consolas", 10)
FONT_BTN    = ctk.CTkFont("Georgia", 13, "bold")
FONT_BTN_LG = ctk.CTkFont("Georgia", 16, "bold")
FONT_LOG    = ctk.CTkFont("Consolas", 11)
FONT_MONO   = ctk.CTkFont("Courier New", 10)


def _make_separator(parent, color=None, height=1):
    color = color or COLORS["border_glow"]
    f = ctk.CTkFrame(parent, height=height, fg_color=color, corner_radius=0)
    f.pack(fill="x", padx=0, pady=0)
    return f


class GesiApp(ctk.CTkToplevel):
    def __init__(self, master=None):
        super().__init__(master)

        self.title("GESI — Sistema de Automatización")
        self.geometry("1080x760")
        self.minsize(900, 650)
        self.configure(fg_color=COLORS["bg_void"])

        self.after(100, self.lift)
        self.focus_force()

        # ── Estado interno (lógica original intacta) ──
        self.nombres          = None
        self.driver           = None
        self.captcha_listo    = threading.Event()
        self.confirmacion_si  = threading.Event()

        # ── Layout principal: sidebar | main ──
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._build_sidebar()
        self._build_main_panel()

    # ─────────────────────────────────────────
    #  SIDEBAR
    # ─────────────────────────────────────────
    def _build_sidebar(self):
        self.sidebar = ctk.CTkFrame(
            self, width=300,
            fg_color=COLORS["bg_panel"],
            corner_radius=0
        )
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_propagate(False)

        # ── Encabezado con barra de acento ──
        header = ctk.CTkFrame(self.sidebar, fg_color=COLORS["bg_card"], corner_radius=0)
        header.pack(fill="x")

        accent_bar = ctk.CTkFrame(header, height=3, fg_color=COLORS["accent"], corner_radius=0)
        accent_bar.pack(fill="x")

        ctk.CTkLabel(
            header,
            text="GESI",
            font=ctk.CTkFont("Georgia", 32, "bold"),
            text_color=COLORS["accent"]
        ).pack(pady=(20, 0))

        ctk.CTkLabel(
            header,
            text="AUTOMATIZACIÓN",
            font=ctk.CTkFont("Consolas", 9, "bold"),
            text_color=COLORS["gold"],
            letter_spacing=4
        ).pack(pady=(0, 5))

        ctk.CTkLabel(
            header,
            text="Control de Historias Clínicas",
            font=ctk.CTkFont("Georgia", 10),
            text_color=COLORS["text_dim"]
        ).pack(pady=(0, 20))

        _make_separator(header, COLORS["accent"])

        # ── Cuerpo del sidebar ──
        body = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=22, pady=20)

        # Sección Credenciales
        self._section_label(body, "CREDENCIALES DE ACCESO")

        self.user_var = ctk.StringVar()
        self.pass_var = ctk.StringVar()

        self._field_label(body, "Usuario GESI")
        self.entry_user = self._make_entry(body, self.user_var, placeholder="usuario@dominio")

        self._field_label(body, "Contraseña")
        self.entry_pass = self._make_entry(body, self.pass_var, show="●", placeholder="••••••••")

        # Divider
        ctk.CTkFrame(body, height=1, fg_color=COLORS["border_dim"]).pack(fill="x", pady=20)

        # Sección Acciones
        self._section_label(body, "ACCIONES DE PROCESO")

        self.btn_captcha = self._make_button(
            body,
            text="✦  Confirmar Captcha",
            fg=COLORS["blue_btn"],
            hover=COLORS["blue_hover"],
            state="disabled",
            cmd=lambda: self.captcha_listo.set()
        )

        ctk.CTkFrame(body, height=8, fg_color="transparent").pack()

        self.btn_proceso_si = self._make_button(
            body,
            text="✦  Incluir Digitado (SÍ)",
            fg=COLORS["accent_dim"],
            hover=COLORS["accent"],
            state="disabled",
            cmd=lambda: self.confirmacion_si.set()
        )

        # ── Footer ──
        footer = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        footer.pack(side="bottom", fill="x", padx=22, pady=14)

        _make_separator(footer, COLORS["border_dim"])

        ctk.CTkLabel(
            footer,
            text="SECRETARÍA DISTRITAL DE SALUD",
            font=ctk.CTkFont("Consolas", 8),
            text_color=COLORS["text_dim"]
        ).pack(pady=(8, 0))

        ctk.CTkLabel(
            footer,
            text="Sistema GESI  ·  v2.0",
            font=ctk.CTkFont("Consolas", 8),
            text_color=COLORS["border_glow"]
        ).pack()

    # ─────────────────────────────────────────
    #  PANEL PRINCIPAL
    # ─────────────────────────────────────────
    def _build_main_panel(self):
        self.main_view = ctk.CTkFrame(self, fg_color="transparent")
        self.main_view.grid(row=0, column=1, sticky="nsew")
        self.main_view.grid_rowconfigure(1, weight=1)
        self.main_view.grid_columnconfigure(0, weight=1)

        # ── Barra superior ──
        topbar = ctk.CTkFrame(
            self.main_view,
            fg_color=COLORS["bg_card"],
            corner_radius=0
        )
        topbar.grid(row=0, column=0, sticky="ew")

        accent_top = ctk.CTkFrame(topbar, height=2, fg_color=COLORS["accent"], corner_radius=0)
        accent_top.pack(fill="x")

        topbar_inner = ctk.CTkFrame(topbar, fg_color="transparent")
        topbar_inner.pack(fill="x", padx=28, pady=12)

        ctk.CTkLabel(
            topbar_inner,
            text="CONSOLA DE OPERACIONES",
            font=ctk.CTkFont("Georgia", 15, "bold"),
            text_color=COLORS["text_main"]
        ).pack(side="left")

        # Indicador de estado
        self.status_indicator = ctk.CTkLabel(
            topbar_inner,
            text="● EN ESPERA",
            font=ctk.CTkFont("Consolas", 10, "bold"),
            text_color=COLORS["text_dim"]
        )
        self.status_indicator.pack(side="right")

        # ── Área central ──
        center = ctk.CTkFrame(self.main_view, fg_color="transparent")
        center.grid(row=1, column=0, sticky="nsew", padx=24, pady=(18, 0))
        center.grid_rowconfigure(1, weight=1)
        center.grid_columnconfigure(0, weight=1)

        # Botón seleccionar archivo
        self.btn_file = ctk.CTkButton(
            center,
            text="◈  Seleccionar Archivo Excel (.xlsx)",
            fg_color=COLORS["bg_card"],
            hover_color=COLORS["border_glow"],
            text_color=COLORS["text_mid"],
            border_width=1,
            border_color=COLORS["border_dim"],
            height=42,
            corner_radius=4,
            font=ctk.CTkFont("Consolas", 12),
            command=self.seleccionar_archivo
        )
        self.btn_file.grid(row=0, column=0, sticky="ew", pady=(0, 12))

        # ── Log / Consola ──
        log_frame = ctk.CTkFrame(
            center,
            fg_color=COLORS["bg_card"],
            corner_radius=6,
            border_width=1,
            border_color=COLORS["border_glow"]
        )
        log_frame.grid(row=1, column=0, sticky="nsew")
        log_frame.grid_rowconfigure(1, weight=1)
        log_frame.grid_columnconfigure(0, weight=1)

        log_header = ctk.CTkFrame(log_frame, fg_color=COLORS["bg_void"], corner_radius=0)
        log_header.grid(row=0, column=0, sticky="ew")

        ctk.CTkFrame(log_header, height=1, fg_color=COLORS["border_glow"], corner_radius=0).pack(fill="x")

        log_title_row = ctk.CTkFrame(log_header, fg_color="transparent")
        log_title_row.pack(fill="x", padx=16, pady=8)

        ctk.CTkLabel(
            log_title_row,
            text="◉ OUTPUT LOG",
            font=ctk.CTkFont("Consolas", 10, "bold"),
            text_color=COLORS["accent"]
        ).pack(side="left")

        ctk.CTkLabel(
            log_title_row,
            text="GESI_AUTOMATION_ENGINE",
            font=ctk.CTkFont("Consolas", 9),
            text_color=COLORS["border_dim"]
        ).pack(side="right")

        self.txt_log = ctk.CTkTextbox(
            log_frame,
            fg_color=COLORS["bg_input"],
            font=FONT_LOG,
            text_color=COLORS["console_txt"],
            corner_radius=0,
            border_width=0,
            scrollbar_button_color=COLORS["accent_dim"],
            scrollbar_button_hover_color=COLORS["accent"]
        )
        self.txt_log.grid(row=1, column=0, sticky="nsew", padx=1, pady=(0, 1))
        self.txt_log.configure(state="disabled")

        # ── Botón principal ──
        btn_frame = ctk.CTkFrame(self.main_view, fg_color="transparent")
        btn_frame.grid(row=2, column=0, sticky="ew", padx=24, pady=14)

        self.btn_action = ctk.CTkButton(
            btn_frame,
            text="⬡  INICIAR AUTOMATIZACIÓN",
            fg_color=COLORS["accent"],
            hover_color=COLORS["accent_light"],
            text_color=COLORS["text_main"],
            height=52,
            corner_radius=4,
            font=FONT_BTN_LG,
            border_width=0,
            command=self.start_thread
        )
        self.btn_action.pack(fill="x")

    # ─────────────────────────────────────────
    #  HELPERS DE UI
    # ─────────────────────────────────────────
    def _section_label(self, parent, text):
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", pady=(6, 10))

        ctk.CTkLabel(
            row,
            text=text,
            font=ctk.CTkFont("Consolas", 9, "bold"),
            text_color=COLORS["gold"]
        ).pack(side="left")

        ctk.CTkFrame(row, height=1, fg_color=COLORS["border_dim"]).pack(
            side="left", fill="x", expand=True, padx=(8, 0), pady=6
        )

    def _field_label(self, parent, text):
        ctk.CTkLabel(
            parent,
            text=text,
            font=ctk.CTkFont("Consolas", 10),
            text_color=COLORS["text_mid"]
        ).pack(anchor="w", pady=(4, 2))

    def _make_entry(self, parent, var, show=None, placeholder=""):
        kwargs = dict(
            textvariable=var,
            height=36,
            fg_color=COLORS["bg_input"],
            border_color=COLORS["border_dim"],
            border_width=1,
            text_color=COLORS["text_main"],
            placeholder_text=placeholder,
            placeholder_text_color=COLORS["text_dim"],
            font=ctk.CTkFont("Consolas", 12),
            corner_radius=4
        )
        if show:
            kwargs["show"] = show
        e = ctk.CTkEntry(parent, **kwargs)
        e.pack(fill="x", pady=(0, 6))
        return e

    def _make_button(self, parent, text, fg, hover, state, cmd):
        btn = ctk.CTkButton(
            parent,
            text=text,
            fg_color=fg,
            hover_color=hover,
            text_color=COLORS["text_main"],
            height=38,
            corner_radius=4,
            font=ctk.CTkFont("Consolas", 11, "bold"),
            state=state,
            command=cmd
        )
        btn.pack(fill="x")
        return btn

    # ─────────────────────────────────────────
    #  LOG HELPER
    # ─────────────────────────────────────────
    def log(self, message):
        t = datetime.now().strftime("%H:%M:%S")
        self.txt_log.configure(state="normal")
        self.txt_log.insert("end", f"[{t}]  {message}\n")
        self.txt_log.see("end")
        self.txt_log.configure(state="disabled")
        self.update_idletasks()

    def _set_status(self, text, color):
        self.status_indicator.configure(text=text, text_color=color)

    # ─────────────────────────────────────────
    #  LÓGICA DE AUTOMATIZACIÓN (sin cambios)
    # ─────────────────────────────────────────
    def seleccionar_archivo(self):
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo de Excel",
            filetypes=[("Archivos de Excel", "*.xlsx;*.xls")]
        )
        if archivo:
            try:
                wb = load_workbook(archivo)
                self.nombres = wb['Hoja1']
                self.log(f"Archivo cargado: {Path(archivo).name}")
                self.btn_file.configure(
                    text=f"✔  {Path(archivo).name}",
                    text_color=COLORS["accent_light"],
                    border_color=COLORS["accent"]
                )
            except Exception as e:
                self.log(f"Error cargando Excel: {e}")

    def wait_for_element(self, by, value, timeout=10):
        return WebDriverWait(self.driver, timeout).until(
            EC.presence_of_element_located((by, value))
        )

    def start_thread(self):
        if not self.nombres:
            return messagebox.showwarning("Falta Info", "Carga el Excel primero.")
        threading.Thread(target=self.hc_crear, daemon=True).start()

    def hc_crear(self):
        self.btn_action.configure(state="disabled")
        self._set_status("● EJECUTANDO", COLORS["accent_light"])
        try:
            self.log("Abriendo Chrome...")
            options = webdriver.ChromeOptions()
            options.add_argument('--start-maximized')
            self.driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()),
                options=options
            )

            self.driver.get("https://gesiapps.saludcapital.gov.co/GESI_sistemas/login")
            time.sleep(2)

            try:
                self.wait_for_element(By.XPATH, '/html/body/div/div[2]/button[3]').click()
                time.sleep(1)
                self.wait_for_element(By.XPATH, '/html/body/div/div[3]/p[2]/a').click()
                time.sleep(3)
            except:
                self.log("Omitiendo bypass de seguridad inicial.")

            self.log("Enviando credenciales de usuario...")
            self.wait_for_element(By.ID, 'usuario').send_keys(self.user_var.get())
            self.wait_for_element(By.ID, 'password').send_keys(self.pass_var.get())

            codigo = self.wait_for_element(By.ID, 'tokenAcceso').get_attribute('value')
            self.wait_for_element(By.ID, 'valorCodigoSeguridad').send_keys(codigo)

            self.log("⚠️  POR FAVOR, RESUELVE EL CAPTCHA Y PRESIONA 'Confirmar Captcha'.")
            self.btn_captcha.configure(state="normal", fg_color=COLORS["blue_hover"])

            self.captcha_listo.wait()
            self.captcha_listo.clear()
            self.btn_captcha.configure(state="disabled", fg_color=COLORS["blue_btn"])

            self.next_step()

        except Exception as e:
            self.log(f"Error: {e}")
            self._set_status("● ERROR", COLORS["accent"])
            self.btn_action.configure(state="normal")

    def next_step(self):
        self.log("Navegando al módulo de Historias Clínicas...")
        self.wait_for_element(By.XPATH, '/html/body/section/div/div/form/div/div/div[7]/div/div/button').click()
        time.sleep(4)
        self.wait_for_element(By.XPATH, '/html/body/div/div/nav/div/div[4]/ul/li[7]').click()
        time.sleep(1)
        self.wait_for_element(By.XPATH, '/html/body/div/div/nav/div/div[4]/ul/li[7]/div/ul/li[1]/a').click()
        time.sleep(2)
        self.wait_for_element(By.XPATH, '/html/body/div/div/main/div/div/div/div[1]/div/div/table/tbody/tr/td[7]/input').click()
        time.sleep(1)
        self.wait_for_element(By.XPATH, '/html/body/div[2]/div/div[3]/button[1]').click()

        self.log("Esperando confirmación de inicio de proceso...")
        self.btn_proceso_si.configure(state="normal", fg_color=COLORS["accent"])

        self.confirmacion_si.wait()
        self.confirmacion_si.clear()
        self.btn_proceso_si.configure(state="disabled", fg_color=COLORS["accent_dim"])

        self.si_proceso()

    def si_proceso(self):
        total_filas = self.nombres.max_row
        try:
            for i in range(1, total_filas + 1):
                fila = self.nombres[f'A{i}:F{i}'][0]
                ficha_val = fila[0].value
                fecha_val = fila[1].value
                formato   = fecha_val.strftime('%d/%m/%Y')

                datos = [ficha_val, formato, fila[2].value, fila[3].value, fila[4].value, fila[5].value]
                self.log(f"Ingresando ficha: {ficha_val}")

                self.DatosCrearSi(datos)

                self.wait_for_element(By.XPATH, '/html/body/div[2]/div/div[3]/button[1]').click()
                self.wait_for_element(By.XPATH, '/html/body/div/div/main/div/div/div/div[1]/div/div/table/tbody/tr/td[5]/input').click()
                self.wait_for_element(By.XPATH, '/html/body/div[2]/div/div[3]/button[1]').click()

            self.log("✅  PROCESO COMPLETADO CORRECTAMENTE.")
            self._set_status("● COMPLETADO", COLORS["gold"])
            messagebox.showinfo("Éxito", "Proceso completado.")
        except Exception as e:
            self.log(f"Error en bucle: {e}")
            self._set_status("● ERROR", COLORS["accent"])
        finally:
            self.btn_action.configure(state="normal")

    def DatosCrearSi(self, Datos):
        Select(self.wait_for_element(By.ID, 'Digitado')).select_by_visible_text('Si')
        fecha_fields = ['Fecha_entrega_tecnologo', 'Fecha_actualizacion', 'Fecha_entrega_digitacion']
        for field in fecha_fields:
            self.wait_for_element(By.ID, field).send_keys(Datos[1])
        self.wait_for_element(By.ID, 'Nro_actualizacion').send_keys('1')

        self.llenar(Datos)
        self.wait_for_element(By.XPATH, '/html/body/div/div/main/div/div/div/div[2]/form/div[12]/div/center/input').click()

    def llenar(self, Datos):
        element_ficha = self.wait_for_element(By.ID, 'Ficha_fic')
        element_ficha.clear()
        element_ficha.send_keys(Datos[0])

        profesional = self.wait_for_element(By.ID, 'Nombre_profesional')
        profesional.clear()
        profesional.send_keys(Datos[2])

        fecha_fields = ['Fecha_ingreso', 'Fecha_entrega_profesional']
        for field in fecha_fields:
            element = self.wait_for_element(By.ID, field)
            element.clear()
            element.send_keys(Datos[1])

        Select(self.wait_for_element(By.ID, 'Id_perfil')).select_by_visible_text(Datos[5])

        for attempt in range(10):
            try:
                espacio = Select(self.wait_for_element(By.ID, 'Espacio_fic'))
                espacio.select_by_visible_text('1 -Hogar')
                time.sleep(1)
                espacio.select_by_visible_text(Datos[3])
                time.sleep(1)
                break
            except:
                time.sleep(1)

        Select(self.wait_for_element(By.ID, 'Id_Base')).select_by_visible_text(Datos[4])


# ─────────────────────────────────────────────
#  ARRANQUE
# ─────────────────────────────────────────────
def main(master=None):
    if master is None:
        root = ctk.CTk()
        app = GesiApp(root)
        root.mainloop()
    else:
        app = GesiApp(master)
        return app


if __name__ == "__main__":
    main()