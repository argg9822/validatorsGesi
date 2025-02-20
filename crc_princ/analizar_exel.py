import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from tkinter import filedialog, messagebox
from datetime import datetime
from fuzzywuzzy import fuzz
import random

def analizar_excel_2(validador):
    print("Realizando validación de datos por favor espere...")
    archivo_excel = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx *.xls")]
        )
    if archivo_excel:
        try:
            # Leer el archivo Excel
            df = pd.read_excel(archivo_excel)

            # Cargar el archivo Excel en openpyxl para aplicar formato
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active

            # Color de fondo rojo para las celdas que no cumplen con la condición
            rojo_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for regla in validador["reglas"]:
                columna = regla.get("columna")
            
                
                tipo = regla.get("tipo")

                if columna in df.columns:
                    col_idx = df.columns.get_loc(columna) + 1  # Obtener el índice de la columna en openpyxl (1-based)
                    

                    if tipo == "longitud":
                        df_original = df.copy()
                        max_longitud = int(regla["condicion"].split("<= ")[1])
                        violaciones = df[columna][df[columna].astype(str).str.len() > max_longitud]
                        for idx in violaciones.index:
                            print("Errores longitud")
                            ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill  # +2 por el encabezado
                            ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                        df = df_original.copy()
                        ws.auto_filter.ref = None
                            

                    elif tipo == "numerico":
                        try:
                            df_original = df.copy()
                            operador, valor = regla["condicion"].split(" ")
                            valor = int(valor)
                            # Convertir la columna a numérico, forzando errores a NaN
                            df[columna] = pd.to_numeric(df[columna], errors='coerce')
                            
                            if operador == "mayor":
                                violaciones = df[columna][df[columna] > valor]
                            elif operador == "menor":
                                violaciones = df[columna][df[columna] < valor]

                            for idx in violaciones.index:
                                ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill  # Marcar en rojo
                                ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                                
                            df = df_original.copy()
                            ws.auto_filter.ref = None
                                

                        except ValueError:
                            pass

                    elif tipo == "patron":
                        print("Validando Patrones")
                        
                        patron = regla["patron"]
                        df_original = df.copy()
                        # Normalizar los datos
                        df[columna] = df[columna].astype(str).str.strip()
                        
                        # Filtrar las filas que no cumplen con el patrón
                        violaciones = df[columna][df[columna].str.fullmatch(patron) == False]
                        
                        for idx in violaciones.index:
                            ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill  # Marcar en rojo
                            ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                        df = df_original.copy()
                        ws.auto_filter.ref = None
                            

                    elif tipo == "unico":
                        print("Validando valores Unicos")
                        
                        df_original = df.copy()
                        duplicados = df[columna][df[columna].duplicated()]
                        for idx in duplicados.index:
                            ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill
                            ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                        df = df_original.copy()
                        ws.auto_filter.ref = None
                        
                            # Marcar en rojo

                    elif tipo == "dependiente positivo":
                        print("Validando valores Dependientes")
                        columna_dependiente = regla.get("columna_dependiente")
                        valor_dependiente = regla.get("valor_dependiente")
                        valor_esperado = regla.get("valor_esperado")
                        columna_dependiente1 = regla.get("columna_dependiente")
                        idx_dependiente1 = df.columns.get_loc(columna_dependiente1) + 1
                        df_original = df.copy()
                        if columna_dependiente in df.columns:
                            # Filtrar las filas donde la columna dependiente tenga el valor esperado
                            filas_dependientes = df[df[columna_dependiente] == valor_dependiente]

                            # Filtrar las filas que NO cumplen con el valor esperado en la columna principal
                            violaciones = filas_dependientes[filas_dependientes[columna] != valor_esperado]

                            # Solo marcar en rojo las filas que no cumplen con la condición
                            for idx in violaciones.index:
                               
                                ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                                ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill  
                                ws.cell(row=idx + 2, column=idx_dependiente1).fill = rojo_fill
                                
                            df = df_original.copy()
                            ws.auto_filter.ref = None
                                
   
     
                        else:
                            messagebox.showinfo("Advertencia", f"Columna dependiente '{columna_dependiente}' no encontrada en el archivo Excel.")
                            
                    elif tipo == "no_vacio":
                        columnas = regla.get("columnas")
                        df_original = df.copy()
                        # Asegúrate de que 'columna' sea una lista
                        if isinstance(columnas, str):  # Si 'columna' es una cadena en lugar de lista
                            columnas = [columnas]  # Convertirla en una lista
                        
                        for col_idx, columna in enumerate(columnas, start=1):  # Enumerar las columnas con índice
                            if not columna.strip():  # Verifica si está vacía o contiene solo espacios
                                ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill
                                ws.cell(row=idx + 2, column=2).fill = rojo_fill
                        df = df_original.copy()
                        ws.auto_filter.ref = None
                        
                                

                    elif tipo == "dependiente_error":
                    
                        columna_dependiente = regla.get("columna_dependiente")
                        valor_dependiente = regla.get("valor_dependiente")
                        valor_esperado = regla.get("valor_esperado")
                        columna_dependiente1 = regla.get("columna_dependiente")
                        idx_dependiente1 = df.columns.get_loc(columna_dependiente1) + 1
                        df_original = df.copy()
                        if columna_dependiente in df.columns:
                            # Filtrar las filas donde la columna dependiente tenga el valor esperado
                            filas_dependientes = df[df[columna_dependiente] == valor_dependiente]

                            # Filtrar las filas que NO cumplen con el valor esperado en la columna principal
                            violaciones = filas_dependientes[filas_dependientes[columna] == valor_esperado]

                            # Solo marcar en rojo las filas que no cumplen con la condición
                            for idx in violaciones.index:
                                # Marcar en rojo las celdas que no cumplen la condición (solo las filas con violaciones)
                                ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                                ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill
                                ws.cell(row=idx + 2, column=idx_dependiente1).fill = rojo_fill
                            df = df_original.copy()
                            ws.auto_filter.ref = None
                                
                            
                                
                        else:
                            messagebox.showinfo("Advertencia", f"Columna dependiente '{columna_dependiente}' no encontrada en el archivo Excel.")
                        
                    elif tipo == "dependiente longitud":
                    
                        columna_dependiente = regla.get("columna_dependiente")
                        valor_dependiente = regla.get("valor_dependiente")
                        valor_esperado = regla.get("valor_esperado")
                        columna_dependiente1 = regla.get("columna_dependiente")
                        idx_dependiente1 = df.columns.get_loc(columna_dependiente1) + 1
                        df_original = df.copy()
                        if columna_dependiente in df.columns:
                            # Filtrar las filas donde la columna dependiente tenga el valor esperado
                            filas_dependientes = df[df[columna_dependiente] == valor_dependiente]
                            
                            max_longitud = int(regla["valor_esperado"].split("<= ")[1])
                            
                            violaciones = filas_dependientes[filas_dependientes[columna] .astype(str).str.len() > max_longitud]
                            
                            for idx in violaciones.index:
                                # Marcar en rojo las celdas que violan la regla de longitud
                                ws.cell(row=idx + 2, column=col_idx).fill = rojo_fill  # +2 por el encabezado
                                ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                                ws.cell(row=idx + 2, column=idx_dependiente1).fill = rojo_fill
                            df = df_original.copy()
                            ws.auto_filter.ref = None

                        else:
                            messagebox.showinfo("Advertencia", f"Columna dependiente '{columna_dependiente}' no encontrada en el archivo Excel.")
                            
                    elif tipo == "dependiente_edad_positivo":
                        columna_dependiente = regla.get("columna_dependiente")  # Fecha de nacimiento
                        valor_dependiente = regla.get("valor_dependiente")  # Rango o edad específica
                        valor_esperado = regla.get("valor_esperado")  # Valor esperado en la columna principal
                        fecha_intervencion = regla.get("Fecha_int")  # Columna con la fecha de referencia
                        nacionalidad = regla.get("nacionalidad") 
                        valor_esperado_nacionalidad = regla.get("nacionalidad_value")# Columna para filtrar primero por nacionalidad dato puede ser "Colombia" o 50
                        
                        df_original = df.copy()
                        # Verificar que las columnas necesarias estén en el DataFrame
                        if columna in df.columns and columna_dependiente in df.columns and fecha_intervencion in df.columns:
                            
                            
                            # Filtrar por nacionalidad si es que se ha especificado
                            if nacionalidad and nacionalidad in df.columns:
                                df[nacionalidad] = df[nacionalidad].astype(str)  # Convertir la columna a texto
                                valor_esperado_nacionalidad = str(valor_esperado_nacionalidad)  # Convertir el valor esperado a texto

                                # Filtrar los datos
                                df = df[df[nacionalidad] == valor_esperado_nacionalidad] # Filtrar por la nacionalidad deseada

                        
                            # Convertir las columnas a datetime si no lo están
                            df[columna_dependiente] = pd.to_datetime(df[columna_dependiente], errors='coerce')
                            df[fecha_intervencion] = pd.to_datetime(df[fecha_intervencion], errors='coerce')

                         
                            if columna_dependiente in df.columns and fecha_intervencion in df.columns:
                                df["edad_calculada"] = df.apply(
                                    lambda row: calcular_edad(row[columna_dependiente], row[fecha_intervencion]) 
                                    if pd.notnull(row[columna_dependiente]) and pd.notnull(row[fecha_intervencion]) else None, axis=1
                                )
                            else:
                                print("Una de las columnas no existe en el DataFrame.")
                            
                        
                            
                            # Identificar filas que no cumplen con la regla
                            if "," in valor_dependiente:  # Rango de edades (e.g., "0,13")
                                min_edad, max_edad = map(int, valor_dependiente.split(","))
                                violaciones = df[
                                    (df["edad_calculada"] >= min_edad) &
                                    (df["edad_calculada"] <= max_edad) &
                                    (df[columna] != valor_esperado)
                                ]
                            else:  # Edad específica (e.g., "14")
                                edad_especifica = int(valor_dependiente)
                                violaciones = df[
                                    (df["edad_calculada"] == edad_especifica) &
                                    (df[columna] != valor_esperado)
                                ]

                            # Marcar las celdas que no cumplen con la regla
                            for idx in violaciones.index:
                                ws.cell(row=idx + 2, column=df.columns.get_loc(columna) + 1).fill = rojo_fill
                                ws.cell(row=idx + 2, column=df.columns.get_loc(columna_dependiente) + 1).fill = rojo_fill
                                ws.cell(row=idx + 2, column=df.columns.get_loc(fecha_intervencion) + 1).fill = rojo_fill
                                ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                            df = df_original.copy()
                            ws.auto_filter.ref = None
                            
                        else:
                            # Mostrar mensaje de advertencia si las columnas no existen
                            print(f"Advertencia: Una de las columnas especificadas no existe en el archivo Excel.")

                                        
                    elif tipo == "dependiente edad error":
                        print("validando edades errores")
                        columna_dependiente = regla.get("columna_dependiente")  # Fecha de nacimiento
                        valor_dependiente = regla.get("valor_dependiente" ) # Rango o edad específica
                        valor_esperado = regla.get("valor_esperado")  # Valor esperado en la columna principal
                        fecha_intervencion = regla.get("Fecha_int")  # Columna con la fecha de referencia
                        df_original = df.copy()
                        # Verificar que las columnas necesarias estén en el DataFrame
                        if columna in df.columns and columna_dependiente in df.columns and fecha_intervencion in df.columns:
                            # Convertir las columnas a datetime si no lo están
                            df[columna_dependiente] = pd.to_datetime(df[columna_dependiente], errors='coerce')
                            df[fecha_intervencion] = pd.to_datetime(df[fecha_intervencion], errors='coerce')

                            # Calcular la edad usando la fecha de referencia
                            df["edad_calculada"] = df.apply(
                                lambda row: calcular_edad(row[columna_dependiente], row[fecha_intervencion]) 
                                if pd.notnull(row[columna_dependiente]) and pd.notnull(row[fecha_intervencion]) else None, axis=1
                            )

                            # Identificar filas que no cumplen con la regla
                            if "," in valor_dependiente:  # Rango de edades (e.g., "0,13")
                                min_edad, max_edad = map(int, valor_dependiente.split(","))
                                violaciones = df[
                                    (df["edad_calculada"] >= min_edad) &
                                    (df["edad_calculada"] <= max_edad) &
                                    (df[columna] == valor_esperado)
                                ]
                            else:  # Edad específica (e.g., "14")
                                edad_especifica = int(valor_dependiente)
                                violaciones = df[
                                    (df["edad_calculada"] == edad_especifica) &
                                    (df[columna] == valor_esperado)
                                ]

                            # Marcar las celdas que no cumplen con la regla
                            for idx in violaciones.index:
                                ws.cell(row=idx + 2, column=df.columns.get_loc(columna) + 1).fill = rojo_fill
                                ws.cell(row=idx + 2, column=df.columns.get_loc(columna_dependiente) + 1).fill = rojo_fill
                                ws.cell(row=idx + 2, column=df.columns.get_loc(fecha_intervencion) + 1).fill = rojo_fill
                                ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en rojo
                            
                            df = df_original.copy()
                            ws.auto_filter.ref = None
                            
                                
                        else:
                            # Mostrar mensaje de advertencia si las columnas no existen
                            print(f"Advertencia: Una de las columnas especificadas no existe en el archivo Excel.")
                            
                    elif tipo == "fechas menor/mayor que":
                        # Fecha con la que se compara
                        df_original = df.copy()
                        
                        fecha_comparar = regla.get("fecha_comparar")
                        comparacion = regla.get("comparacion")
                        # Asegurarse de que las columnas son tipo datetime
                        df[columna] = pd.to_datetime(df[columna], errors="coerce")
                        df[fecha_comparar] = pd.to_datetime(df[fecha_comparar], errors="coerce")

                        # Eliminar filas con valores vacíos o NaT en las columnas relevantes
                        df = df.dropna(subset=[columna])

                        # Aplicar la comparación según el operador
                        if comparacion == ">":
                            violaciones = df[df[columna] > df[fecha_comparar]]
                        elif comparacion == "<":
                            violaciones = df[df[columna] < df[fecha_comparar]]
                        else:
                            raise ValueError("Tipo de comparación no válido. Use '>' o '<'.")

                        # Resaltar en rojo las violaciones en el archivo Excel
                        for idx in violaciones.index:
                            # Colorear las celdas de las columnas involucradas
                            ws.cell(row=idx + 2, column=df.columns.get_loc(columna) + 1).fill = rojo_fill
                            ws.cell(row=idx + 2, column=df.columns.get_loc(fecha_comparar) + 1).fill = rojo_fill
                            ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar en columna fija (columna 2)
                            
                        df = df_original.copy()  
                        ws.auto_filter.ref = None

                    elif tipo == "dependiente_vacio":
                        # Cambiar el patrón para filtrar celdas que no están vacías y no contienen solo espacios
                        patron = r"^\S+$"  # Esta expresión regular asegura que la celda no esté vacía y no contenga solo espacios
                        columna_dependiente = regla.get("columna_dependiente")
                        valor_dependiente = regla.get("valor_dependiente")
                        df_original = df.copy()
                        
                        # Filtrar el DataFrame para las filas donde columna_dependiente coincide con valor_dependiente
                        df_filtrado = df[df[columna_dependiente] == valor_dependiente]

                        ## Filtrar las filas que no cumplen con el patrón
                        violaciones = df_filtrado[df_filtrado[columna].str.fullmatch(patron) == False]

                        # Iterar sobre las filas con violaciones
                        for idx in violaciones.index:
                            # Colorear las celdas de las columnas involucradas
                            ws.cell(row=idx + 2, column=df.columns.get_loc(columna) + 1).fill = rojo_fill  # Columna verificada
                            ws.cell(row=idx + 2, column=df.columns.get_loc(columna_dependiente) + 1).fill = rojo_fill  # Columna dependiente
                            ws.cell(row=idx + 2, column=2).fill = rojo_fill  # Marcar columna fija (columna 2)
                        
                        df = df_original.copy()  
                        ws.auto_filter.ref = None
                        
                

                    elif tipo == "coincidencia_textos":
                        print("Validando valores de textos")

                        # Extraer los valores junto con su índice
                        texto = df[[columna]].dropna().reset_index().values.tolist()  # Obtiene índice y valores

                        grupos_similares = []  # Lista para almacenar grupos de textos similares

                        # Conjunto para rastrear los índices ya agrupados
                        indices_agrupados = set()

                        for i in range(len(texto)):
                            idx_i, valor_i = texto[i]

                            # Si ya está en un grupo, lo saltamos
                            if idx_i in indices_agrupados:
                                continue  

                            grupo_actual = [(idx_i, valor_i)]  # Grupo de similitud actual

                            for j in range(i + 1, len(texto)):
                                idx_j, valor_j = texto[j]

                                # Si el índice ya está agrupado, lo saltamos
                                if idx_j in indices_agrupados:
                                    continue

                                similitud = fuzz.ratio(str(valor_i), str(valor_j))

                                if 80 <= similitud < 100:
                                    grupo_actual.append((idx_j, valor_j))
                                    indices_agrupados.add(idx_j)  # Marcar como agrupado

                            # Si encontramos coincidencias, agregamos el grupo
                            if len(grupo_actual) > 1:
                                grupos_similares.append(grupo_actual)
                                indices_agrupados.update(idx for idx, _ in grupo_actual)  # Marcar todos los índices del grupo
                                
                        print("-" * 50)
                        print("Analizando datos de textos..." )
                        print("-" * 50)

                        colores_usados = set()
                        # Imprimir los grupos encontrados
                         # Pintar las celdas en rojo
                        for grupo in grupos_similares:
                            color = generar_color_unico()
        
                            # Evitar colores repetidos
                            while color in colores_usados:
                                color = generar_color_unico()
                            
                            color_hex = generar_color_unico()
                            fill_color = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

                            print("Grupo de textos similares:")
                            for idx, valor in grupo:
                                fila_excel = df.index[idx] + 2  # Ajustar fila real en Excel
                                print(f"Fila {fila_excel}: {valor} ")
                                ws.cell(row=fila_excel , column=df.columns.get_loc(columna) + 1).fill = fill_color  # Columna verificada
                                ws.cell(row=fila_excel , column=2).fill = rojo_fill  # Marcar columna fija (columna 2)
                                
                                print(f"celta pintada {fila_excel} de {color} ")
                                
                            print("-" * 50)

     
                else: 
                    messagebox.showinfo("Advertencia", f"Columna '{columna}' no encontrada en el archivo Excel.")

            # Guardar el nuevo archivo Excel con las celdas marcadas
            nuevo_archivo = filedialog.asksaveasfilename(
                title="Guardar archivo Excel con validaciones",
                defaultextension=".xlsx",
                filetypes=[("Archivos Excel", "*.xlsx")]
            )

            if nuevo_archivo:
                wb.save(nuevo_archivo)
                print("Finalizado")
                
                messagebox.showinfo("Éxito", "Se ha creado un nuevo archivo con las validaciones marcadas en rojo.")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo analizar el archivo Excel:\n{e}")

def generar_color_unico():
    """Genera un color aleatorio en formato ARGB compatible con openpyxl."""
    color = "{:02X}{:02X}{:02X}".format(random.randint(150, 255), random.randint(10, 255), random.randint(10, 255))
    return f"FF{color}"  # Añadir "FF" para el canal de transparencia

def calcular_edad(fecha_nacimiento, fecha_referencia):
    

        edad = fecha_referencia.year - fecha_nacimiento.year
        if (fecha_referencia.month, fecha_referencia.day) < (fecha_nacimiento.month, fecha_nacimiento.day):
            edad -= 1
        print(f"verificando edades {edad}")
        return edad
